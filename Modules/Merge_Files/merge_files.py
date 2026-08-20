"""
==================================================
MERGE FILES TOOL (MASTER COLUMN & DUPES ENHANCED)
==================================================

PURPOSE
--------------------------------------------------
Merges multiple Excel files (.xlsx / .xls) uploaded simultaneously.
Automatically selects the file with the highest number of columns as the
master structure, maps all matching columns regardless of order, leaves
missing fields blank, and removes blank rows.

Preserves large numeric identifiers (e.g., 16-digit PRNs, Roll Nos)
without converting them into scientific notation (e.g., 2.01902E+15).

Provides advanced duplicate row detection (highlighting and/or removal)
using all or custom selected columns, outputting a styled 'Merged Data' sheet
and an organized 'Duplicates' sheet with thick group dividers.
==================================================
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------
# Helpers
# --------------------------------------------------
def clean_column_name(name: str) -> str:
    """Trim leading/trailing spaces and collapse internal whitespace."""
    name = str(name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def sanitize_cell_value(val):
    """
    Cleans up string representations of numbers, stripping
    erroneous '.0' artifacts while preserving large integer IDs as text.
    """
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str


def style_preview(df: pd.DataFrame):
    """Formats values cleanly for preview display and applies italic style to Source File."""
    def apply_italic_source(data):
        styles = pd.DataFrame("", index=data.index, columns=data.columns)
        if "Source File" in data.columns:
            styles["Source File"] = "font-style: italic;"
        return styles

    display_df = df.copy().fillna("")
    display_df.index = range(1, len(display_df) + 1)
    return display_df.style.apply(apply_italic_source, axis=None)


def format_merged_sheet(writer, sheet_name: str, df: pd.DataFrame, dup_indices_set=None):
    """
    Applies professional styling to the 'Merged Data' sheet:
    - Sets cells to explicit Text format ('@') to protect 16+ digit numbers
    - Bold white-on-navy header
    - Auto-filter
    - Frozen panes at row 2
    - Auto-sized columns and text wrapping
    - Soft red/orange highlight for duplicate records
    """
    worksheet = writer.sheets[sheet_name]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    italic_font = Font(name="Calibri", size=11, italic=True)
    regular_font = Font(name="Calibri", size=11)
    
    dup_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    source_col_idx = None
    max_cols = len(df.columns)
    max_rows = len(df) + 1

    # Header styling & column width calculation
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_name == "Source File":
            source_col_idx = col_idx

        max_len = len(str(col_name))
        for value in df[col_name].head(250).tolist():
            if pd.isna(value) or value == "":
                continue
            val_str = str(value)
            if len(val_str) > max_len:
                max_len = len(val_str)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 4, 45), 12)

    # Data row styling
    for row_idx in range(2, max_rows + 1):
        df_row_idx = row_idx - 2
        is_dup_row = dup_indices_set and (df_row_idx in dup_indices_set)
        
        for col_idx in range(1, max_cols + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.number_format = "@"  # Explicit Text format to avoid scientific notation
            
            if col_idx == source_col_idx:
                cell.font = italic_font
            else:
                cell.font = regular_font

            if is_dup_row:
                cell.fill = dup_fill

    # Enable Excel Auto-Filter and Freeze Top Row
    if max_cols > 0:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_cols)}{max_rows}"
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 28


def format_duplicates_sheet(writer, sheet_name: str, dupes_df: pd.DataFrame, dup_group_series: pd.Series):
    """
    Applies structured styling to the 'Duplicates' sheet:
    - Preserves numbers as exact text
    - Dark red header
    - Alternating soft colors per duplicate group
    - Thick bottom borders separating distinct duplicate sets
    - Auto-filters & auto column sizing
    """
    worksheet = writer.sheets[sheet_name]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    italic_font = Font(name="Calibri", size=11, italic=True)
    regular_font = Font(name="Calibri", size=11)

    group_fill_a = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    group_fill_b = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    thin_border_side = Side(style="thin", color="BFBFBF")
    thick_bottom_side = Side(style="medium", color="000000")

    max_cols = len(dupes_df.columns)
    max_rows = len(dupes_df) + 1
    source_col_idx = None

    # Headers
    for col_idx, col_name in enumerate(dupes_df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_name == "Source File":
            source_col_idx = col_idx

        max_len = len(str(col_name))
        for value in dupes_df[col_name].head(250).tolist():
            if pd.isna(value) or value == "":
                continue
            val_str = str(value)
            if len(val_str) > max_len:
                max_len = len(val_str)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 4, 45), 12)

    # Data styling with group separation
    unique_groups = list(dup_group_series.unique())
    group_to_color_idx = {grp: idx % 2 for idx, grp in enumerate(unique_groups)}

    for row_idx in range(2, max_rows + 1):
        df_row_pos = row_idx - 2
        current_group = dup_group_series.iloc[df_row_pos]
        fill_color = group_fill_a if group_to_color_idx[current_group] == 0 else group_fill_b

        is_last_in_group = (
            df_row_pos == len(dupes_df) - 1 or
            dup_group_series.iloc[df_row_pos + 1] != current_group
        )

        row_bottom_border = thick_bottom_side if is_last_in_group else thin_border_side

        for col_idx in range(1, max_cols + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill_color
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.number_format = "@"  # Explicit Text format
            cell.border = Border(
                left=thin_border_side,
                right=thin_border_side,
                top=thin_border_side,
                bottom=row_bottom_border,
            )

            if col_idx == source_col_idx:
                cell.font = italic_font
            else:
                cell.font = regular_font

    if max_cols > 0:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_cols)}{max_rows}"
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 28


# --------------------------------------------------
# Main entry point
# --------------------------------------------------
def show():
    st.title("🔗 Master Merge & Duplicate Auditor")
    st.caption(
        "Upload multiple Excel files simultaneously. The file with the most columns sets the "
        "master column structure. Large numeric IDs (PRNs, Roll numbers) are protected and "
        "prevented from converting to scientific notation."
    )
    st.markdown("---")

    uploaded_files = st.file_uploader(
        "Upload Excel files to merge",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        st.info("Upload two or more Excel files to get started.")
        return

    if len(uploaded_files) == 1:
        st.warning("Only one file was uploaded — upload at least two files for a multi-file merge.")

    # --------------------------------------------------
    # Step 1: Pre-read uploaded files (dtype=str avoids large number truncation)
    # --------------------------------------------------
    raw_file_records = []
    read_errors = []
    all_discovered_cols = []
    total_raw_rows = 0

    for f in uploaded_files:
        try:
            # Read all columns as string to protect long identifiers
            df = pd.read_excel(f, sheet_name=0, dtype=str)
        except Exception as e:
            read_errors.append((f.name, str(e)))
            continue

        # Drop completely blank rows
        df = df.dropna(how="all")
        # Apply string sanitation across all cells
        df = df.map(sanitize_cell_value)
        # Drop rows that became completely empty after sanitation
        df = df[(df != "").any(axis=1)]

        total_raw_rows += len(df)

        if df.empty:
            st.warning(f"File **'{f.name}'** contains no valid data rows and was skipped.")
            continue

        # Clean column names (strip spaces, single whitespace)
        df.columns = [clean_column_name(c) for c in df.columns]

        for col in df.columns:
            if col not in all_discovered_cols:
                all_discovered_cols.append(col)

        raw_file_records.append({
            "file_name": f.name,
            "df": df,
            "col_count": len(df.columns),
        })

    if read_errors:
        st.error("Some files could not be read and were skipped:")
        for name, err in read_errors:
            st.write(f"- **{name}**: {err}")

    if not raw_file_records:
        st.stop()

    # --------------------------------------------------
    # Step 2: Simplified Merge & Audit Configuration UI
    # --------------------------------------------------
    with st.expander("⚙️ Merge & Duplicate Settings", expanded=True):
        col_left, col_right = st.columns(2)

        with col_left:
            st.markdown("##### 📁 Merge Options")
            merge_mode = st.radio(
                "Merge Strategy:",
                options=[
                    "Master / Union (Keep all columns)",
                    "Intersection (Keep common columns only)",
                ],
                index=0,
            )
            add_source_col = st.checkbox(
                "Add 'Source File' column",
                value=True,
            )
            case_insensitive = st.checkbox(
                "Standardize column casing (e.g., 'Roll No' = 'roll no')",
                value=True,
            )

        with col_right:
            st.markdown("##### 🔍 Duplicate Detection")
            selected_dedup_cols = st.multiselect(
                "Select column(s) to determine duplicates (leave empty to check entire row):",
                options=all_discovered_cols,
                default=[],
                help="Choose specific ID/Key columns like PRN or Roll No. Leaving this empty evaluates the complete row."
            )
            highlight_dupes = st.checkbox(
                "Highlight duplicates in output sheet",
                value=True,
            )
            remove_dupes = st.checkbox(
                "Remove duplicate records",
                value=False,
            )
            keep_option = st.selectbox(
                "When removing duplicates, keep:",
                options=["First Record", "Last Record"],
                index=0,
                disabled=not remove_dupes,
            )

    # --------------------------------------------------
    # Step 3: Canonical casing adjustments (if enabled)
    # --------------------------------------------------
    if case_insensitive:
        canonical_map = {}
        for item in raw_file_records:
            df = item["df"]
            new_cols = []
            for c in df.columns:
                key = c.lower()
                if key not in canonical_map:
                    canonical_map[key] = c
                new_cols.append(canonical_map[key])
            df.columns = new_cols

    # --------------------------------------------------
    # Step 4: Identify Master Column Structure
    # --------------------------------------------------
    master_record = max(raw_file_records, key=lambda x: x["col_count"])
    master_columns = list(master_record["df"].columns)
    master_file_name = master_record["file_name"]

    per_file_report = []
    for item in raw_file_records:
        per_file_report.append({
            "File": item["file_name"],
            "Rows": len(item["df"]),
            "Columns Found": item["col_count"],
            "Is Master Column Base": "⭐ Yes" if item["file_name"] == master_file_name else "No",
            "Cleaned Headers": ", ".join(item["df"].columns),
        })

    st.subheader("📋 Per-File Column Audit")
    report_df = pd.DataFrame(per_file_report)
    report_df.index = range(1, len(report_df) + 1)
    st.dataframe(report_df, use_container_width=True)

    st.info(
        f"👑 **Master Column Structure Base:** `{master_file_name}` "
        f"({len(master_columns)} columns). Missing columns in other files will be left blank."
    )

    # --------------------------------------------------
    # Step 5: Align and Standardize DataFrames
    # --------------------------------------------------
    is_intersection = merge_mode.startswith("Intersection")

    if is_intersection:
        common_cols = set(raw_file_records[0]["df"].columns)
        for item in raw_file_records[1:]:
            common_cols &= set(item["df"].columns)
        final_column_order = [c for c in master_columns if c in common_cols]
        for c in all_discovered_cols:
            if c in common_cols and c not in final_column_order:
                final_column_order.append(c)

        if not final_column_order:
            st.error("No common columns exist across all files for an intersection merge. Please switch to Master / Union mode.")
            st.stop()
    else:
        final_column_order = list(master_columns)
        for c in all_discovered_cols:
            if c not in final_column_order:
                final_column_order.append(c)

    aligned_frames = []
    for item in raw_file_records:
        current_df = item["df"].copy()
        for target_col in final_column_order:
            if target_col not in current_df.columns:
                current_df[target_col] = ""

        current_df = current_df[final_column_order]

        if add_source_col:
            current_df["Source File"] = item["file_name"]

        aligned_frames.append(current_df)

    merged_raw_df = pd.concat(aligned_frames, ignore_index=True, sort=False).fillna("")

    # --------------------------------------------------
    # Step 6: Duplicate Evaluation & Audit
    # --------------------------------------------------
    eval_cols = selected_dedup_cols if selected_dedup_cols else [c for c in merged_raw_df.columns if c != "Source File"]

    all_dup_mask = merged_raw_df.duplicated(subset=eval_cols, keep=False)
    total_duplicates_found = int(all_dup_mask.sum())

    duplicates_df = pd.DataFrame()
    dup_group_series = pd.Series(dtype=object)

    if total_duplicates_found > 0:
        raw_dupes = merged_raw_df[all_dup_mask].copy()

        def make_group_key(row):
            return " | ".join([f"{col}:{str(row[col])}" for col in eval_cols])

        raw_dupes["_dup_group_key"] = raw_dupes.apply(make_group_key, axis=1)
        sort_by = list(eval_cols) + (["Source File"] if add_source_col else [])
        raw_dupes = raw_dupes.sort_values(by=sort_by)

        dup_group_series = raw_dupes["_dup_group_key"].copy()
        duplicates_df = raw_dupes.drop(columns=["_dup_group_key"]).reset_index(drop=True)

    # --------------------------------------------------
    # Step 7: Final Data Generation (Deduplication)
    # --------------------------------------------------
    final_merged_df = merged_raw_df.copy()
    duplicates_removed_count = 0

    if remove_dupes:
        before_len = len(final_merged_df)
        keep_val = "first" if keep_option.startswith("First") else "last"
        final_merged_df = final_merged_df.drop_duplicates(
            subset=eval_cols,
            keep=keep_val,
        ).reset_index(drop=True)
        duplicates_removed_count = before_len - len(final_merged_df)

    highlight_indices_set = set()
    if highlight_dupes and not remove_dupes:
        highlight_indices_set = set(final_merged_df[final_merged_df.duplicated(subset=eval_cols, keep=False)].index)

    # --------------------------------------------------
    # Step 8: Executive Processing Summary
    # --------------------------------------------------
    st.markdown("---")
    st.subheader("📊 Processing Summary")
    
    kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
    kpi1.metric("Uploaded Files", len(raw_file_records))
    kpi2.metric("Total Rows (Before)", total_raw_rows)
    kpi3.metric("Total Rows (After)", len(final_merged_df))
    kpi4.metric("Duplicates Found", total_duplicates_found)
    kpi5.metric("Duplicates Removed", duplicates_removed_count)

    # --------------------------------------------------
    # Step 9: Previews
    # --------------------------------------------------
    st.markdown("---")
    st.subheader("✅ Merged Data Preview")
    st.write(
        f"**{len(final_merged_df)} total row(s)** across **{len(final_merged_df.columns)} column(s)**."
    )

    if "Source File" in final_merged_df.columns:
        preview_samples = []
        for _, grp in final_merged_df.groupby("Source File", sort=False):
            preview_samples.append(grp.head(10))
        preview_df = pd.concat(preview_samples, ignore_index=True)
        st.caption("🔍 Showing sample rows (up to 10) from each uploaded file:")
    else:
        preview_df = final_merged_df.head(50)

    st.write(style_preview(preview_df))

    if not duplicates_df.empty:
        with st.expander(f"⚠️ View Detected Duplicate Records ({len(duplicates_df)} rows)", expanded=False):
            st.dataframe(duplicates_df, use_container_width=True)

    # --------------------------------------------------
    # Step 10: Excel Export Builder
    # --------------------------------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: Merged Data
        final_merged_df.to_excel(writer, index=False, sheet_name="Merged Data")
        format_merged_sheet(writer, "Merged Data", final_merged_df, dup_indices_set=highlight_indices_set)

        # Sheet 2: Duplicates (Only generated if duplicates exist)
        if not duplicates_df.empty:
            duplicates_df.to_excel(writer, index=False, sheet_name="Duplicates")
            format_duplicates_sheet(writer, "Duplicates", duplicates_df, dup_group_series)

    output.seek(0)

    st.markdown("---")
    st.download_button(
        label="⬇️ Download Processed Excel Workbook (.xlsx)",
        data=output,
        file_name="master_merged_workbook.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Maintain Session State Counters if configured
    if "files_uploaded" in st.session_state:
        st.session_state.files_uploaded += len(raw_file_records)
    if "students_processed" in st.session_state:
        st.session_state.students_processed += len(final_merged_df)