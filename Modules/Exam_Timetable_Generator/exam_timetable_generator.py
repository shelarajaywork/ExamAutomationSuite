import datetime
import re
import io
from difflib import SequenceMatcher
import networkx as nx
import pandas as pd
import streamlit as st

# Distinct soft pastel hex colors for common-module groups
PALETTE_HEX = [
    "6ECDDB",  # Pastel Ocean
    "FFB347",  # Pastel Orange
    "EBCCFF",  # Pastel Mauve
    "B1C086",  # Pastel Avocado
    "E99FAA",  # Pastel Rouge
    "A4D8D8",  # Pastel Cyan
    "CFAC94",  # Pastel Cinnamon
    "AFC0EA",  # Pastel Denim
    "D1FEB8",  # Pastel Lime
    "F6B8D0",  # Pastel Rose
    "DBDBDC",  # Pastel Silver
    "FFA38C",  # Pastel Coral
    "A5E3E0",  # Pastel Blue green
    "FFDDB3",  # Pastel Caramel
    "B0E9D5",  # Pastel Aquamarine
    "FFA4A9",  # Pastel Tulip
    "D0E9C0",  # Pastel Pistachio
]

# ===================== HELPER FUNCTIONS =====================

def canonical_subject(s: str) -> str:
    """Normalises module descriptions for scheduling by removing common language/level suffixes."""
    if not isinstance(s, str):
        s = str(s) if pd.notna(s) else ""
    t = s.strip().upper()
    
    # Remove specific language variations
    t = t.replace(" IN ENGLISH", "").replace(" ENGLISH", "")
    
    # Trim trailing Roman numerals (I, II, III)
    while t.endswith(" I") or t.endswith(" II") or t.endswith(" III"):
        t = t.rsplit(" ", 1)[0]
        
    # Standardise multiple spaces
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def cluster_canonical_subjects(canon_subjects, threshold: float, canon_students: dict = None) -> dict:
    """Groups canonical subject strings that are similar to each other into the same cluster.

    Subjects that share at least one enrolled student are NEVER merged into the same cluster,
    even when their text similarity meets/exceeds the threshold. Two modules with a common
    student are genuinely different exams for that student (e.g. "Synthetic Organic Chemistry"
    vs "Synthetic And Sustainable Organic Chemistry") - merging them into one cluster/node would
    schedule both on the same day/time, causing a real clash for that student. `canon_students`
    maps each canonical subject string to the set of student identifiers enrolled in it, and is
    used purely to veto such unsafe merges; it does not affect merges between subjects that
    don't share any students.
    """
    canon_students = canon_students or {}

    if threshold >= 1.0:
        return {s: s for s in canon_subjects}

    representatives = []
    rep_students = {}  # representative -> union of student sets merged into its cluster so far
    mapping = {}
    for subj in sorted(canon_subjects):
        subj_students = canon_students.get(subj, set())
        best_ratio = 0.0
        best_rep = None
        for rep in representatives:
            # Never merge into a representative's cluster if doing so would put a shared
            # student into two different exams on the same day - irrespective of similarity.
            if subj_students and rep_students.get(rep) and (subj_students & rep_students[rep]):
                continue
            ratio = SequenceMatcher(None, subj, rep).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_rep = rep
        if best_rep is not None and best_ratio >= threshold:
            mapping[subj] = best_rep
            rep_students.setdefault(best_rep, set()).update(subj_students)
        else:
            representatives.append(subj)
            mapping[subj] = subj
            rep_students[subj] = set(subj_students)
    return mapping


def short_year(y) -> str:
    """Formats academic year string to short format (e.g. '2024-2025' -> '24-25')."""
    y_str = str(y).strip()
    if "-" in y_str:
        parts = y_str.split("-")
        if len(parts) == 2 and len(parts[0]) >= 2 and len(parts[1]) >= 2:
            return f"{parts[0][-2:]}-{parts[1][-2:]}"
    return y_str


def parse_duration_to_seconds(dur_val) -> int:
    """Parses duration string (HH:MM or HH:MM:SS) or timedelta object into seconds."""
    if pd.isna(dur_val) or str(dur_val).strip() == "":
        return 0
    
    dur_str = str(dur_val).strip()
    parts = dur_str.split(":")
    try:
        if len(parts) >= 2:
            hrs = int(parts[0])
            mins = int(parts[1])
            return hrs * 3600 + mins * 60
        else:
            return int(float(dur_str)) * 3600
    except ValueError:
        return 0


def format_duration(seconds: int) -> str:
    """Formats total seconds into 'HH:MM:SS' string format."""
    hrs = seconds // 3600
    mins = (seconds % 3600) // 60
    return f"{hrs:02d}:{mins:02d}:00"


def format_time_range(start_time_str: str, duration_seconds: int) -> str:
    """Calculates end time and returns a formatted range '10:00 AM - 12:00 PM'."""
    start_dt = datetime.datetime.strptime(start_time_str, "%I:%M %p")
    end_dt = start_dt + datetime.timedelta(seconds=duration_seconds)
    return f"{start_dt.strftime('%I:%M %p')} - {end_dt.strftime('%I:%M %p')}"


def format_time_for_timetable(time_str: str) -> str:
    """Formats a time range for the Timetable sheet:
    - Replaces ' - ' with ' to '
    - Lowercases AM/PM
    - Displays 12:00 pm (start or end) as '12:00 noon'
    """
    if not isinstance(time_str, str) or not time_str.strip():
        return str(time_str) if time_str is not None else ""

    s = time_str.replace(" - ", " to ")
    s = s.replace("AM", "am").replace("PM", "pm")

    if " to " in s:
        start_part, end_part = s.split(" to ", 1)
        if start_part.strip() == "12:00 pm":
            start_part = "12:00 noon"
        if end_part.strip() == "12:00 pm":
            end_part = "12:00 noon"
        s = f"{start_part} to {end_part}"
    else:
        if s.strip() == "12:00 pm":
            s = "12:00 noon"

    return s


def map_colors_to_dates(color_dict: dict, start_date: datetime.date, excluded_dates=None) -> dict:
    """Maps color/slot integers (1, 2, 3...) to actual date objects, automatically
    skipping Sundays and any additional dates supplied in excluded_dates."""
    excluded_set = set(excluded_dates) if excluded_dates else set()
    max_slot = max(color_dict.values()) if color_dict else 0
    day_date_map = {}
    current_date = start_date
    slot = 1

    while slot <= max_slot:
        if current_date.weekday() != 6 and current_date not in excluded_set:  # 6 = Sunday
            day_date_map[slot] = current_date
            slot += 1
        current_date += datetime.timedelta(days=1)
        
    return day_date_map


def generate_time_options():
    """Generates time options from 07:00 AM to 03:00 PM with 30-minute intervals."""
    times = []
    current = datetime.datetime.strptime("07:00 AM", "%I:%M %p")
    end = datetime.datetime.strptime("03:00 PM", "%I:%M %p")
    while current <= end:
        times.append(current.strftime("%I:%M %p"))
        current += datetime.timedelta(minutes=30)
    return times


# ===================== RE-EXAMINATION: DATA MERGE =====================

def _normalize_header(h) -> str:
    s = str(h).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _find_column(columns, exact_candidates, prefix_candidates=None):
    norm_map = {}
    for c in columns:
        norm_map.setdefault(_normalize_header(c), c)

    for cand in exact_candidates:
        if cand in norm_map:
            return norm_map[cand]

    if prefix_candidates:
        for norm, orig in norm_map.items():
            for prefix in prefix_candidates:
                if norm.startswith(prefix):
                    return orig
    return None


def normalize_reexam_acad_year(val) -> str:
    match = re.search(r"(\d{4})", str(val))
    return match.group(1) if match else str(val).strip()


def _clean_acad_report_year(val) -> str:
    try:
        return str(int(float(val))).strip()
    except (ValueError, TypeError):
        return str(val).strip().upper()


def build_reexam_merged_dataset(reexam_df: pd.DataFrame, acad_df: pd.DataFrame):
    r_prog_code = _find_column(reexam_df.columns, ["progcode"])
    r_prog_name = _find_column(reexam_df.columns, ["progname"])
    r_module_code = _find_column(reexam_df.columns, ["modulecode"])
    r_module_name = _find_column(reexam_df.columns, ["modulename"])
    r_student_number = _find_column(reexam_df.columns, ["studentnumber"])
    r_roll_number = _find_column(reexam_df.columns, ["rollnumber"])
    r_student_name = _find_column(reexam_df.columns, ["studentname"])
    r_acad_year = _find_column(reexam_df.columns, ["acadyear"])
    r_acad_session = _find_column(reexam_df.columns, ["acadsession"])
    r_gr_applicable = _find_column(reexam_df.columns, ["grapplicable"])
    r_appraisal_desc = _find_column(
        reexam_df.columns, ["appraisaldescription"], prefix_candidates=["appraisaldescri"]
    )

    missing_reexam = []
    for label, col in [
        ("Prog. Code", r_prog_code), ("Prog. Name", r_prog_name),
        ("Module code", r_module_code), ("Module name", r_module_name),
        ("Student number", r_student_number), ("Roll number", r_roll_number),
        ("Student name", r_student_name), ("Acad. Year", r_acad_year),
        ("Acad. Session", r_acad_session), ("GR Applicable", r_gr_applicable),
        ("Appraisal Description", r_appraisal_desc),
    ]:
        if not col:
            missing_reexam.append(label)

    a_school = _find_column(acad_df.columns, ["schoolname"])
    a_prog_abbr = _find_column(acad_df.columns, ["programabbreviation"])
    a_ay = _find_column(acad_df.columns, ["currentacademicyear"])
    a_session = _find_column(acad_df.columns, ["currentsession"])
    a_mod_abbr = _find_column(acad_df.columns, ["moduleabbreviation"])
    a_credit = _find_column(acad_df.columns, ["credit"])
    a_gr_applicable = _find_column(acad_df.columns, ["isgrapplicable"])
    a_exam_mode = _find_column(acad_df.columns, ["exammode"])
    a_exam_duration = _find_column(acad_df.columns, ["examduration"])

    missing_acad = []
    for label, col in [
        ("School Name", a_school), ("Program Abbreviation", a_prog_abbr),
        ("Current Academic Year", a_ay), ("Current Session", a_session),
        ("Module Abbreviation", a_mod_abbr), ("Credit", a_credit),
        ("Is GR Applicable?", a_gr_applicable), ("Exam Mode", a_exam_mode),
        ("Exam Duration", a_exam_duration),
    ]:
        if not col:
            missing_acad.append(label)

    if missing_reexam or missing_acad:
        parts = []
        if missing_reexam:
            parts.append(f"Re-exam Applications List is missing: {', '.join(missing_reexam)}")
        if missing_acad:
            parts.append(f"Academic Report is missing: {', '.join(missing_acad)}")
        return None, " | ".join(parts)

    reexam_f = reexam_df[
        (reexam_df[r_gr_applicable].astype(str).str.strip().str.upper() == "YES") &
        (reexam_df[r_appraisal_desc].astype(str).str.strip().str.upper() == "SEM ACTUAL")
    ].copy()

    acad_f = acad_df[
        (acad_df[a_exam_mode].astype(str).str.strip().str.upper() == "WRIT") &
        (acad_df[a_gr_applicable].astype(str).str.strip().str.upper() == "Y")
    ].copy()

    if reexam_f.empty:
        return pd.DataFrame(), None

    reexam_f["_Normalized Acad Year"] = reexam_f[r_acad_year].apply(normalize_reexam_acad_year)
    reexam_f["_lookup_key"] = (
        reexam_f[r_prog_code].astype(str).str.strip().str.upper() + "|" +
        reexam_f[r_module_code].astype(str).str.strip().str.upper() + "|" +
        reexam_f[r_acad_session].astype(str).str.strip().str.upper() + "|" +
        reexam_f["_Normalized Acad Year"]
    )

    acad_f["_lookup_key"] = (
        acad_f[a_prog_abbr].astype(str).str.strip().str.upper() + "|" +
        acad_f[a_mod_abbr].astype(str).str.strip().str.upper() + "|" +
        acad_f[a_session].astype(str).str.strip().str.upper() + "|" +
        acad_f[a_ay].apply(_clean_acad_report_year)
    )

    acad_lookup = acad_f.drop_duplicates(subset="_lookup_key", keep="first").set_index("_lookup_key")

    merged_rows = []
    for _, r in reexam_f.iterrows():
        key = r["_lookup_key"]
        match = acad_lookup.loc[key] if key in acad_lookup.index else None
        incomplete = match is None

        merged_rows.append({
            "School Name": match[a_school] if match is not None else "",
            "Program Abbreviation": str(r[r_prog_code]).strip(),
            "Program Name": str(r[r_prog_name]).strip(),
            "Current Academic Year": r["_Normalized Acad Year"],
            "Current Session": str(r[r_acad_session]).strip(),
            "Module Abbreviation": str(r[r_module_code]).strip(),
            "Module Description": str(r[r_module_name]).strip(),
            "Credit": match[a_credit] if match is not None else "",
            "Student Number": str(r[r_student_number]).strip(),
            "Student Roll Number": str(r[r_roll_number]).strip(),
            "Student Name": str(r[r_student_name]).strip(),
            "Exam Mode": match[a_exam_mode] if match is not None else "",
            "Exam Duration": match[a_exam_duration] if match is not None else "",
            "Incomplete Data": "Yes" if incomplete else "No",
        })

    merged_df = pd.DataFrame(merged_rows)
    return merged_df, None


# ===================== CORE SCHEDULING ALGORITHM =====================

def generate_timetable(
    df: pd.DataFrame,
    is_reexam: bool,
    start_date: datetime.date,
    start_time_str: str,
    fuzzy_threshold: float = 1.0,
    manual_module_map: dict = None,
    excluded_dates=None,
):
    col_map = {str(col).strip().lower().replace(" ", ""): col for col in df.columns}

    c_sub = col_map.get("moduledescription", col_map.get("subjectname"))
    c_stud = col_map.get("studentnumber")
    c_prog = col_map.get("programname", col_map.get("programmename"))
    c_college = col_map.get("schoolname", col_map.get("collegename"))
    c_ay = col_map.get("currentacademicyear", col_map.get("academicyear"))
    c_sem = col_map.get("currentsession", col_map.get("semester"))
    c_dur = col_map.get("examduration", col_map.get("duration"))
    
    c_prog_abbr = col_map.get("programabbreviation")
    c_mod_abbr = col_map.get("moduleabbreviation")
    c_credit = col_map.get("credit")
    c_student_na = col_map.get("studentna")

    missing = []
    if not c_sub: missing.append("Module Description")
    if not c_stud: missing.append("Student Number")
    if not c_prog: missing.append("Program Name")
    if not c_college: missing.append("School Name")
    if not c_ay: missing.append("Current Academic Year")
    if not c_sem: missing.append("Current Session")
    if not c_dur: missing.append("Exam Duration")

    if missing:
        st.error(f"Missing required columns in uploaded Excel file(s): {', '.join(missing)}")
        return None, {}

    if c_student_na and c_student_na in df.columns:
        df = df[df[c_student_na].astype(str).str.strip().str.upper() == "WRIT"].copy()

    df = df.dropna(subset=[c_sub, c_stud]).copy()

    manual_map_clean = {}
    if manual_module_map:
        for k, v in manual_module_map.items():
            k_clean = str(k).strip().upper()
            v_clean = str(v).strip()
            if k_clean and v_clean:
                manual_map_clean[k_clean] = v_clean

    canon_subject_pool = set()
    canon_to_students = {}  # canonical subject (pre-clustering) -> set of enrolled student ids
    for _, row in df.iterrows():
        subject_raw = str(row[c_sub]).strip()
        if not subject_raw:
            continue
        if subject_raw.upper() in manual_map_clean:
            continue
        base_canon = canonical_subject(subject_raw)
        canon_subject_pool.add(base_canon)
        stud_val = str(row[c_stud]).strip() if pd.notna(row[c_stud]) else ""
        if stud_val:
            canon_to_students.setdefault(base_canon, set()).add(stud_val)

    fuzzy_cluster_map = cluster_canonical_subjects(canon_subject_pool, fuzzy_threshold, canon_to_students)

    def effective_canon(subject_text: str) -> str:
        raw_key = str(subject_text).strip().upper()
        if raw_key in manual_map_clean:
            return canonical_subject(manual_map_clean[raw_key])
        base_canon = canonical_subject(subject_text)
        return fuzzy_cluster_map.get(base_canon, base_canon)

    subj_data = {}         
    subj_sem = {}          
    subj_dur = {}          
    student_to_nodes = {}  

    re_row_students = {}   
    re_row_ays = {}        

    for _, row in df.iterrows():
        subject = str(row[c_sub]).strip()
        stud = str(row[c_stud]).strip()
        if not subject or not stud:
            continue

        canon_sub = effective_canon(subject)
        sem = str(row[c_sem]).strip()
        ay = short_year(row[c_ay])
        dur_txt = str(row[c_dur]).strip()
        dur_sec = parse_duration_to_seconds(dur_txt)
        dur_fmt = format_duration(dur_sec)
        college = str(row[c_college]).strip()
        prog = str(row[c_prog]).strip()

        prog_abbr = str(row[c_prog_abbr]).strip() if c_prog_abbr and pd.notna(row[c_prog_abbr]) else ""
        mod_abbr = str(row[c_mod_abbr]).strip() if c_mod_abbr and pd.notna(row[c_mod_abbr]) else ""
        credit_val = row[c_credit] if c_credit and pd.notna(row[c_credit]) else ""

        node_key = f"{canon_sub}|{sem}|{dur_txt}"

        if node_key not in subj_data:
            subj_data[node_key] = {}
            subj_sem[node_key] = sem
            subj_dur[node_key] = dur_fmt

        if not is_reexam:
            row_key = f"{prog_abbr}|{prog}|{ay}|{mod_abbr}|{subject}|{credit_val}"
            if college not in subj_data[node_key]:
                subj_data[node_key][college] = {}
            if row_key not in subj_data[node_key][college]:
                subj_data[node_key][college][row_key] = set()
            subj_data[node_key][college][row_key].add(stud)
        else:
            row_key = f"{college}|{prog_abbr}|{prog}|{sem}|{mod_abbr}|{subject}|{credit_val}|{dur_txt}"
            if row_key not in re_row_students:
                re_row_students[row_key] = set()
                re_row_ays[row_key] = set()
            re_row_students[row_key].add(stud)
            re_row_ays[row_key].add(ay)

        if stud not in student_to_nodes:
            student_to_nodes[stud] = set()
        student_to_nodes[stud].add(node_key)

    G = nx.Graph()
    for node_key in subj_data.keys():
        if not is_reexam:
            total_students = sum(
                len(studs)
                for col in subj_data[node_key].values()
                for studs in col.values()
            )
        else:
            total_students = sum(
                len(re_row_students[rk])
                for rk in re_row_students
                if node_key in rk
            )
        G.add_node(node_key, weight=total_students)

    for stud, nodes in student_to_nodes.items():
        node_list = list(nodes)
        for i in range(len(node_list)):
            for j in range(i + 1, len(node_list)):
                G.add_edge(node_list[i], node_list[j])

    color_assignment = nx.coloring.greedy_color(G, strategy="DSATUR")
    color_assignment = {node: c + 1 for node, c in color_assignment.items()}

    day_date_map = map_colors_to_dates(color_assignment, start_date, excluded_dates=excluded_dates)

    output_rows = []

    if is_reexam:
        for row_key, stud_set in re_row_students.items():
            parts = row_key.split("|")
            coll_val, p_abbr, prog_val, sem_val, m_abbr, sub_val, cred_val, dur_val = (
                parts[0], parts[1], parts[2], parts[3], parts[4], parts[5], parts[6], parts[7]
            )
            
            node_key = f"{effective_canon(sub_val)}|{sem_val}|{dur_val}"
            assigned_slot = color_assignment.get(node_key, 1)
            dur_sec = parse_duration_to_seconds(dur_val)

            output_rows.append({
                "School Name": coll_val,
                "Program Abbreviation": p_abbr,
                "Program Name": prog_val,
                "Current Session": sem_val,
                "Module Abbreviation": m_abbr,
                "Module Description": sub_val,
                "Current Academic Year": ", ".join(sorted(re_row_ays[row_key])),
                "Credit": cred_val,
                "Students": len(stud_set),
                "Exam Duration": format_duration(dur_sec),
                "Exam Date": day_date_map.get(assigned_slot, ""),
                "Time": format_time_range(start_time_str, dur_sec),
                "Day": f"Day {assigned_slot}"
            })
    else:
        for node_key, colleges in subj_data.items():
            assigned_slot = color_assignment.get(node_key, 1)
            dur_sec = parse_duration_to_seconds(subj_dur[node_key])

            for college, row_keys in colleges.items():
                for row_key, stud_set in row_keys.items():
                    p = row_key.split("|")
                    p_abbr, prog_val, ay_val, m_abbr, sub_val, cred_val = (
                        p[0], p[1], p[2], p[3], p[4], p[5]
                    )
                    output_rows.append({
                        "School Name": college,
                        "Program Abbreviation": p_abbr,
                        "Program Name": prog_val,
                        "Current Session": subj_sem[node_key],
                        "Module Abbreviation": m_abbr,
                        "Module Description": sub_val,
                        "Current Academic Year": ay_val,
                        "Credit": cred_val,
                        "Students": len(stud_set),
                        "Exam Duration": subj_dur[node_key],
                        "Exam Date": day_date_map.get(assigned_slot, ""),
                        "Time": format_time_range(start_time_str, dur_sec),
                        "Day": f"Day {assigned_slot}"
                    })

    target_columns = [
        "School Name", "Program Abbreviation", "Program Name", "Current Session",
        "Module Abbreviation", "Module Description", "Current Academic Year",
        "Credit", "Students", "Exam Duration", "Exam Date", "Time", "Day"
    ]
    
    out_df = pd.DataFrame(output_rows)
    for col in target_columns:
        if col not in out_df.columns:
            out_df[col] = ""

    out_df = out_df[target_columns]

    # ---- IDENTIFY COMMON MODULE GROUPS & ASSIGN HEX COLORS ----
    canon_to_raws = {}
    for sub in out_df["Module Description"].unique():
        s_clean = str(sub).strip()
        if not s_clean:
            continue
        c_sub = effective_canon(s_clean)
        canon_to_raws.setdefault(c_sub, set()).add(s_clean)

    common_module_color_map = {}
    color_index = 0

    for c_sub, raw_set in canon_to_raws.items():
        is_common = False
        if len(raw_set) > 1:
            is_common = True
        else:
            matching_rows = out_df[out_df["Module Description"].isin(raw_set)]
            if len(matching_rows["Program Name"].unique()) > 1 or len(matching_rows) > 1:
                is_common = True

        if is_common:
            chosen_hex = PALETTE_HEX[color_index % len(PALETTE_HEX)]
            color_index += 1
            for raw_s in raw_set:
                common_module_color_map[raw_s] = chosen_hex

    return out_df, common_module_color_map


# ===================== TIMETABLE (PIVOT-STYLE) SHEET =====================

def build_timetable_sheet(
    workbook,
    result_df: pd.DataFrame,
    common_module_color_map: dict = None,
    exam_type: str = "",
    exam_start_date: datetime.date = None,
    show_academic_years: bool = True,
    show_student_count: bool = True,
):
    """Adds a 'Timetable' worksheet to the given workbook in a pivot-style layout.
    Header Row 2 displays Current Session (Semester) - suffixed with the selected
    Examination Type and the month/year of the Exam Start Date, e.g.
    "Semester III - Regular Examination - October 2026" - and merges across
    consecutive columns belonging to the same School + Semester. Header Rows 3 and 4
    display Program Abbreviation and Program Name as individual cells. Each Module
    Description cell can optionally also show that module's Current Academic Year
    value(s) (show_academic_years) and/or its total Students count
    (show_student_count), per the Timetable Sheet Display Options checkboxes."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet("Timetable")

    HEADER_ROWS = 4
    DATE_COL = 1
    TIME_COL = 2
    DATA_COL_START = 3
    FONT_NAME = "Times New Roman"

    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    day_separator_border = Border(left=thin, right=thin, top=thin, bottom=thick)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font_bold = Font(name=FONT_NAME, bold=True)
    font_normal = Font(name=FONT_NAME, bold=False)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    if result_df is None or result_df.empty:
        ws.sheet_view.showGridLines = False
        return ws

    col_df = result_df[["School Name", "Current Session", "Program Abbreviation", "Program Name"]].astype(str)
    col_df = col_df.sort_values(["School Name", "Current Session", "Program Abbreviation", "Program Name"])
    col_keys = []
    seen_cols = set()
    for _, r in col_df.iterrows():
        key = (r["School Name"], r["Current Session"], r["Program Abbreviation"], r["Program Name"])
        if key not in seen_cols:
            seen_cols.add(key)
            col_keys.append(key)

    def _time_sort_key(t):
        try:
            return datetime.datetime.strptime(str(t).split(" - ")[0].strip(), "%I:%M %p")
        except Exception:
            return datetime.datetime.min

    row_df = result_df[["Exam Date", "Time", "Exam Duration"]].copy()
    row_df["_dur_sort"] = row_df["Exam Duration"].apply(parse_duration_to_seconds)
    row_df["_time_sort"] = row_df["Time"].apply(_time_sort_key)
    row_df = row_df.sort_values(
        ["Exam Date", "_dur_sort", "_time_sort"],
        ascending=[True, False, True]
    )
    row_keys = []
    seen_rows = set()
    for _, r in row_df.iterrows():
        key = (r["Exam Date"], r["Time"])
        if key not in seen_rows:
            seen_rows.add(key)
            row_keys.append(key)

    cell_map = {}
    for _, r in result_df.iterrows():
        mk = (
            r["Exam Date"], r["Time"],
            str(r["School Name"]), str(r["Current Session"]), str(r["Program Abbreviation"]),
            str(r["Program Name"])
        )
        mod_desc = str(r["Module Description"]).strip()
        if not mod_desc:
            continue
        dur_seconds = parse_duration_to_seconds(r.get("Exam Duration", ""))

        try:
            students_val = int(r.get("Students", 0))
        except (ValueError, TypeError):
            students_val = 0

        cell_map.setdefault(mk, {})
        if mod_desc not in cell_map[mk]:
            cell_map[mk][mod_desc] = {"dur": dur_seconds, "ay": set(), "students": 0}

        # Current Academic Year may already be a comma-separated list (Re-Examination
        # combines multiple academic years for the same module into one cell).
        ay_raw = str(r.get("Current Academic Year", "")).strip()
        for ay_part in ay_raw.split(","):
            ay_part = ay_part.strip()
            if ay_part:
                cell_map[mk][mod_desc]["ay"].add(ay_part)

        # Sum students across any duplicate rows for the same module in the same cell
        # (e.g. separate cohorts of the same module within one program/date/time slot).
        cell_map[mk][mod_desc]["students"] += students_val

    ws.cell(row=1, column=DATE_COL, value="Exam Date")
    ws.cell(row=1, column=TIME_COL, value="Time")
    ws.merge_cells(start_row=1, start_column=DATE_COL, end_row=HEADER_ROWS, end_column=DATE_COL)
    ws.merge_cells(start_row=1, start_column=TIME_COL, end_row=HEADER_ROWS, end_column=TIME_COL)

    session_suffix_parts = [p for p in [exam_type, exam_start_date.strftime("%B %Y") if exam_start_date else ""] if p]
    session_suffix = " - " + " - ".join(session_suffix_parts) if session_suffix_parts else ""

    for idx, (school, session, prog_abbr, prog_name) in enumerate(col_keys):
        col_num = DATA_COL_START + idx
        ws.cell(row=2, column=col_num, value=f"{session}{session_suffix}")
        ws.cell(row=3, column=col_num, value=prog_abbr)
        ws.cell(row=4, column=col_num, value=prog_name)

    # Row 1 Merge: School Name across all its program columns
    if col_keys:
        start_idx = 0
        current_school = col_keys[0][0]
        for i in range(1, len(col_keys) + 1):
            if i == len(col_keys) or col_keys[i][0] != current_school:
                start_col = DATA_COL_START + start_idx
                end_col = DATA_COL_START + (i - 1)
                ws.cell(row=1, column=start_col, value=current_school)
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                if i < len(col_keys):
                    current_school = col_keys[i][0]
                    start_idx = i

        # Row 2 Merge: Current Session (Semester) across consecutive columns for same School + Session
        start_idx = 0
        current_session = (col_keys[0][0], col_keys[0][1])
        for i in range(1, len(col_keys) + 1):
            key_session = (col_keys[i][0], col_keys[i][1]) if i < len(col_keys) else None
            if i == len(col_keys) or key_session != current_session:
                start_col = DATA_COL_START + start_idx
                end_col = DATA_COL_START + (i - 1)
                if end_col > start_col:
                    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
                if i < len(col_keys):
                    current_session = key_session
                    start_idx = i

    date_groups = []
    if row_keys:
        start_idx = 0
        current_date = row_keys[0][0]
        for i in range(1, len(row_keys) + 1):
            if i == len(row_keys) or row_keys[i][0] != current_date:
                date_groups.append((current_date, start_idx, i - 1))
                if i < len(row_keys):
                    current_date = row_keys[i][0]
                    start_idx = i

    date_group_start = {g[1]: g[0] for g in date_groups}
    date_merge_ranges = [
        (HEADER_ROWS + 1 + g[1], HEADER_ROWS + 1 + g[2]) for g in date_groups if g[2] > g[1]
    ]
    day_separator_rows = {HEADER_ROWS + 1 + g[2] for g in date_groups[:-1]}

    cell_fill_map = {}

    for r_idx, (exam_date, time_val) in enumerate(row_keys):
        row_num = HEADER_ROWS + 1 + r_idx

        if r_idx in date_group_start:
            if isinstance(exam_date, (datetime.date, datetime.datetime)):
                day_line = exam_date.strftime("%A,")
                date_line = exam_date.strftime("%d.%m.%Y")
                date_text = f"{day_line}\n{date_line}"
            else:
                date_text = str(exam_date)
            ws.cell(row=row_num, column=DATE_COL, value=date_text)

        ws.cell(row=row_num, column=TIME_COL, value=format_time_for_timetable(time_val))

        for c_idx, (school, session, prog_abbr, prog_name) in enumerate(col_keys):
            col_num = DATA_COL_START + c_idx
            mods = cell_map.get((exam_date, time_val, school, session, prog_abbr, prog_name), {})

            if mods:
                sorted_mods = sorted(mods.items(), key=lambda kv: kv[1]["dur"], reverse=True)
                lines = []
                for m_desc, info in sorted_mods:
                    extras = []
                    if show_academic_years and info["ay"]:
                        extras.append(", ".join(sorted(info["ay"])))
                    if show_student_count:
                        extras.append(str(info["students"]))
                    line = f"{m_desc}\n({' | '.join(extras)})" if extras else m_desc
                    lines.append(line)
                value = " / \n".join(lines)

                if common_module_color_map:
                    for m_desc, _info in sorted_mods:
                        if m_desc in common_module_color_map:
                            cell_fill_map[(row_num, col_num)] = common_module_color_map[m_desc]
                            break
            else:
                value = "--"

            ws.cell(row=row_num, column=col_num, value=value)

    for start_row, end_row in date_merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=DATE_COL, end_row=end_row, end_column=DATE_COL)

    total_rows = HEADER_ROWS + len(row_keys)
    total_cols = DATA_COL_START + len(col_keys) - 1 if col_keys else DATA_COL_START - 1

    for col_num in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 30
    for row_num in range(1, total_rows + 1):
        ws.row_dimensions[row_num].height = 50

    ws.column_dimensions[get_column_letter(DATE_COL)].width = 18
    ws.column_dimensions[get_column_letter(TIME_COL)].width = 26

    for row_num in range(1, total_rows + 1):
        for col_num in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_align
            cell.border = day_separator_border if row_num in day_separator_rows else border
            is_header_row = row_num <= HEADER_ROWS
            is_date_time_col = col_num in (DATE_COL, TIME_COL)
            cell.font = font_bold if (is_header_row or is_date_time_col) else font_normal
            
            if is_header_row:
                cell.fill = header_fill
            elif (row_num, col_num) in cell_fill_map:
                fill_hex = cell_fill_map[(row_num, col_num)]
                cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    return ws


# ===================== ANALYSIS (PIVOT-STYLE) SHEET =====================

def build_analysis_sheet(workbook, result_df: pd.DataFrame):
    """Adds an 'Analysis' worksheet to the given workbook with row height 30 across all rows
    and auto-fitted column widths."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet("Analysis")

    HEADER_ROWS = 2
    DATE_COL = 1
    TIME_COL = 2
    DATA_COL_START = 3
    FONT_NAME = "Times New Roman"

    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    day_separator_border = Border(left=thin, right=thin, top=thin, bottom=thick)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font_bold = Font(name=FONT_NAME, bold=True)
    font_normal = Font(name=FONT_NAME, bold=False)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    if result_df is None or result_df.empty:
        ws.sheet_view.showGridLines = False
        return ws

    col_df = result_df[["School Name", "Current Session"]].astype(str)
    col_df = col_df.sort_values(["School Name", "Current Session"])
    col_keys = []
    seen_cols = set()
    for _, r in col_df.iterrows():
        key = (r["School Name"], r["Current Session"])
        if key not in seen_cols:
            seen_cols.add(key)
            col_keys.append(key)

    school_order = []
    seen_schools = set()
    for school, _ in col_keys:
        if school not in seen_schools:
            seen_schools.add(school)
            school_order.append(school)

    col_entries = []
    for school in school_order:
        for sch, session in col_keys:
            if sch == school:
                col_entries.append(("session", school, session))
        col_entries.append(("total", school, None))

    def _time_sort_key(t):
        try:
            return datetime.datetime.strptime(str(t).split(" - ")[0].strip(), "%I:%M %p")
        except Exception:
            return datetime.datetime.min

    row_df = result_df[["Exam Date", "Time", "Exam Duration"]].copy()
    row_df["_dur_sort"] = row_df["Exam Duration"].apply(parse_duration_to_seconds)
    row_df["_time_sort"] = row_df["Time"].apply(_time_sort_key)
    row_df = row_df.sort_values(
        ["Exam Date", "_dur_sort", "_time_sort"],
        ascending=[True, False, True]
    )
    row_keys = []
    seen_rows = set()
    for _, r in row_df.iterrows():
        key = (r["Exam Date"], r["Time"])
        if key not in seen_rows:
            seen_rows.add(key)
            row_keys.append(key)

    value_map = {}
    for _, r in result_df.iterrows():
        mk = (r["Exam Date"], r["Time"], str(r["School Name"]), str(r["Current Session"]))
        try:
            students_val = int(r.get("Students", 0) or 0)
        except (ValueError, TypeError):
            students_val = 0
        value_map[mk] = value_map.get(mk, 0) + students_val

    ws.cell(row=1, column=DATE_COL, value="Exam Date")
    ws.cell(row=1, column=TIME_COL, value="Time")
    ws.merge_cells(start_row=1, start_column=DATE_COL, end_row=HEADER_ROWS, end_column=DATE_COL)
    ws.merge_cells(start_row=1, start_column=TIME_COL, end_row=HEADER_ROWS, end_column=TIME_COL)

    total_cols_set = set()
    for idx, (kind, school, session) in enumerate(col_entries):
        col_num = DATA_COL_START + idx
        if kind == "total":
            ws.cell(row=2, column=col_num, value="Total Students")
            total_cols_set.add(col_num)
        else:
            ws.cell(row=2, column=col_num, value=session)

    if col_entries:
        start_idx = 0
        current_school = col_entries[0][1]
        for i in range(1, len(col_entries) + 1):
            if i == len(col_entries) or col_entries[i][1] != current_school:
                start_col = DATA_COL_START + start_idx
                end_col = DATA_COL_START + (i - 1)
                ws.cell(row=1, column=start_col, value=current_school)
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                if i < len(col_entries):
                    current_school = col_entries[i][1]
                    start_idx = i

    date_groups = []
    if row_keys:
        start_idx = 0
        current_date = row_keys[0][0]
        for i in range(1, len(row_keys) + 1):
            if i == len(row_keys) or row_keys[i][0] != current_date:
                date_groups.append((current_date, start_idx, i - 1))
                if i < len(row_keys):
                    current_date = row_keys[i][0]
                    start_idx = i

    date_group_start = {g[1]: g[0] for g in date_groups}
    date_merge_ranges = [
        (HEADER_ROWS + 1 + g[1], HEADER_ROWS + 1 + g[2]) for g in date_groups if g[2] > g[1]
    ]
    day_separator_rows = {HEADER_ROWS + 1 + g[2] for g in date_groups[:-1]}

    for r_idx, (exam_date, time_val) in enumerate(row_keys):
        row_num = HEADER_ROWS + 1 + r_idx

        if r_idx in date_group_start:
            if isinstance(exam_date, (datetime.date, datetime.datetime)):
                day_line = exam_date.strftime("%A,")
                date_line = exam_date.strftime("%d.%m.%Y")
                date_text = f"{day_line}\n{date_line}"
            else:
                date_text = str(exam_date)
            ws.cell(row=row_num, column=DATE_COL, value=date_text)

        ws.cell(row=row_num, column=TIME_COL, value=format_time_for_timetable(time_val))

        for c_idx, (kind, school, session) in enumerate(col_entries):
            col_num = DATA_COL_START + c_idx
            if kind == "total":
                total_students = sum(
                    value_map.get((exam_date, time_val, school, sess), 0)
                    for sch, sess in col_keys if sch == school
                )
            else:
                total_students = value_map.get((exam_date, time_val, school, session), 0)
            value = total_students if total_students else "--"
            ws.cell(row=row_num, column=col_num, value=value)

    for start_row, end_row in date_merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=DATE_COL, end_row=end_row, end_column=DATE_COL)

    total_rows = HEADER_ROWS + len(row_keys)
    total_cols = DATA_COL_START + len(col_entries) - 1 if col_entries else DATA_COL_START - 1

    # Set row height to 30 for all rows
    for row_num in range(1, total_rows + 1):
        ws.row_dimensions[row_num].height = 30

    # Apply cell styling and alignments
    for row_num in range(1, total_rows + 1):
        for col_num in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_align
            cell.border = day_separator_border if row_num in day_separator_rows else border
            is_header_row = row_num <= HEADER_ROWS
            is_date_time_col = col_num in (DATE_COL, TIME_COL)
            is_total_col = col_num in total_cols_set
            cell.font = font_bold if (is_header_row or is_date_time_col or is_total_col) else font_normal

            if is_header_row:
                cell.fill = header_fill
            elif is_total_col:
                cell.fill = total_fill

    # Auto-fit Column Widths dynamically
    for col_num in range(1, total_cols + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for row_num in range(1, total_rows + 1):
            val = ws.cell(row=row_num, column=col_num).value
            if val:
                lines = str(val).split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    return ws


# ===================== STREAMLIT USER INTERFACE =====================

def main():
    st.set_page_config(page_title="Exam Timetable Generator", layout="wide", page_icon="📅")
    
    st.title("📅 Exam Timetable Generator")
    st.divider()

    # --- CONFIGURATION PANEL IN MAIN WINDOW ---
    with st.container():
        left_col, right_col = st.columns([1, 3])

        with left_col:
            exam_type = st.radio(
                "Select Examination Type:",
                ("Regular Examination", "Re-Examination")
            )
            is_reexam = (exam_type == "Re-Examination")

            exam_start_date = st.date_input(
                "Exam Start Date:",
                value=datetime.date.today(),
                format="DD/MM/YYYY",
                help="Select the starting date for the examination schedule."
            )

            time_options = generate_time_options()
            default_index = time_options.index("10:00 AM") if "10:00 AM" in time_options else 0
            
            exam_start_time_str = st.selectbox(
                "Exam Start Time:",
                options=time_options,
                index=default_index,
                help="Select standard start time between 07:00 AM and 03:00 PM (30 min intervals)."
            )

            if not is_reexam:
                uploaded_files = st.file_uploader(
                    "Upload All Subjects Students Data files [ZACAD_REPORT]. (.xlsx, .xls)",
                    type=["xlsx", "xls"],
                    accept_multiple_files=True,
                    key="regular_uploader"
                )
            else:
                reexam_list_files = st.file_uploader(
                    "Upload Re-exam Applications List Data [ZREEXAM_REPORT] (.xlsx, .xls)",
                    type=["xlsx", "xls"],
                    accept_multiple_files=True,
                    key="reexam_list_uploader"
                )
                acad_report_files = st.file_uploader(
                    "Upload Acad Report of All Colleges and Years [ZACAD_REPORT] (.xlsx, .xls)",
                    type=["xlsx", "xls"],
                    accept_multiple_files=True,
                    key="reexam_acad_uploader"
                )
                uploaded_files = (reexam_list_files or []) + (acad_report_files or [])

        with right_col:
            with st.expander("⚙️ Advanced Scheduling Options (optional)", expanded=True):
                
                opt_col1, opt_col2 = st.columns(2)

                with opt_col1:
                    st.markdown("**Common Module Matching**")
                    similarity_pct = st.slider(
                        "Treat Module Descriptions as common when similarity is at least:",
                        min_value=50,
                        max_value=100,
                        value=75,
                        step=5,
                        format="%d%%",
                        help=(
                            "100% requires an exact match (current/default behaviour). Lower this to "
                            "automatically group subjects with minor spelling, singular/plural, or "
                            "wording differences, e.g. 'Business Law' vs 'Business Laws'."
                        ),
                        key="fuzzy_threshold_pct"
                    )
                    fuzzy_threshold = similarity_pct / 100.0

                with opt_col2:
                    st.markdown("**Additional Non-Working Days**")
                    holiday_text = st.text_area(
                        "(public/institutional holidays), one per line — DD-MM-YYYY:",                        
                        value="",
                        height=110,
                        help="Sundays are always excluded automatically. Dates listed here are also skipped when assigning exam days.",
                        key="holiday_exclusion_text"
                    )

                excluded_dates = []
                invalid_holiday_lines = []
                for line in holiday_text.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    parsed_date = None
                    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
                        try:
                            parsed_date = datetime.datetime.strptime(line, fmt).date()
                            break
                        except ValueError:
                            continue
                    if parsed_date:
                        excluded_dates.append(parsed_date)
                    else:
                        invalid_holiday_lines.append(line)

                if invalid_holiday_lines:
                    st.warning(
                        "Could not parse the following exclusion date(s) — they will be ignored. "
                        f"Please use DD-MM-YYYY format: {', '.join(invalid_holiday_lines)}"
                    )

                st.markdown("**Timetable Sheet Display Options**")
                tt_opt_col1, tt_opt_col2 = st.columns(2)
                with tt_opt_col1:
                    show_academic_years = st.checkbox(
                        "Show Academic Years",
                        value=True,
                        help=(
                            "Show each module's Current Academic Year value(s) alongside its name in "
                            "every Timetable sheet cell."
                        ),
                        key="show_academic_years_chk"
                    )
                with tt_opt_col2:
                    show_student_count = st.checkbox(
                        "Show Student Count",
                        value=True,
                        help=(
                            "Show each module's total Students count alongside its name in every "
                            "Timetable sheet cell."
                        ),
                        key="show_student_count_chk"
                    )

                st.markdown("**Manual Common Module Mapping**")
                st.caption(
                    "If the common modules have variations in module descriptions across programs/colleges, "
                    "give them the same Common Group Name below to make sure those modules "
                    "are scheduled on the same day and time."                    
                )
                empty_mapping_df = pd.DataFrame(
                    {
                        "Module Description (as in file)": pd.Series([], dtype="object"),
                        "Common Group Name": pd.Series([], dtype="object"),
                    }
                )
                mapping_df = st.data_editor(
                    empty_mapping_df,
                    num_rows="dynamic",
                    use_container_width=True,
                    column_config={
                        "Module Description (as in file)": st.column_config.TextColumn(
                            "Module Description (as in file)",
                            help="Exact Module Description text as it appears in the uploaded file.",
                            width="large",
                        ),
                        "Common Group Name": st.column_config.TextColumn(
                            "Common Group Name",
                            help="A shared label — any modules mapped to the same group name are scheduled together.",
                            width="medium",
                        ),
                    },
                    key="common_module_mapping_editor"
                )

                manual_module_map = {}
                if mapping_df is not None and not mapping_df.empty:
                    for _, m_row in mapping_df.iterrows():
                        desc = str(m_row.get("Module Description (as in file)", "")).strip()
                        group = str(m_row.get("Common Group Name", "")).strip()
                        if desc and desc.lower() != "nan" and group and group.lower() != "nan":
                            manual_module_map[desc] = group

    st.divider()

    # --- CONDITIONAL DATA PROCESSING SECTIONS ---
    if uploaded_files:
        try:
            incomplete_df = None

            if not is_reexam:
                df_list = [pd.read_excel(file) for file in uploaded_files]
                df_input = pd.concat(df_list, ignore_index=True)
                st.success(f"Successfully loaded {len(uploaded_files)} file(s) with a total of {len(df_input)} rows.")

            else:
                if not reexam_list_files or not acad_report_files:
                    st.warning(
                        "Please upload both the Re-exam Applications List [ZREEXAM_REPORT] and the "
                        "Academic Report [ZACAD_REPORT] to proceed."
                    )
                    df_input = pd.DataFrame()
                else:
                    reexam_raw_df = pd.concat(
                        [pd.read_excel(f) for f in reexam_list_files], ignore_index=True
                    )
                    acad_raw_df = pd.concat(
                        [pd.read_excel(f) for f in acad_report_files], ignore_index=True
                    )

                    merged_df, merge_error = build_reexam_merged_dataset(reexam_raw_df, acad_raw_df)

                    if merge_error:
                        st.error(f"Could not process the uploaded files: {merge_error}")
                        df_input = pd.DataFrame()
                    elif merged_df.empty:
                        st.warning(
                            "No records matched the required filters (GR Applicable = YES and "
                            "Appraisal Description = Sem Actual in the Re-exam Applications List)."
                        )
                        df_input = pd.DataFrame()
                    else:
                        incomplete_df = merged_df[merged_df["Incomplete Data"] == "Yes"].reset_index(drop=True)
                        df_input = (
                            merged_df[merged_df["Incomplete Data"] == "No"]
                            .drop(columns=["Incomplete Data"])
                            .reset_index(drop=True)
                        )

                        summary_msg = (
                            f"Successfully processed {len(reexam_list_files)} Re-exam Applications List "
                            f"file(s) and {len(acad_report_files)} Academic Report file(s): "
                            f"{len(df_input)} record(s) ready for scheduling"
                        )
                        if not incomplete_df.empty:
                            summary_msg += f", {len(incomplete_df)} flagged as Incomplete Data."
                        else:
                            summary_msg += "."
                        st.success(summary_msg)

                        if not incomplete_df.empty:
                            st.warning(
                                f"⚠️ {len(incomplete_df)} record(s) from the Re-exam Applications List could "
                                "not be matched with a corresponding Academic Report record (missing or "
                                "unmatched Prog. Code + Module code + Acad. Session + Acad. Year key). "
                                "These records are flagged as **Incomplete Data**, excluded from the "
                                "generated timetable, and listed below for review. They are also included "
                                "in a separate 'Incomplete Data' sheet in the downloaded workbook."
                            )
                            with st.expander(f"⚠️ View Incomplete Data ({len(incomplete_df)} record(s))"):
                                st.dataframe(incomplete_df, use_container_width=True)

            if not df_input.empty:
                with st.expander("Preview Combined Data"):
                    st.dataframe(df_input.head(10))

                if st.button("🚀 Generate Timetable", type="primary"):
                    with st.spinner("Building conflict graph and calculating schedules..."):
                        result_df, common_module_color_map = generate_timetable(
                            df=df_input,
                            is_reexam=is_reexam,
                            start_date=exam_start_date,
                            start_time_str=exam_start_time_str,
                            fuzzy_threshold=fuzzy_threshold,
                            manual_module_map=manual_module_map,
                            excluded_dates=excluded_dates
                        )

                    if result_df is not None and not result_df.empty:
                        st.success("Timetable generated successfully!")
                        
                        st.subheader(f"Generated Schedule ({exam_type})")
                        st.dataframe(result_df, use_container_width=True)

                        # Export to Excel with custom formatting
                        output_bytes = io.BytesIO()
                        with pd.ExcelWriter(output_bytes, engine="openpyxl") as writer:
                            sheet_name = "Re-Examination" if is_reexam else "Regular Examination"
                            result_df.to_excel(writer, index=False, sheet_name=sheet_name)
                            
                            workbook = writer.book
                            worksheet = writer.sheets[sheet_name]
                            
                            # Freeze top header row
                            worksheet.freeze_panes = "A2"
                            
                            from openpyxl.styles import Font, PatternFill
                            
                            header_font = Font(bold=True)
                            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            
                            for col_num in range(1, len(result_df.columns) + 1):
                                cell = worksheet.cell(row=1, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill

                            # Apply Short Date format (DD-MM-YYYY) to 'Exam Date' column
                            date_col_idx = result_df.columns.get_loc("Exam Date") + 1
                            for row_num in range(2, len(result_df) + 2):
                                cell = worksheet.cell(row=row_num, column=date_col_idx)
                                if cell.value:
                                    cell.number_format = "DD-MM-YYYY"

                            # Apply Color-Coding for Common Modules in Main Output Sheet
                            if common_module_color_map:
                                desc_col_idx = result_df.columns.get_loc("Module Description") + 1
                                for row_num in range(2, len(result_df) + 2):
                                    cell = worksheet.cell(row=row_num, column=desc_col_idx)
                                    mod_val = str(cell.value).strip() if cell.value else ""
                                    if mod_val in common_module_color_map:
                                        fill_hex = common_module_color_map[mod_val]
                                        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

                            # Add the pivot-style "Timetable" worksheet
                            build_timetable_sheet(
                                workbook, result_df, common_module_color_map,
                                exam_type=exam_type, exam_start_date=exam_start_date,
                                show_academic_years=show_academic_years,
                                show_student_count=show_student_count
                            )

                            # Add the "Analysis" worksheet
                            build_analysis_sheet(workbook, result_df)

                            # Include unmatched Re-exam Applications List records for transparency
                            if is_reexam and incomplete_df is not None and not incomplete_df.empty:
                                incomplete_df.to_excel(writer, index=False, sheet_name="Incomplete Data")
                                incomplete_ws = writer.sheets["Incomplete Data"]
                                incomplete_ws.freeze_panes = "A2"
                                for col_num in range(1, len(incomplete_df.columns) + 1):
                                    cell = incomplete_ws.cell(row=1, column=col_num)
                                    cell.font = header_font
                                    cell.fill = header_fill

                        # Dynamic Output File Name
                        curr_date_str = datetime.date.today().strftime("%d-%m-%Y")
                        if not is_reexam:
                            file_name = f"Regular Examination Timetable {curr_date_str}.xlsx"
                        else:
                            file_name = f"Re-examination Timetable {curr_date_str}.xlsx"
                        
                        st.download_button(
                            label="📥 Download Timetable Excel",
                            data=output_bytes.getvalue(),
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        except Exception as e:
            st.error(f"Error processing files: {e}")

    st.divider()

    render_help_section()


def render_help_section():
    """
    Renders a collapsible 'Tool Description / Help' section. Collapsed by
    default (expanded=False) - it only opens when the user clicks it. Purely
    documentation; no processing logic lives here.
    """
    with st.expander("ℹ️ Tool Description / Help - click to expand", expanded=False):
        st.markdown(
            """
### 1. Purpose of the Tool
The Exam Timetable Generator builds a clash-free examination schedule from student-module
enrolment data. It groups students by the modules they're enrolled in, treats any two modules
shared by at least one common student as a scheduling conflict, and uses a graph-coloring
algorithm to assign each module a day so that no student is ever scheduled for two exams on
the same day. It supports both **Regular Examination** and **Re-Examination** scheduling, and
exports a fully formatted, multi-sheet Excel workbook (raw schedule, pivot-style Timetable
view, and a summary Analysis sheet).

---
### 2. Required Input Files & Formats
**Regular Examination** - one or more Excel files (`.xlsx`, `.xls`) exported as
**[ZACAD_REPORT]** ("All Subjects Students Data"), each row representing one student's
enrolment in one module. Required columns (matched flexibly by name/casing/spacing):
`Module Description`, `Student Number`, `Program Name`, `School Name`,
`Current Academic Year`, `Current Session`, `Exam Duration`. Optional columns
`Program Abbreviation`, `Module Abbreviation`, `Credit`, and `Student NA` (rows are filtered
to `WRIT` if present) are used when available.

**Re-Examination** - two separate sets of files are required together:
- **Re-exam Applications List [ZREEXAM_REPORT]** - student-level re-exam applications. Needs
  `Prog. Code`, `Prog. Name`, `Module code`, `Module name`, `Student number`, `Roll number`,
  `Student name`, `Acad. Year`, `Acad. Session`, `GR Applicable`, `Appraisal Description`.
- **Academic Report [ZACAD_REPORT]** - the master module/credit reference data. Needs
  `School Name`, `Program Abbreviation`, `Current Academic Year`, `Current Session`,
  `Module Abbreviation`, `Credit`, `Is GR Applicable?`, `Exam Mode`, `Exam Duration`.

Both file types can be uploaded as multiple files each; all files of the same type are
concatenated before processing.

---
### 3. Description of All Upload/Configuration Options
- **Select Examination Type** - toggles between Regular Examination and Re-Examination, which
  changes which file uploader(s) are shown and how records are merged/filtered.
- **Exam Start Date** - the first calendar date exams may be scheduled on; scheduling always
  moves forward in time from this date.
- **Exam Start Time** - a fixed daily start time (07:00 AM-03:00 PM, 30-minute steps) used for
  every exam's time slot; end time is derived by adding that module's Exam Duration.
- **Upload All Subjects Students Data files [ZACAD_REPORT]** *(Regular only)* - the enrolment
  data described above.
- **Upload Re-exam Applications List [ZREEXAM_REPORT]** / **Upload Acad Report
  [ZACAD_REPORT]** *(Re-Examination only)* - the two file sets merged together (see Matching
  Logic below).
- **Advanced Scheduling Options (optional)**:
  - **Common Module Matching (similarity slider)** - at 100% (default-equivalent), only
    modules with identical descriptions are grouped together. Lowering it clusters module
    descriptions that are textually similar (e.g. "Business Law" vs "Business Laws") using
    fuzzy string matching, so near-duplicate module names are scheduled on the same day/slot.
    Two modules are never clustered together, regardless of similarity score, if they share
    at least one enrolled student - since merging them would schedule two different exams for
    that student at the same day/time (see Grouping, Clustering & Other Calculations below).
  - **Additional Non-Working Days** - a free-text box, one date per line in `DD-MM-YYYY`
    format, listing public/institutional holidays to skip when assigning exam dates. Sundays
    are always skipped automatically regardless of this list. Unparseable lines are flagged
    with a warning and ignored.
  - **Timetable Sheet Display Options** - two checkboxes, both ticked by default:
    - **Show Academic Years** - when ticked, each module's Current Academic Year value(s) are
      shown alongside its name in every Timetable sheet cell.
    - **Show Student Count** - when ticked, each module's total Students count is shown
      alongside its name in every Timetable sheet cell.
  - **Manual Common Module Mapping** - an editable table for explicitly pairing exact module
    description strings (as they appear in the uploaded file) to a shared "Common Group
    Name", overriding the automatic fuzzy clustering for those specific modules.

---
### 4. Matching Logic Between Re-Exam Applications List & Academic Report
Records are merged using a composite key of **Program Abbreviation/Code + Module Abbreviation/
Code + Session + Academic Year** (all upper-cased and stripped), built independently on each
file using its own column names, then matched against each other. The Re-exam Applications
List is first filtered to rows where `GR Applicable = YES` and `Appraisal Description = SEM
ACTUAL`; the Academic Report is filtered to `Exam Mode = WRIT` and `Is GR Applicable? = Y`.
Any Re-exam Applications List row whose key finds no match in the filtered Academic Report is
flagged **Incomplete Data** and excluded from scheduling (see Validation Rules below).

---
### 5. Conditions & Filters Applied During Processing
- Rows missing a Module Description or Student Number are dropped.
- Where a `Student NA` column is present, only rows marked `WRIT` are kept.
- (Re-Examination) Only `GR Applicable = YES` + `Appraisal Description = SEM ACTUAL`
  applications, matched against `Exam Mode = WRIT` + `Is GR Applicable? = Y` academic
  records, are scheduled.

---
### 6. Grouping, Clustering & Other Calculations
- **Canonical subject name** - each Module Description is normalized (uppercased, language
  suffixes like "IN ENGLISH" removed, trailing Roman numerals I/II/III trimmed, whitespace
  collapsed) so that level/variant labels don't accidentally split or merge modules.
- **Fuzzy clustering** - canonical names are clustered using text-similarity ratio against the
  similarity threshold set in Advanced Scheduling Options; a name only joins an existing
  cluster if its similarity to that cluster's representative meets the threshold **and** the
  two share no enrolled student in common. A shared student always blocks the merge, however
  high the similarity score, so two genuinely different modules (e.g. "Synthetic Organic
  Chemistry" vs "Synthetic And Sustainable Organic Chemistry") are never forced onto the same
  day/slot just because a fuzzy match grouped their names together.
- **Manual mapping** - always takes priority over automatic canonicalisation/clustering for
  any module description explicitly listed in the mapping table.
- **Conflict graph & day assignment** - one graph node per canonical module (+ session +
  duration); an edge is added between two nodes whenever at least one student is enrolled in
  both. A DSATUR greedy graph-coloring algorithm assigns each node a "color" (day slot) such
  that no two connected nodes share a color, guaranteeing no student has two exams the same
  day. Colors are then mapped to actual calendar dates starting from the Exam Start Date,
  automatically skipping Sundays and any configured holiday dates.
- **Exam Duration / Time** - parsed from `HH:MM`/`HH:MM:SS` text (or a numeric hour count) into
  seconds, then combined with the fixed Exam Start Time to produce a display time range (e.g.
  "10:00 AM - 12:00 PM", shown as "10:00 am to 12:00 noon" on the Timetable sheet).
- **Common module identification & colour-coding** - a canonical module is treated as
  "common" (shared across programs/colleges) if it maps from more than one raw description, or
  appears against more than one Program Name, or has more than one enrolment row. Each common
  module group is assigned a distinct pastel colour from a fixed palette, applied as a cell
  fill on its Module Description cells throughout the output.

---
### 7. Re-Examination vs Regular Examination Processing Logic
- **Regular Examination** rows are grouped per School + Program + Academic Year + Module +
  Credit combination, each becoming its own output row with its own student count.
- **Re-Examination** rows are grouped per School + Program + Session + Module + Credit +
  Duration combination (Academic Years for that group are combined into one comma-separated
  cell), since re-exam applicants for the same module can span multiple academic years.
- Both modes ultimately produce the same output column layout: School Name, Program
  Abbreviation, Program Name, Current Session, Module Abbreviation, Module Description,
  Current Academic Year, Credit, Students, Exam Duration, Exam Date, Time, Day.

---
### 8. Validation Rules & the "Incomplete Data" Sheet *(Re-Examination only)*
Any Re-exam Applications List record that cannot be matched to a corresponding Academic
Report record (via the composite key above) is flagged **Incomplete Data**, shown in an
on-screen warning + expandable preview, excluded from the generated timetable, and written to
a separate **"Incomplete Data"** sheet in the downloaded workbook for manual follow-up.

---
### 9. Output Excel Workbook - Sheets & Contents
- **"Regular Examination" / "Re-Examination"** (main data sheet) - one row per scheduled
  module/program/college combination with the columns listed in section 7. Header row is
  bold with a grey fill and frozen; `Exam Date` is formatted `DD-MM-YYYY`; Module Description
  cells for common (shared) modules are filled with their assigned pastel colour.
- **"Timetable"** - a pivot-style grid: rows are Exam Date + Time (sorted by date, then by
  longest duration, then by time), columns are School → Session → Program Abbreviation →
  Program Name (merged header cells across rows 1-4), and each cell lists the module(s)
  scheduled for that program at that date/time. Row 2 shows Current Session suffixed with the
  selected Examination Type and the month/year of Exam Start Date, e.g. "Semester III - Regular
  Examination - October 2026". Each module's name in a cell is optionally followed, on the next
  line, by its Current Academic Year value(s) and/or total Students count in parentheses -
  e.g. "Communication Skills I" on one line and "(2025 | 2)" on the next - per the Timetable
  Sheet Display Options checkboxes (both on by default). Thick borders mark day boundaries;
  common modules keep their colour-coding.
- **"Analysis"** - a summary grid (School → Session columns, with a Total column per school)
  showing module/enrolment counts per day/time slot, for a quick at-a-glance load check.
- **"Incomplete Data"** *(Re-Examination only, when applicable)* - the unmatched Re-exam
  Applications List records described in section 8.

---
### 10. Preview Sections Shown On-Screen
- **Preview Combined Data** - the first 10 rows of the combined/merged input, before
  scheduling, so you can sanity-check the uploaded files were read correctly.
- **⚠️ View Incomplete Data** *(Re-Examination only)* - an expandable table of any unmatched
  applications, shown before you generate the timetable.
- **Generated Schedule** - the full scheduled output table, shown on-screen immediately after
  clicking "🚀 Generate Timetable", matching the main sheet of the downloaded workbook.

---
### 11. Downloadable Output
- **📥 Download Timetable Excel** - the complete workbook described in section 9, named
  `Regular Examination Timetable DD-MM-YYYY.xlsx` or `Re-examination Timetable
  DD-MM-YYYY.xlsx` depending on the selected exam type, using today's date.

---
### 12. Assumptions & Processing Workflow
1. Select the examination type and set the start date/time and any advanced options.
2. Upload the required file(s) for that exam type.
3. The tool loads and combines the files (for Re-Examination, merges the Applications List
   against the Academic Report and flags unmatched rows as Incomplete Data).
4. Click **🚀 Generate Timetable** to build the conflict graph, run the day-assignment
   algorithm, and produce the scheduled output.
5. Review the on-screen schedule, then download the formatted multi-sheet Excel workbook.

**Assumptions**: column names in uploaded files may vary in casing/spacing/prefix and are
matched flexibly rather than requiring an exact header match; a student is uniquely identified
by Student Number within a file; two modules are only treated as a scheduling conflict if they
share at least one enrolled student in common.
            """
        )


if __name__ == "__main__":
    main()