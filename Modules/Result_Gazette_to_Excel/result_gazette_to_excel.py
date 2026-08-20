import io
import re
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# HELPER PARSING FUNCTIONS
# ==============================================================================


def clean_text_preserve_symbols(val):
    """Clean whitespace while preserving academic notations ($, ~, /, #, +)."""
    if val is None:
        return ""
    return str(val).strip()


def clean_dashes_and_artifacts(text):
    """Remove continuous dash strings (---...) and redundant whitespace."""
    if not text:
        return ""
    text = re.sub(r"[-_=]{3,}", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_page_1(text):
    """Extract metadata key-values and accurately identify Regular vs Re-examination."""
    metadata = []
    text_clean = re.sub(r"\s+", " ", text)

    # Institution / Trust
    inst_match = re.search(
        r"(Shri Vile Parle Kelavani Mandal[^\n\r]*[\s\S]*?(?:AUTONOMOUS|\(Empowered Autonomous\)|\(Autonomous\)))",
        text,
        re.IGNORECASE,
    )
    inst_name = (
        inst_match.group(1).replace("\n", " ")
        if inst_match
        else "SVKM Educational Institution (AUTONOMOUS)"
    )
    metadata.append({"Property": "Institution / Trust", "Value": inst_name})

    # Programme
    prog_match = re.search(
        r"(Bachelor of [^\n\r,]+|Master of [^\n\r,]+|M\.Com[^\n\r,]*|B\.Sc[^\n\r,]*|B\.Com[^\n\r,]*)",
        text,
        re.IGNORECASE,
    )
    metadata.append(
        {
            "Property": "Programme",
            "Value": prog_match.group(1).strip() if prog_match else "N/A",
        }
    )

    # Semester
    sem_match = re.search(r"Semester\s+([IVXLCDM\d]+)", text, re.IGNORECASE)
    metadata.append(
        {
            "Property": "Semester",
            "Value": f"Semester {sem_match.group(1)}" if sem_match else "N/A",
        }
    )

    # NCrF Level
    ncrf_match = re.search(r"NCrF\s*Level[:\s]*([0-9.]+)", text, re.IGNORECASE)
    metadata.append(
        {
            "Property": "NCrF Level",
            "Value": ncrf_match.group(1) if ncrf_match else "N/A",
        }
    )

    # Academic Year
    ay_match = re.search(r"A\.Y\.?\s*([0-9]{4}\s*-\s*[0-9]{4})", text, re.IGNORECASE)
    metadata.append(
        {
            "Property": "Academic Year",
            "Value": ay_match.group(1) if ay_match else "N/A",
        }
    )

    # Examination Type: Accurately identify only Regular Examination or Re-examination
    if re.search(r"Re\s*[-–]?\s*examination", text_clean, re.IGNORECASE):
        exam_type = "Re-examination"
    elif re.search(r"Regular\s+examination", text_clean, re.IGNORECASE):
        exam_type = "Regular Examination"
    elif re.search(r"ATKT\s+examination", text_clean, re.IGNORECASE):
        exam_type = "ATKT Examination"
    else:
        exam_type = (
            "Re-examination"
            if re.search(r"\bRe\b", text_clean, re.I)
            else "Regular Examination"
        )

    # Examination Session
    exam_session_match = re.search(
        r"(?:examination\s+held\s+in|held\s+in)\s+([A-Za-z]+[\s,]*(?:\d{4})?)",
        text_clean,
        re.IGNORECASE,
    )
    exam_session = (
        f"{exam_type} held in {exam_session_match.group(1)}".strip()
        if exam_session_match
        else exam_type
    )

    metadata.append({"Property": "Examination Type", "Value": exam_type})
    metadata.append({"Property": "Examination Session", "Value": exam_session})

    return pd.DataFrame(metadata), exam_type


def parse_page_2(page):
    """Extract course catalogue and schema from Page 2."""
    text = page.extract_text() or ""
    lines = text.split("\n")
    courses = []

    course_pattern = re.compile(
        r"^([A-Z0-9]{5,12})\s+(.+?)\s+(\d+\.\d{2})\s+(\d+|--)\s+(\d+|--)\s+(\d+)"
    )

    for line in lines:
        m = course_pattern.match(line.strip())
        if m:
            code, title, creds, int_m, sem_m, tot_m = m.groups()
            courses.append(
                {
                    "Course Code": code.strip(),
                    "Course Title": title.strip(),
                    "Credits": float(creds),
                    "Internal Max Marks": int_m if int_m != "--" else 0,
                    "Sem End Max Marks": sem_m if sem_m != "--" else 0,
                    "Total Max Marks": int(tot_m),
                }
            )

    df = pd.DataFrame(courses)
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "Course Code",
                "Course Title",
                "Credits",
                "Internal Max Marks",
                "Sem End Max Marks",
                "Total Max Marks",
            ]
        )
    return df


def clean_page_lines(raw_text):
    """Filter out non-student header/footer lines, separator dashes, and execution logs."""
    lines = raw_text.split("\n")
    cleaned_lines = []

    for line in lines:
        l_str = line.strip()
        if not l_str:
            continue

        # Drop lines containing only dashes/underscores/equals
        if re.fullmatch(r"[-_=\s]+", l_str):
            continue

        # Skip headers
        if any(
            h in l_str
            for h in [
                "Shri Vile Parle Kelavani Mandal",
                "SVKM's",
                "Narsee Monjee College",
                "UPG College",
                "MITHIBAI COLLEGE",
                "AFFILIATED TO UNIVERSITY",
                "NAAC RE-ACCREDITED",
                "NAAC Reaccredited",
                "Module Credits",
                "Consolidated Result",
                "COURSE DETAILS",
                "Candidate's Full Name",
                "MRK G GP C",
                "CG ; CG ;",
            ]
        ):
            continue

        # Skip execution metadata and signatures
        if re.search(
            r"^\s*/\s*:\s*Female|Grade\s+Marks\s+GP|Date:\s*\(Execution|Page:\s*\d+/\d+|CONTROLLER OF EXAMINATIONS|PRINCIPAL",
            l_str,
            re.IGNORECASE,
        ):
            continue

        cleaned_lines.append(l_str)

    return cleaned_lines


def parse_header_line(line):
    """Accurately extracts SNo, SAP, Roll, PRN, Name, Gender, ABC ID, Abeyance, and LD/PwD flag."""
    line_clean = clean_dashes_and_artifacts(line)
    is_abeyance = "Result in abeyance" in line_clean
    if is_abeyance:
        line_clean = line_clean.replace("Result in abeyance", "").strip()

    parts = line_clean.split()
    if len(parts) < 3 or not parts[0].isdigit() or not parts[1].isdigit():
        return None

    s_no = parts[0]
    sap_no = parts[1]
    roll_no = parts[2]

    rem_parts = parts[3:]
    prn_no = ""
    abc_id = ""

    # Check for PRN
    if rem_parts and (
        rem_parts[0].startswith("MU")
        or (rem_parts[0].isdigit() and len(rem_parts[0]) >= 14)
    ):
        prn_no = rem_parts.pop(0)

    # Check for 12-digit ABC ID at the end
    if rem_parts and rem_parts[-1].isdigit() and len(rem_parts[-1]) == 12:
        abc_id = rem_parts.pop()

    full_name = " ".join(rem_parts)
    gender = "Female" if "/" in full_name else "Male"

    # Identify LD / PwD candidates (marked with ~)
    is_pwd = bool(re.search(r"~", full_name))

    return {
        "SNo": s_no,
        "SAP No": sap_no,
        "Roll No": roll_no,
        "PRN No": prn_no,
        "Candidate Full Name": full_name,
        "Gender": gender,
        "ABC ID": abc_id,
        "Is_PwD": is_pwd,
        "Is_Abeyance": is_abeyance,
    }


def split_pipe_row(raw_line, label):
    """Split row on '|' preserving empty cells for practicals alignment."""
    if raw_line is None:
        return []
    body = raw_line.strip()
    if label and body.startswith(label):
        body = body[len(label) :]
    segs = body.split("|")
    if segs and segs[-1].strip() == "":
        segs = segs[:-1]
    return [s.strip() for s in segs]


def robust_parse_total_segment(seg):
    """Parse a single subject's Total-line segment preserving all marks, grades, and $ symbols."""
    seg = (seg or "").strip()
    if not seg:
        return "", "", "", ""

    m_start = re.match(r"^([0-9\+ABNV]+|\-\-)\s+(.*)$", seg)
    if not m_start:
        tokens = seg.split()
        if len(tokens) == 1:
            return tokens[0], "", "", ""
        return "", "", "", ""

    marks = m_start.group(1)
    rest = m_start.group(2).strip()

    # Normalize $ spacing
    rest = re.sub(r"\$([0-9]+)", r"$ \1", rest)
    rest = re.sub(r"([A-Za-z0-9\+\-]+)\s+\$", r"\1$", rest)
    rest = re.sub(r"\bF00\b", "F 0 0", rest)
    rest = re.sub(r"\bF\s+00\b", "F 0 0", rest)
    rest = re.sub(r"\bPS(\d+)\b", r"P$ \1", rest)

    tokens = rest.split()

    grade, gp, creds = "", "", ""
    if len(tokens) == 1:
        grade = tokens[0]
    elif len(tokens) == 2:
        grade = tokens[0]
        if len(tokens[1]) == 2 and tokens[1].isdigit():
            gp, creds = tokens[1][0], tokens[1][1]
        else:
            gp = tokens[1]
    elif len(tokens) == 3:
        grade, gp, creds = tokens[0], tokens[1], tokens[2]
    elif len(tokens) >= 4:
        if tokens[1] == "$":
            grade = tokens[0] + "$"
            gp, creds = tokens[2], tokens[3]
        else:
            grade, gp, creds = tokens[0], tokens[-2], tokens[-1]

    # Clean OCR artifact for Grade '0' -> 'O'
    if grade == "0" and gp:
        grade = "O"
    elif grade == "0$" and gp:
        grade = "O$"

    return marks, grade, gp, creds


def extract_subjects_universal(chunk_lines):
    """Universally extracts Subject Codes, IA Marks, TH Marks, Total details, and Grand Total across all exam formats."""
    subj_codes = []
    ia_scores = []
    th_scores = []
    tot_info = []
    cg_scores = []
    gt_score = ""
    percentage = ""

    # 1. Search for Grand Total and Percentage anywhere in the student block
    for l in chunk_lines:
        gt_match = re.search(r"(\d+/\d+)\s*([\d\.]+%?)?", l)
        if gt_match:
            gt_score = gt_match.group(1)
            percentage = gt_match.group(2) or ""
            break

    # 2. Check for Embedded Subject Codes in IA Line (Re-exam layout)
    ia_with_codes = None
    for l in chunk_lines:
        if l.strip().startswith("IA") and re.search(r"[A-Z]{3,8}\d{2,4}", l):
            ia_with_codes = l
            break

    if ia_with_codes:
        segments = [
            s.strip()
            for s in ia_with_codes.split("|")
            if s.strip() and not s.strip().startswith("IA")
        ]
        for seg in segments:
            if "/" in seg and any(c.isdigit() for c in seg):
                m_gt = re.search(r"(\d+/\d+)\s*([\d\.]+%?)?", seg)
                if m_gt and not gt_score:
                    gt_score = m_gt.group(1)
                    percentage = m_gt.group(2) or ""
                seg = seg.split(gt_score)[0].strip() if gt_score else seg
                if not seg:
                    continue

            m_code_ia = re.search(
                r"([A-Z0-9]{5,12})\s*([0-9\$\+ABNV\s\-]+)?", seg
            )
            if m_code_ia:
                code = m_code_ia.group(1).strip()
                ia_val = m_code_ia.group(2).strip() if m_code_ia.group(2) else ""
                subj_codes.append(code)
                ia_scores.append(ia_val)
    else:
        # Standard layout: Dedicated line for Subject Codes
        for l in chunk_lines:
            codes = re.findall(r"\b[A-Z]{2,8}\d{2,4}[A-Z0-9]?\b", l)
            if len(codes) >= 3:
                subj_codes = codes
                break
        for l in chunk_lines:
            if l.startswith("IA"):
                ia_scores = split_pipe_row(l, "IA")

    # TH and Total line extraction
    total_line_idx = None
    for i, l in enumerate(chunk_lines):
        if l.startswith("TH"):
            th_scores = split_pipe_row(l, "TH")
        elif l.startswith("Total"):
            tot_info = split_pipe_row(l, "Total")
            total_line_idx = i

    # Extract CG row
    if total_line_idx is not None and total_line_idx + 1 < len(chunk_lines):
        candidate = chunk_lines[total_line_idx + 1]
        if candidate.strip() and re.fullmatch(
            r"[\d\-\s\|\.]+", candidate.strip()
        ):
            cg_scores = split_pipe_row(candidate, "")

    return (
        subj_codes,
        ia_scores,
        th_scores,
        tot_info,
        cg_scores,
        gt_score,
        percentage,
    )


def extract_semester_and_final_metrics(rem_str):
    """Dynamically parses semester progression metrics, individual grade value, and clean status."""
    rem_clean = clean_dashes_and_artifacts(rem_str)

    sem_data = {}
    sem_tokens = re.split(r"\[?\bSemester\s+", rem_clean, flags=re.IGNORECASE)

    for token in sem_tokens[1:]:
        raw_body = token.split("]")[0]
        m_num = re.match(r"^([IVXLCDM\d]+)", raw_body, re.I)
        if not m_num:
            continue

        sem_raw = m_num.group(1).upper()
        if (
            sem_raw.endswith("C")
            and len(sem_raw) > 1
            and sem_raw[:-1]
            in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
        ):
            sem_num = sem_raw[:-1]
        else:
            sem_num = sem_raw

        body = raw_body[len(m_num.group(1)) :]

        # Parse Credits
        c_m = re.search(r"\bC\s*[-:\.]\s*(\d+|--)", body)
        if not c_m:
            c_lead = re.match(r"^[-:\.]\s*(\d+)", body)
            c_val = c_lead.group(1) if c_lead else ""
        else:
            c_val = c_m.group(1)

        cg_m = re.search(r"\bCG\s*[-:\.]\s*(\d+|--)", body)
        sgpa_m = re.search(r"\bSGPA\s*[-:\.]\s*([0-9\.]+|AD|--)", body)
        att_m = re.search(r"Attempt[/s]*[-:\.]\s*(\d+)", body, re.I)
        fail_m = re.search(r"\bF\s*\(\s*(\d+)\s*\)", body, re.I)

        sem_key = f"Sem_{sem_num}"
        sem_data[f"{sem_key}_Credits"] = c_val
        sem_data[f"{sem_key}_CG"] = cg_m.group(1) if cg_m else ""
        sem_data[f"{sem_key}_SGPA"] = sgpa_m.group(1) if sgpa_m else ""
        sem_data[f"{sem_key}_Attempts"] = att_m.group(1) if att_m else ""
        sem_data[f"{sem_key}_Fail_Info"] = (
            f"F({fail_m.group(1)})" if fail_m else ""
        )

    # Cumulative metrics
    tot_cred = re.search(r"Total\s+Credit\s*:\s*(\d+|--)?", rem_clean, re.I)
    tot_cg = re.search(r"Total\s+CG\s*:\s*(\d+|--)?", rem_clean, re.I)
    cgpa = re.search(r"\bCGPA\s*:\s*([0-9\.]+|--)?", rem_clean, re.I)

    final_grade_m = re.search(
        r"Final\s+Grade\s*:\s*([OABCDPF][\+\-!]?|--)", rem_clean, re.I
    )
    final_grade_val = (
        final_grade_m.group(1).strip()
        if (final_grade_m and final_grade_m.group(1) != "--")
        else ""
    )

    # Semester Grade: Extract standalone grade (e.g. B+, A, A+) from [Remark: ... Grade-B+]
    sem_grade_m = re.search(
        r"Grade[-:\s]+([OABCDPF][\+\-!]?|--)(?:\s|$|\]|;|,)", rem_clean, re.I
    )
    extracted_grade = (
        sem_grade_m.group(1).strip()
        if (sem_grade_m and sem_grade_m.group(1) != "--")
        else ""
    )

    cum_perc = re.search(
        r"Cum\.?\s*Perc\.?\s*:\s*([0-9\.]+%?|--)?", rem_clean, re.I
    )

    # Result Status (Successful / Unsuccessful)
    remark_m = re.search(r"Remark\s*:\s*([^\];]+)", rem_clean, re.I)
    raw_remark = remark_m.group(1).strip() if remark_m else ""
    status_m = re.search(
        r"\b(Successful[\$]*|Unsuccessful[\$]*)\b", raw_remark, re.I
    )
    if status_m:
        result_status = status_m.group(1).strip()
    elif (
        "Unsuccessful" in rem_clean
        or "Not Eligible" in rem_clean
        or "F(" in rem_clean
    ):
        result_status = "Unsuccessful"
    elif "Successful" in rem_clean or "Successfully" in rem_clean:
        result_status = "Successful"
    else:
        result_status = ""

    # Eligibility / Progression Status
    prog_m = re.search(
        r"(Program\s+Completed\s+Successfully|Eligible\s+for\s+next\s+semester|Not\s+Eligible\s+for\s+next\s+semester|Not\s+Eligible|Successfully\s+Completed)",
        rem_clean,
        re.I,
    )
    eligibility_status = prog_m.group(1).strip() if prog_m else ""

    return {
        "semesters": sem_data,
        "Total Credit": (
            tot_cred.group(1)
            if (tot_cred and tot_cred.group(1) and tot_cred.group(1) != "--")
            else ""
        ),
        "Total CG": (
            tot_cg.group(1)
            if (tot_cg and tot_cg.group(1) and tot_cg.group(1) != "--")
            else ""
        ),
        "CGPA": (
            cgpa.group(1)
            if (cgpa and cgpa.group(1) and cgpa.group(1) != "--")
            else ""
        ),
        "Final Grade": final_grade_val,
        "Cumulative Percentage": (
            cum_perc.group(1)
            if (cum_perc and cum_perc.group(1) and cum_perc.group(1) != "--")
            else ""
        ),
        "Result Status": result_status,
        "Grade": extracted_grade,
        "Eligibility Status": eligibility_status,
    }


def parse_result_pages(
    pdf_pages, course_title_map=None, exam_type="Regular"
):
    """Extract all student records and generate a structured Result_Data dataset."""
    if course_title_map is None:
        course_title_map = {}

    all_lines = []

    for p_idx, page in enumerate(pdf_pages):
        p_no = p_idx + 1
        p_text = page.extract_text() or ""

        # Content-aware page routing
        if "Consolidated Result" not in p_text and not re.search(
            r"^\d{1,4}\s+\d{9,14}", p_text, re.M
        ):
            if len(pdf_pages) > 2 and p_no < 3:
                continue

        cleaned = clean_page_lines(p_text)
        all_lines.extend(cleaned)

    # Group lines by student block
    student_chunks = []
    current_chunk = []

    for line in all_lines:
        if re.match(r"^\d{1,4}\s+\d{9,14}\s+[A-Z0-9]+", line):
            if current_chunk:
                student_chunks.append(current_chunk)
                current_chunk = []
        if current_chunk or re.match(r"^\d{1,4}\s+\d{9,14}\s+[A-Z0-9]+", line):
            current_chunk.append(line)

    if current_chunk:
        student_chunks.append(current_chunk)

    detailed_records = []
    all_discovered_sem_keys = set()

    for chunk in student_chunks:
        if not chunk:
            continue

        header_info = parse_header_line(chunk[0])
        if not header_info:
            continue

        chunk_str = " ".join(chunk)
        is_abeyance = (
            header_info["Is_Abeyance"] or "abeyance" in chunk_str.lower()
        )

        # Handle Result in Abeyance Cases explicitly
        if is_abeyance:
            detailed_row = {
                "SNo": header_info["SNo"],
                "SAP No": header_info["SAP No"],
                "Roll No": header_info["Roll No"],
                "PRN No": header_info["PRN No"],
                "Candidate Full Name": header_info["Candidate Full Name"],
                "Gender": header_info["Gender"],
                "ABC ID": header_info["ABC ID"],
                "Examination Type": exam_type,
                "Is_PwD": "Yes" if header_info["Is_PwD"] else "No",
                "Result Status": "Result in abeyance",
                "Grade": "Result in abeyance",
                "Eligibility Status": "Result in abeyance",
            }
            detailed_records.append(detailed_row)
            continue

        # Extract Universal Subject Details & Marks
        (
            subj_codes,
            ia_scores,
            th_scores,
            tot_info,
            cg_scores,
            gt_score,
            percentage,
        ) = extract_subjects_universal(chunk)

        # Multi-semester & degree evaluation metrics
        rem_str = " ".join(chunk)
        eval_metrics = extract_semester_and_final_metrics(rem_str)
        all_discovered_sem_keys.update(eval_metrics["semesters"].keys())

        # Split Grand Total
        marks_obt, max_marks = "", ""
        if gt_score and "/" in gt_score:
            parts = gt_score.split("/")
            marks_obt = parts[0].strip()
            max_marks = parts[1].strip()

        if gt_score and not percentage:
            try:
                num, den = float(marks_obt), float(max_marks)
                percentage = f"{round((num / den) * 100, 2)}%"
            except Exception:
                percentage = ""

        # Detailed Result Row
        detailed_row = {
            "SNo": header_info["SNo"],
            "SAP No": header_info["SAP No"],
            "Roll No": header_info["Roll No"],
            "PRN No": header_info["PRN No"],
            "Candidate Full Name": header_info["Candidate Full Name"],
            "Gender": header_info["Gender"],
            "ABC ID": header_info["ABC ID"],
            "Examination Type": exam_type,
            "Is_PwD": "Yes" if header_info["Is_PwD"] else "No",
        }

        # Extract per-subject fields
        for idx, code in enumerate(subj_codes):
            ia = clean_text_preserve_symbols(
                ia_scores[idx] if idx < len(ia_scores) else ""
            )
            th = clean_text_preserve_symbols(
                th_scores[idx] if idx < len(th_scores) else ""
            )

            t_entry = tot_info[idx] if idx < len(tot_info) else ""
            marks, grade, gp, creds = robust_parse_total_segment(t_entry)

            course_title = course_title_map.get(code, code)

            # Direct CG extraction from source line
            cg_entry = cg_scores[idx] if idx < len(cg_scores) else ""
            cg_clean = cg_entry.replace("$", "").replace("#", "").strip()
            if cg_clean.isdigit():
                cg_val = int(cg_clean)
            elif (
                gp.replace("$", "").replace("#", "").isdigit()
                and creds.replace("$", "").replace("#", "").isdigit()
            ):
                cg_val = int(gp.replace("$", "").replace("#", "")) * int(
                    creds.replace("$", "").replace("#", "")
                )
            else:
                cg_val = ""

            sub_num = idx + 1
            detailed_row[f"Sub{sub_num} Code"] = code
            detailed_row[f"Sub{sub_num} Course Title"] = course_title
            detailed_row[f"Sub{sub_num} IA Marks"] = ia
            detailed_row[f"Sub{sub_num} TH Marks"] = th
            detailed_row[f"Sub{sub_num} Total Marks"] = marks
            detailed_row[f"Sub{sub_num} Grade"] = grade
            detailed_row[f"Sub{sub_num} GP"] = gp
            detailed_row[f"Sub{sub_num} Credits (C)"] = creds
            detailed_row[f"Sub{sub_num} CG"] = cg_val

        # Store semester metrics
        for sem_col, val in eval_metrics["semesters"].items():
            detailed_row[sem_col] = val

        # Final-Result Section moved to the end
        detailed_row["Total Marks Obt."] = marks_obt
        detailed_row["Max Total Marks"] = max_marks
        detailed_row["Percentage"] = percentage.replace("%", "").strip()
        detailed_row["Total Credit"] = eval_metrics["Total Credit"]
        detailed_row["Total CG"] = eval_metrics["Total CG"]
        detailed_row["CGPA"] = eval_metrics["CGPA"]
        detailed_row["Final Grade"] = eval_metrics["Final Grade"]
        detailed_row["Cumulative Percentage"] = eval_metrics[
            "Cumulative Percentage"
        ]
        detailed_row["Result Status"] = eval_metrics["Result Status"]
        detailed_row["Grade"] = eval_metrics["Grade"]
        detailed_row["Eligibility Status"] = eval_metrics["Eligibility Status"]

        detailed_records.append(detailed_row)

    # Sort discovered semester columns in ascending sequence
    roman_order = [
        "I",
        "II",
        "III",
        "IV",
        "V",
        "VI",
        "VII",
        "VIII",
        "IX",
        "X",
    ]

    def sem_sort_key(col_name):
        m = re.search(r"Sem_([IVXLCDM\d]+)_", col_name)
        if m:
            num = m.group(1).upper()
            return (
                roman_order.index(num)
                if num in roman_order
                else (int(num) if num.isdigit() else 99)
            )
        return 99

    sorted_sem_cols = sorted(all_discovered_sem_keys, key=sem_sort_key)

    df_detailed = pd.DataFrame(detailed_records)

    # Order columns and automatically remove empty columns
    if not df_detailed.empty:
        base_id_cols = [
            c
            for c in [
                "SNo",
                "SAP No",
                "Roll No",
                "PRN No",
                "Candidate Full Name",
                "Gender",
                "ABC ID",
                "Examination Type",
                "Is_PwD",
            ]
            if c in df_detailed.columns
        ]

        sub_cols = [
            c
            for c in df_detailed.columns
            if re.match(r"^Sub\d+", str(c))
        ]

        gt_cols = [
            c
            for c in ["Total Marks Obt.", "Max Total Marks", "Percentage"]
            if c in df_detailed.columns
        ]

        final_result_cols = [
            c
            for c in [
                "Total Credit",
                "Total CG",
                "CGPA",
                "Final Grade",
                "Cumulative Percentage",
                "Result Status",
                "Grade",
                "Eligibility Status",
            ]
            if c in df_detailed.columns
        ]

        ordered_cols = (
            base_id_cols
            + sub_cols
            + [c for c in sorted_sem_cols if c in df_detailed.columns]
            + gt_cols
            + final_result_cols
        )
        all_ordered = ordered_cols + [
            c for c in df_detailed.columns if c not in ordered_cols
        ]
        df_detailed = df_detailed.reindex(columns=all_ordered)

        # Automatically remove any columns that contain no data across all rows
        non_empty_cols = [
            c
            for c in df_detailed.columns
            if not df_detailed[c]
            .fillna("")
            .astype(str)
            .str.strip()
            .eq("")
            .all()
        ]
        df_detailed = df_detailed[non_empty_cols]

    return df_detailed


def save_to_formatted_excel(sheets_dict, exam_type="Regular"):
    """Export formatted Excel workbook with vibrant distinct pastel palette, borders, freezing, and styling."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Hide gridlines in all sheets
            ws.views.sheetView[0].showGridLines = False

            # Borders definitions
            thin_side = Side(style="thin", color="D9D9D9")
            thick_right_side = Side(style="medium", color="1F4E79")
            thin_border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )

            # Vibrant, High-Contrast Distinct Section Header Fills
            header_fill_navy = PatternFill(
                start_color="1B365D", end_color="1B365D", fill_type="solid"
            )  # Classic Navy for Demographics
            header_fill_subject = PatternFill(
                start_color="1A5276", end_color="1A5276", fill_type="solid"
            )  # Deep Ocean Blue for Subjects
            header_fill_semester = PatternFill(
                start_color="5B2C6F", end_color="5B2C6F", fill_type="solid"
            )  # Rich Royal Purple for Semesters
            header_fill_marks = PatternFill(
                start_color="2E4053", end_color="2E4053", fill_type="solid"
            )  # Slate for Totals
            header_fill_final = PatternFill(
                start_color="145A32", end_color="145A32", fill_type="solid"
            )  # Forest Emerald for Final Result
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            # Layered Cell Background Fills
            pwd_row_fill = PatternFill(
                start_color="E8F4F8", end_color="E8F4F8", fill_type="solid"
            )  # Light Ice Blue
            carried_forward_fill = PatternFill(
                start_color="EAECEE", end_color="EAECEE", fill_type="solid"
            )  # Light Slate Grey
            fresh_mark_fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )  # Light Amber/Yellow for fresh marks in Re-exam
            fail_fill = PatternFill(
                start_color="FADBD8", end_color="FADBD8", fill_type="solid"
            )  # Soft Pastel Red
            fail_font = Font(name="Calibri", size=11, color="9C0006")
            unsuccessful_row_font = Font(name="Calibri", size=11, color="C00000")
            normal_font = Font(name="Calibri", size=11, color="000000")

            if sheet_name == "Page_1":
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill_navy
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=1).border = thin_border
                    ws.cell(row=row_idx, column=1).alignment = Alignment(
                        vertical="center", horizontal="left"
                    )
                    ws.cell(row=row_idx, column=2).border = thin_border
                    ws.cell(row=row_idx, column=2).alignment = Alignment(
                        vertical="center", horizontal="left"
                    )

            elif sheet_name == "Page_2":
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill_navy
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                # Center align values in numeric columns & apply full borders
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if col_name in [
                            "Credits",
                            "Internal Max Marks",
                            "Sem End Max Marks",
                            "Total Max Marks",
                            "Course Code",
                        ]:
                            cell.alignment = Alignment(
                                vertical="center", horizontal="center"
                            )
                        else:
                            cell.alignment = Alignment(
                                vertical="center", horizontal="left"
                            )

            elif sheet_name == "Result_Data":
                # Freeze Panes: freeze row 1 and columns up to Candidate Full Name
                name_col_idx = (
                    list(df.columns).index("Candidate Full Name") + 1
                    if "Candidate Full Name" in df.columns
                    else 5
                )
                freeze_col_letter = get_column_letter(name_col_idx + 1)
                ws.freeze_panes = f"{freeze_col_letter}2"

                final_section_cols = [
                    "Total Credit",
                    "Total CG",
                    "CGPA",
                    "Final Grade",
                    "Cumulative Percentage",
                    "Result Status",
                    "Grade",
                    "Eligibility Status",
                ]

                # Identify subject block boundaries
                subject_end_col_indices = [
                    idx + 1
                    for idx, c in enumerate(df.columns)
                    if str(c).endswith(" CG")
                ]

                # Identify semester block boundaries (bold vertical separator between semester blocks)
                sem_end_col_indices = []
                current_sem = None
                last_idx = None
                for idx, c in enumerate(df.columns, start=1):
                    m = re.match(r"^Sem_([IVXLCDM\d]+)_", str(c))
                    if m:
                        s_name = m.group(1)
                        if current_sem and s_name != current_sem:
                            sem_end_col_indices.append(last_idx)
                        current_sem = s_name
                        last_idx = idx
                if last_idx and last_idx not in sem_end_col_indices:
                    sem_end_col_indices.append(last_idx)

                # Move separator to appear before Total Credit column
                pre_total_credit_col_indices = []
                if "Total Credit" in df.columns:
                    tc_idx = list(df.columns).index("Total Credit")
                    if tc_idx > 0:
                        pre_total_credit_col_indices.append(tc_idx)

                all_thick_right_col_indices = set(
                    subject_end_col_indices
                    + sem_end_col_indices
                    + pre_total_credit_col_indices
                )

                # Header Formatting with vibrant distinct section fills
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font

                    if str(col_name).startswith("Sem_"):
                        cell.fill = header_fill_semester
                    elif col_name in final_section_cols:
                        cell.fill = header_fill_final
                    elif col_name in [
                        "Total Marks Obt.",
                        "Max Total Marks",
                        "Percentage",
                    ]:
                        cell.fill = header_fill_marks
                    elif str(col_name).startswith("Sub"):
                        cell.fill = header_fill_subject
                    else:
                        cell.fill = header_fill_navy

                    r_side = (
                        thick_right_side
                        if col_idx in all_thick_right_col_indices
                        else thin_side
                    )
                    cell.border = Border(
                        left=thin_side,
                        right=r_side,
                        top=thin_side,
                        bottom=thin_side,
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                # Subject block ranges for cell evaluation
                grade_col_indices = [
                    idx + 1
                    for idx, c in enumerate(df.columns)
                    if str(c).endswith(" Grade")
                    and not str(c).startswith("Final")
                    and not str(c).startswith("Overall")
                    and str(c) != "Grade"
                ]

                sub_blocks = []
                for g_col in grade_col_indices:
                    start_col = g_col - 5
                    end_col = g_col + 3
                    sub_blocks.append((g_col, start_col, end_col))

                is_pwd_col_idx = (
                    list(df.columns).index("Is_PwD") + 1
                    if "Is_PwD" in df.columns
                    else None
                )

                # Row-by-row data formatting
                for row_idx in range(2, ws.max_row + 1):
                    status_val = str(
                        ws.cell(
                            row=row_idx,
                            column=list(df.columns).index("Result Status") + 1,
                        ).value
                        or ""
                        if "Result Status" in df.columns
                        else ""
                    )
                    grade_val = str(
                        ws.cell(
                            row=row_idx,
                            column=list(df.columns).index("Grade") + 1,
                        ).value
                        or ""
                        if "Grade" in df.columns
                        else ""
                    )
                    final_grade_val = str(
                        ws.cell(
                            row=row_idx,
                            column=list(df.columns).index("Final Grade") + 1,
                        ).value
                        or ""
                        if "Final Grade" in df.columns
                        else ""
                    )
                    eligibility_val = str(
                        ws.cell(
                            row=row_idx,
                            column=list(df.columns).index("Eligibility Status")
                            + 1,
                        ).value
                        or ""
                        if "Eligibility Status" in df.columns
                        else ""
                    )

                    is_unsuccessful_row = (
                        "Unsuccessful" in status_val
                        or "Not Eligible" in eligibility_val
                        or grade_val.strip() == "F"
                        or final_grade_val.strip() == "F"
                    )

                    is_pwd_row = False
                    if is_pwd_col_idx:
                        is_pwd_row = (
                            str(
                                ws.cell(
                                    row=row_idx, column=is_pwd_col_idx
                                ).value
                                or ""
                            ).strip()
                            == "Yes"
                        )

                    failed_cols_in_row = set()
                    carried_cols_in_row = set()
                    fresh_mark_cols_in_row = set()

                    for g_col, start_col, end_col in sub_blocks:
                        g_val = str(
                            ws.cell(row=row_idx, column=g_col).value or ""
                        ).strip()
                        if "F" in g_val:
                            for c in range(start_col, end_col + 1):
                                failed_cols_in_row.add(c)

                        for c in range(start_col, end_col + 1):
                            c_name = df.columns[c - 1]
                            c_val = str(
                                ws.cell(row=row_idx, column=c).value or ""
                            ).strip()

                            if "$" in c_val:
                                carried_cols_in_row.add(c)
                            elif (
                                exam_type == "Re-examination"
                                and (
                                    "IA Marks" in c_name or "TH Marks" in c_name
                                )
                                and c_val
                                and c_val not in ["--", "NV", "AB", "NaN", "nan"]
                            ):
                                fresh_mark_cols_in_row.add(c)

                    # Apply Cell Borders, Alignments, Fonts, and Highlights
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        r_side = (
                            thick_right_side
                            if col_idx in all_thick_right_col_indices
                            else thin_side
                        )
                        cell.border = Border(
                            left=thin_side,
                            right=r_side,
                            top=thin_side,
                            bottom=thin_side,
                        )

                        col_name = df.columns[col_idx - 1]
                        if col_name == "Candidate Full Name":
                            cell.alignment = Alignment(
                                vertical="center", horizontal="left"
                            )
                        elif "Course Title" in col_name:
                            cell.alignment = Alignment(
                                vertical="center", horizontal="left"
                            )
                        else:
                            cell.alignment = Alignment(
                                vertical="center", horizontal="center"
                            )

                        # Color-based separation without forced bold text
                        if col_idx in failed_cols_in_row:
                            cell.font = fail_font
                        elif is_unsuccessful_row:
                            cell.font = unsuccessful_row_font
                        else:
                            cell.font = normal_font

                        # Fill priorities
                        if col_idx in failed_cols_in_row:
                            cell.fill = fail_fill
                        elif col_idx in fresh_mark_cols_in_row:
                            cell.fill = fresh_mark_fill
                        elif col_idx in carried_cols_in_row:
                            cell.fill = carried_forward_fill
                        elif is_pwd_row:
                            cell.fill = pwd_row_fill

            # Auto-fit columns based strictly on actual content width
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                ws.column_dimensions[col_letter].width = max(max_len + 2, 6)

    output.seek(0)
    return output


def render_detailed_table_html(df, exam_type="Regular"):
    """Renders interactive HTML preview in Streamlit with visual cues and distinct section colors."""
    if df.empty:
        st.info("No result records found.")
        return

    cols = list(df.columns)
    final_section_cols = [
        "Total Credit",
        "Total CG",
        "CGPA",
        "Final Grade",
        "Cumulative Percentage",
        "Result Status",
        "Grade",
        "Eligibility Status",
    ]

    # Block endings
    subject_end_cols = [c for c in cols if str(c).endswith(" CG")]
    sem_end_cols = []
    current_sem = None
    last_col = None
    for c in cols:
        m = re.match(r"^Sem_([IVXLCDM\d]+)_", str(c))
        if m:
            s_name = m.group(1)
            if current_sem and s_name != current_sem:
                sem_end_cols.append(last_col)
            current_sem = s_name
            last_col = c
    if last_col and last_col not in sem_end_cols:
        sem_end_cols.append(last_col)

    pre_tc_cols = []
    if "Total Credit" in cols:
        tc_i = cols.index("Total Credit")
        if tc_i > 0:
            pre_tc_cols.append(cols[tc_i - 1])

    all_thick_cols = set(subject_end_cols + sem_end_cols + pre_tc_cols)

    grade_cols = [
        c
        for c in cols
        if str(c).endswith(" Grade")
        and not str(c).startswith("Final")
        and not str(c).startswith("Overall")
        and str(c) != "Grade"
    ]

    sub_ranges = []
    for g in grade_cols:
        g_idx = cols.index(g)
        s_code_col = cols[g_idx - 5]
        e_cg_col = cols[g_idx + 3]
        sub_ranges.append((g, s_code_col, e_cg_col, g_idx - 5, g_idx + 3))

    html = """
    <style>
        .detailed-table-container {
            width: 100%;
            overflow-x: auto;
            max-height: 550px;
            border: 1px solid #d3d3d3;
            border-radius: 6px;
            margin-top: 10px;
        }
        .detailed-table {
            border-collapse: collapse;
            font-family: Calibri, sans-serif;
            font-size: 13px;
            width: 100%;
            white-space: nowrap;
        }
        .detailed-table th {
            color: white;
            padding: 8px 10px;
            text-align: center;
            position: sticky;
            top: 0;
            z-index: 2;
            border: 1px solid #d9d9d9;
            font-weight: bold;
        }
        .th-navy { background-color: #1B365D; }
        .th-subject { background-color: #1A5276; }
        .th-semester { background-color: #5B2C6F; }
        .th-marks { background-color: #2E4053; }
        .th-final { background-color: #145A32; }

        .detailed-table td {
            padding: 6px 10px;
            text-align: center;
            border: 1px solid #e0e0e0;
            color: #000000;
        }
        .text-left-cell { text-align: left !important; }
        .unsuccessful-row-text { color: #C00000 !important; }

        .thick-right {
            border-right: 3px solid #1F4E79 !important;
        }
        .pwd-row {
            background-color: #E8F4F8 !important;
        }
        .carried-forward-cell {
            background-color: #EAECEE !important;
        }
        .fresh-mark-cell {
            background-color: #FFF3CD !important;
        }
        .fail-cell {
            background-color: #FADBD8 !important;
            color: #9C0006 !important;
        }
    </style>
    <div class="detailed-table-container">
    <table class="detailed-table">
        <thead>
            <tr>
    """
    for col in cols:
        th_cls = ["thick-right"] if col in all_thick_cols else []
        if str(col).startswith("Sem_"):
            th_cls.append("th-semester")
        elif col in final_section_cols:
            th_cls.append("th-final")
        elif col in ["Total Marks Obt.", "Max Total Marks", "Percentage"]:
            th_cls.append("th-marks")
        elif str(col).startswith("Sub"):
            th_cls.append("th-subject")
        else:
            th_cls.append("th-navy")

        html += f'<th class="{" ".join(th_cls)}">{col}</th>'
    html += "</tr></thead><tbody>"

    for _, row in df.iterrows():
        is_pwd = str(row.get("Is_PwD", "")).strip() == "Yes"
        status_val = str(row.get("Result Status", ""))
        grade_val = str(row.get("Grade", ""))
        final_grade_val = str(row.get("Final Grade", ""))
        eligibility_val = str(row.get("Eligibility Status", ""))

        is_unsuccessful_row = (
            "Unsuccessful" in status_val
            or "Not Eligible" in eligibility_val
            or grade_val.strip() == "F"
            or final_grade_val.strip() == "F"
        )

        failed_cols = set()
        carried_cols = set()
        fresh_cols = set()

        for g_col, _, _, start_i, end_i in sub_ranges:
            if "F" in str(row.get(g_col, "")).strip():
                for col_name in cols[start_i : end_i + 1]:
                    failed_cols.add(col_name)

            for col_name in cols[start_i : end_i + 1]:
                c_val = str(row.get(col_name, "")).strip()
                if "$" in c_val:
                    carried_cols.add(col_name)
                elif (
                    exam_type == "Re-examination"
                    and ("IA Marks" in col_name or "TH Marks" in col_name)
                    and c_val
                    and c_val not in ["--", "NV", "AB", "NaN", "nan"]
                ):
                    fresh_cols.add(col_name)

        row_classes = []
        if is_pwd:
            row_classes.append("pwd-row")
        if is_unsuccessful_row:
            row_classes.append("unsuccessful-row-text")
        row_cls_str = (
            f' class="{" ".join(row_classes)}"' if row_classes else ""
        )

        html += f"<tr{row_cls_str}>"
        for col in cols:
            val = row[col] if pd.notna(row[col]) else ""
            classes = []
            if col in all_thick_cols:
                classes.append("thick-right")
            if col == "Candidate Full Name" or "Course Title" in col:
                classes.append("text-left-cell")

            if col in failed_cols:
                classes.append("fail-cell")
            elif col in fresh_cols:
                classes.append("fresh-mark-cell")
            elif col in carried_cols:
                classes.append("carried-forward-cell")

            cls_str = f' class="{" ".join(classes)}"' if classes else ""
            html += f"<td{cls_str}>{val}</td>"
        html += "</tr>"

    html += "</tbody></table></div>"
    st.markdown(html, unsafe_allow_html=True)


# ==============================================================================
# MAIN ENTRY POINTS
# ==============================================================================


def show():
    """Main view function invoked by app.py router."""
    st.title("📑 Result Gazette PDF to Excel")
    st.caption(
        "Dynamic Examination Result Gazette Parser for all Undergraduate & Postgraduate Programmes."
    )

    uploaded_files = st.file_uploader(
        "Upload Result Gazette PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        for uploaded_file in uploaded_files:
            st.subheader(f"📄 Processing: {uploaded_file.name}")

            with pdfplumber.open(uploaded_file) as pdf:
                total_pages = len(pdf.pages)
                st.info(f"Total Pages Detected: **{total_pages}**")

                # 1. Page 1 Metadata & Exam Type Detection
                p1_text = pdf.pages[0].extract_text() or ""
                df_p1, exam_type = parse_page_1(p1_text)

                # 2. Page 2 Course Scheme
                p2_page = pdf.pages[1] if total_pages > 1 else pdf.pages[0]
                df_p2 = parse_page_2(p2_page)

                # Map Course Code -> Course Title
                course_map = {}
                if (
                    not df_p2.empty
                    and "Course Code" in df_p2.columns
                    and "Course Title" in df_p2.columns
                ):
                    course_map = dict(
                        zip(df_p2["Course Code"], df_p2["Course Title"])
                    )

                # 3. Dynamic Results Parsing
                df_result_data = parse_result_pages(
                    pdf.pages,
                    course_title_map=course_map,
                    exam_type=exam_type,
                )

            # 4. Preview Tabs (Page_1, Page_2, Result_Data)
            t1, t2, t3 = st.tabs(
                [
                    "📋 Metadata (Page 1)",
                    "📚 Courses (Page 2)",
                    "📊 Result Data",
                ]
            )

            with t1:
                st.dataframe(df_p1, use_container_width=True)
            with t2:
                st.dataframe(df_p2, use_container_width=True)
            with t3:
                st.markdown("##### 📌 Consolidated & Expanded Student Results")
                render_detailed_table_html(df_result_data, exam_type=exam_type)

            # 5. Save & Download Excel (Consolidated Clean Sheets)
            sheets_data = {
                "Page_1": df_p1,
                "Page_2": df_p2,
                "Result_Data": df_result_data,
            }

            excel_bytes = save_to_formatted_excel(
                sheets_data, exam_type=exam_type
            )

            export_file_name = f"{uploaded_file.name.rsplit('.', 1)[0]}_Export.xlsx"
            st.download_button(
                label=f"⬇️ Download Structured Excel ({export_file_name})",
                data=excel_bytes,
                file_name=export_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.divider()


def main():
    """Alias for show() to support direct script execution or alt import."""
    show()


if __name__ == "__main__":
    st.set_page_config(
        page_title="College Result Ledger Extractor",
        page_icon="📑",
        layout="wide",
    )
    show()