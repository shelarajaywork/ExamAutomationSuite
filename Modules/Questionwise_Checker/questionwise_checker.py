"""
==================================================
QUESTIONWISE CHECKER (MULTI-FILE & GMR INTEGRATED)
==================================================

PURPOSE
--------------------------------------------------
This Streamlit tool helps exam staff:

1. Open password-protected or unprotected Excel mark sheets exported from
   the examination system (supports multiple uploads simultaneously).
2. Optionally upload a GMR Excel file to match records by PRN/Student Number 
   and Subject Code/Modulecode to fetch 'Internal Marks' (Internal Total Obtained),
   'Total Max Marks' (from Composite AGR type), 'Semester Total Max Marks' (from 
   Semester Total AGR type), 'Semester Marks' (Grades value from "Semester Total 
   Marks" AGR type rows - shown only on Moderator rows, or on an Examiner row 
   with no corresponding Moderator record), calculate 'Total Marks Obtained', 
   and derive 'Grade' dynamically based on percentage criteria.
3. Clean the raw export (remove blank/footer rows, drop unused columns).
4. Sort records so each student's rows appear in a fixed order:
   Examiner -> Moderator -> Reval 1 -> Reval 2.
5. Calculate 'Change %' for Moderator rows:
   Change % = (Moderator TotalObtained - Examiner TotalObtained) / Semester Total Max Marks * 100.
6. Check paper eligibility for 100% moderation if >10% mark change occurs 
   in >50% of moderated answer books.
7. Display 5 summary cards (Total Students, Positive Changes, Negative Changes, 
   No Changes, and 10% Change Moderation Metrics) and three preview tabs.
8. Export processed output Excel files:
   - Original Cleaned File: Reval rows removed, Moderator filtered, legacy format.
   - Formatted File: All UserTypes, GMR columns, Change %, center alignment, F2 freeze, 
     thick student borders, green/red text, and automatic Excel pop-up message if eligible.
   - Consolidated Multi-Sheet Excel option for multiple files with pop-up messages per eligible sheet.
==================================================
"""

from datetime import datetime
import io
import os
import re
import tempfile
import zipfile

import msoffcrypto
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st


# ==================================================
# CONFIGURATION
# ==================================================

# Candidate passwords tried (in order) against any password-protected upload.
# decrypt_and_load() stops at the first one that successfully opens the file.
PASSWORDS = [
    "Exam@108",
    "UPG@123",
    "Exam@105"
]

# Canonical set of "known" (non-question) columns that always appear in a
# processed sheet. Anything NOT in this list is treated as a question-mark
# column (see calculate_paper_max_marks / question_columns usage below).
FIXED_COLUMNS = [
    "SrNo",
    "ExamAssingment_ID",
    "Subject Name-Subject code",
    "OnScreenID",
    "SubjectName",
    "Semester",
    "RollNo",
    "PRNNumber",
    "Name",
    "CampusName",
    "UserType",
    "MarkAttendance",
    "Email",
    "Semester Total Max Marks",
    "Semester Marks",
    "Internal Marks",
    "Total Max Marks",
    "Total Marks Obtained",
    "Grade",
    "Change %",
]

# Columns to drop from the original standard clean downloaded Excel files.
# This is the "legacy" export: no GMR-derived columns, no Change %.
COLUMNS_TO_DROP_STANDARD = [
    "Semester",
    "ExamAssingment_ID",
    "OnScreenID",
    "RollNo",
    "Name",
    "UserType",
    "CampusName",
    "MarkAttendance",
    "Email",
    "Grade",
    "Internal Marks",
    "Total Max Marks",
    "Semester Total Max Marks",
    "Semester Marks",
    "Total Marks Obtained",
    "Change %",
]

# Columns to drop from the formatted downloaded Excel files.
# This export keeps UserType, GMR columns, Grade, Change %, etc. - it only
# drops the columns nobody needs to see (identifiers/admin metadata).
COLUMNS_TO_DROP_FORMATTED = [
    "Semester",
    "ExamAssingment_ID",
    "OnScreenID",
    "RollNo",
    "Name",
    "CampusName",
    "MarkAttendance",
]

# Columns hidden from the on-screen Streamlit preview tables (kept in the
# underlying DataFrame / exported Excel files - this only affects display).
PREVIEW_HIDDEN_COLUMNS = [
    "OnScreenID",
    "RollNo",
    "Name",
    "CampusName",
    "Subject Name-Subject code",
    "SubjectName",
    "Semester",
]

# Defines the fixed row order within each student's record block.
# Lower number = appears first. Anything not listed here (shouldn't happen)
# falls back to 99 via .fillna(99) in clean_and_sort().
USER_TYPE_ORDER = {
    "Examiner": 1,
    "Moderator": 2,
    "Reval 1": 3,
    "Reval 2": 4,
}

# Row background colours (CSS strings) used only in the Streamlit preview
# tables (see highlight_rows/style_preview) to visually distinguish
# Moderator/Reval rows from the Examiner row above them.
ROW_COLOURS = {
    "Moderator": "background-color:#143d22;color:#ffffff;",  # Explicit white text for high contrast in light & dark mode
    "Reval 1": "background-color:#162456;color:white;",
    "Reval 2": "background-color:#290245;color:white;",
}

INCREASE_STYLE = "color:#00E676;font-weight:bold;"  # Bright green: mark went up / grade improved
DECREASE_STYLE = "color:#FF4500;font-weight:bold;"  # Red: mark went down / grade degraded

# Ordinal rank of each grade letter, used purely for comparison (e.g. did a
# reviewer's Grade rank higher or lower than the Examiner's Grade).
GRADE_RANKS = {
    "O": 8,
    "A+": 7,
    "A": 6,
    "B+": 5,
    "B": 4,
    "C": 3,
    "P": 2,
    "F": 1
}


# ==================================================
# STEP 1: DECRYPT + LOAD UPLOADED FILES
# ==================================================
def decrypt_and_load(uploaded_file):
    """
    Attempts to load unprotected Excel files directly first.
    If protected, iterates through PASSWORDS list to decrypt.

    Returns:
        (DataFrame, label) on success, where `label` is either
        "Unprotected (Direct)" or the password that worked.
        (None, None) if the file could not be opened by any means.
    """
    # --- Attempt 1: file has no password at all ------------------------
    try:
        uploaded_file.seek(0)  # Streamlit's UploadedFile is a stream; rewind before every read attempt.
        df = pd.read_excel(uploaded_file, dtype=object, keep_default_na=False)
        return df, "Unprotected (Direct)"
    except Exception:
        # Not readable as a plain workbook - most likely it's encrypted. Fall through to password attempts.
        pass

    # --- Attempt 2: try each known password in turn ---------------------
    for password in PASSWORDS:
        decrypted_path = None
        try:
            uploaded_file.seek(0)

            office_file = msoffcrypto.OfficeFile(uploaded_file)
            office_file.load_key(password=password)  # Raises if the password is wrong.

            # msoffcrypto needs to write the decrypted bytes somewhere before pandas can read them,
            # so we decrypt into a throwaway temp file and clean it up in `finally`.
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                office_file.decrypt(tmp)
                decrypted_path = tmp.name

            df = pd.read_excel(decrypted_path, dtype=object, keep_default_na=False)
            return df, password

        except Exception:
            # Wrong password (or some other decrypt/read failure) - try the next candidate.
            continue

        finally:
            # Always scrub the temp file, whether decryption succeeded or failed.
            if decrypted_path and os.path.exists(decrypted_path):
                os.remove(decrypted_path)

    # Exhausted every option - unprotected read failed and no password matched.
    return None, None


def load_gmr_file(gmr_file):
    """Loads GMR file directly or decrypts if password protected."""
    if gmr_file is None:
        return None
    df, _ = decrypt_and_load(gmr_file)  # Reuse the same open/decrypt logic; we don't care which password worked.
    return df


# ==================================================
# STEP 2: CLEAN, SORT, & GMR LOOKUP MATCHING
# ==================================================
def clean_and_sort(df):
    """Clean up raw Questionwise export and sort rows."""
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()  # Guard against stray whitespace in exported headers.

    # Drop rows with a blank PRNNumber - these are usually spacer/footer rows in the raw export.
    if "PRNNumber" in df.columns:
        df = df[df["PRNNumber"].astype(str).str.strip() != ""]

    df = df.dropna(how="all")  # Drop fully-empty rows (e.g. trailing blank lines in the sheet).

    # The export appends a "Count: N" footer row - strip it out using whichever
    # column is available to check ("SrNo" preferred, else the first column).
    if "SrNo" in df.columns:
        df = df[~df["SrNo"].astype(str).str.contains("Count:", case=False, na=False)]
    else:
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains("Count:", case=False, na=False)]

    # PRNNumber should be numeric - coerce and drop anything that isn't
    # (defensive cleanup in case a stray text row slipped through above).
    if "PRNNumber" in df.columns:
        df["PRNNumber"] = pd.to_numeric(df["PRNNumber"], errors="coerce")
        df = df[df["PRNNumber"].notna()]

    # Sort so each student's rows appear in a fixed order: Examiner -> Moderator -> Reval 1 -> Reval 2.
    if "PRNNumber" in df.columns and "UserType" in df.columns:
        df["SortOrder"] = df["UserType"].map(USER_TYPE_ORDER).fillna(99)
        df = df.sort_values(by=["PRNNumber", "SortOrder"]).drop(columns=["SortOrder"])

    df = df.reset_index(drop=True)
    return df


def calculate_grade(total_obtained, total_max):
    """Calculates Grade based on percentage of Total Obtained vs Total Max Marks."""
    try:
        obtained = float(total_obtained)
        max_m = float(total_max)

        if max_m <= 0:
            return ""  # Avoid division by zero / nonsensical percentages.

        pct = (obtained / max_m) * 100.0

        # Standard percentage-to-grade ladder, checked from highest to lowest.
        if pct >= 90:
            return "O"
        elif pct >= 80:
            return "A+"
        elif pct >= 70:
            return "A"
        elif pct >= 60:
            return "B+"
        elif pct >= 55:
            return "B"
        elif pct >= 50:
            return "C"
        elif pct >= 40:
            return "P"
        else:
            return "F"
    except (ValueError, TypeError):
        # total_obtained/total_max were blank, "AB", "NA", etc. - no grade can be derived.
        return ""


def apply_gmr_lookup(df, gmr_df):
    """
    Performs lookup matching between Questionwise and GMR data:
    - Internal Marks: from rows where AGR type contains "Internal Total"
    - Total Max Marks: MAX Marks from rows where AGR type contains "Composite "
    - Semester Total Max Marks: MAX Marks from rows where AGR type contains "Semester Total"
    - Total Marks Obtained: TotalObtainedScore + Internal Marks
    - Grade: Calculated based on percentage criteria
    """
    df = df.copy()

    # Always create these columns (even with no GMR file) so downstream code
    # (reorder_gmr_columns, exports, previews) can assume they exist.
    df["Internal Marks"] = ""
    df["Total Max Marks"] = ""
    df["Semester Total Max Marks"] = ""
    df["Semester Marks"] = ""
    df["Total Marks Obtained"] = ""
    df["Grade"] = ""

    if gmr_df is None or gmr_df.empty:
        return df  # No GMR file uploaded - leave the GMR columns blank.

    gmr_clean = gmr_df.copy()
    gmr_clean.columns = gmr_clean.columns.astype(str).str.strip()

    # Bail out quietly if the GMR file doesn't have the columns we need to match on.
    required_gmr_cols = ["Student Number", "Modulecode", "AGR type", "Grades"]
    if not all(col in gmr_clean.columns for col in required_gmr_cols):
        return df

    # --- Build a composite lookup key (PRN + Subject code) on both sides ---
    # so a single row in the Questionwise sheet can be matched to the right
    # GMR record even though GMR uses different column names for the same IDs.
    df_key = (
        df["PRNNumber"].astype(str).str.strip() +
        df["Subject Name-Subject code"].astype(str).str.strip()
    )

    gmr_clean["LookupKey"] = (
        gmr_clean["Student Number"].astype(str).str.strip() +
        gmr_clean["Modulecode"].astype(str).str.strip()
    )

    # 1. Fetch Internal Marks (GMR rows whose AGR type mentions "Internal Total").
    internal_df = gmr_clean[
        gmr_clean["AGR type"].astype(str).str.contains("Internal Total", case=False, na=False)
    ]
    internal_obtained_map = dict(zip(internal_df["LookupKey"], internal_df["Grades"]))
    df["Internal Marks"] = df_key.map(internal_obtained_map).fillna("")

    # 2. Fetch Total Max Marks from Composite AGR type.
    composite_max_map = {}
    if "MAX Marks" in gmr_clean.columns:
        composite_df = gmr_clean[
            gmr_clean["AGR type"].astype(str).str.contains("Composite ", case=False, na=False)
        ]
        composite_max_map = dict(zip(composite_df["LookupKey"], composite_df["MAX Marks"]))
    df["Total Max Marks"] = df_key.map(composite_max_map).fillna("")

    # 3. Fetch Semester Total Max Marks from Semester Total AGR type.
    sem_total_max_map = {}
    if "MAX Marks" in gmr_clean.columns:
        sem_total_df = gmr_clean[
            gmr_clean["AGR type"].astype(str).str.contains("Semester Total", case=False, na=False)
        ]
        sem_total_max_map = dict(zip(sem_total_df["LookupKey"], sem_total_df["MAX Marks"]))
    df["Semester Total Max Marks"] = df_key.map(sem_total_max_map).fillna("")

    # 3b. Fetch Semester Marks (Grades column) from GMR rows whose AGR type
    # contains "Semester Total Marks" (e.g. "Semester Total Marks(60 )"),
    # using the same PRN + Subject code matching key as the other lookups above.
    # The looked-up value is only ever KEPT on:
    #   (a) every Moderator row, and
    #   (b) an Examiner row whose student+module combination has NO
    #       corresponding Moderator row.
    # All other rows (Examiner rows that DO have a Moderator row, plus any
    # Reval 1 / Reval 2 rows) are left blank.
    sem_marks_df = gmr_clean[
        gmr_clean["AGR type"].astype(str).str.contains("Semester Total Marks", case=False, na=False)
    ]
    sem_marks_map = dict(zip(sem_marks_df["LookupKey"], sem_marks_df["Grades"]))
    raw_semester_marks = df_key.map(sem_marks_map).fillna("")

    if "UserType" in df.columns and "PRNNumber" in df.columns and "Subject Name-Subject code" in df.columns:
        module_key = (
            df["PRNNumber"].astype(str).str.strip() +
            "||" +
            df["Subject Name-Subject code"].astype(str).str.strip()
        )
        moderator_module_keys = set(
            module_key[df["UserType"].astype(str).str.strip() == "Moderator"]
        )

        def keep_semester_marks(idx):
            u_type = str(df.loc[idx, "UserType"]).strip()
            if u_type == "Moderator":
                return raw_semester_marks.loc[idx]
            if u_type == "Examiner" and module_key.loc[idx] not in moderator_module_keys:
                return raw_semester_marks.loc[idx]
            return ""

        df["Semester Marks"] = [keep_semester_marks(i) for i in df.index]

    # 4. Calculate Total Marks Obtained = TotalObtainedScore (exam script) + Internal Marks (GMR).
    def compute_total_obtained(row):
        tot_score = row.get("TotalObtainedScore", "")
        int_marks = row.get("Internal Marks", "")

        try:
            val_tot = float(str(tot_score).strip())
        except (ValueError, TypeError):
            val_tot = 0.0

        try:
            # Clean leading '+' and whitespace before parsing Internal Marks into a float
            cleaned_int_marks = re.sub(r"^\+\s*", "", str(int_marks).strip())
            val_int = float(cleaned_int_marks)
        except (ValueError, TypeError):
            val_int = 0.0

        # Only report a blank total if BOTH inputs were blank - otherwise treat
        # a missing side as 0 (e.g. a student with Internal Marks but no script yet).
        if str(tot_score).strip() == "" and str(int_marks).strip() == "":
            return ""

        return val_tot + val_int

    df["Total Marks Obtained"] = df.apply(compute_total_obtained, axis=1)

    # 5. Calculate Grade from the newly computed Total Marks Obtained / Total Max Marks.
    df["Grade"] = df.apply(
        lambda r: calculate_grade(r["Total Marks Obtained"], r["Total Max Marks"]), axis=1
    )

    return df


def calculate_paper_max_marks(df):
    """Extracts sum of question max marks from question column headers."""
    # Question columns are simply "everything that isn't a known fixed column".
    question_columns = [col for col in df.columns if col not in FIXED_COLUMNS]
    total_q_max = 0.0
    for col in question_columns:
        # Column headers look like "Q1 (10)" - pull the number out of the parentheses.
        match = re.search(r"\(([\d.]+)\)", col)
        if match:
            try:
                total_q_max += float(match.group(1))
            except ValueError:
                pass
    return total_q_max


def calculate_moderator_change_pct(df):
    """
    Calculates Change % for Moderator rows using Semester Total Max Marks:
    Change % = (TotalObtainedScore (Moderator) - TotalObtainedScore (Examiner)) / Semester Total Max Marks * 100
    """
    df = df.copy()
    df["Change %"] = pd.Series([""] * len(df), index=df.index, dtype="object")

    if "PRNNumber" not in df.columns or "UserType" not in df.columns or "TotalObtainedScore" not in df.columns:
        return df  # Can't compute anything meaningful without these three columns.

    # Precompute once - used as the last-resort denominator fallback below.
    paper_q_max = calculate_paper_max_marks(df)

    for prn, group in df.groupby("PRNNumber", dropna=False):
        examiner_rows = group[group["UserType"].astype(str).str.strip() == "Examiner"]
        moderator_rows = group[group["UserType"].astype(str).str.strip() == "Moderator"]

        if examiner_rows.empty or moderator_rows.empty:
            continue  # A student without both an Examiner AND a Moderator row has no "change" to report.

        examiner_row = examiner_rows.iloc[0]
        moderator_row = moderator_rows.iloc[0]

        try:
            ex_score = float(examiner_row["TotalObtainedScore"])
            mod_score = float(moderator_row["TotalObtainedScore"])

            # Denominator selection, in priority order:
            # 1. Try Semester Total Max Marks from GMR lookup (most accurate, whole-semester scale).
            sem_max_marks = 0.0
            sem_max_val = str(moderator_row.get("Semester Total Max Marks", "")).strip()
            if sem_max_val != "":
                try:
                    sem_max_marks = float(sem_max_val)
                except ValueError:
                    sem_max_marks = 0.0

            # 2. Fallback to Total Max Marks if GMR Semester Total is not available.
            if sem_max_marks <= 0:
                tot_max_val = str(moderator_row.get("Total Max Marks", "")).strip()
                if tot_max_val != "":
                    try:
                        sem_max_marks = float(tot_max_val)
                    except ValueError:
                        sem_max_marks = 0.0

            # 3. Fallback to sum of question max marks from column headers (no GMR file at all).
            if sem_max_marks <= 0:
                sem_max_marks = paper_q_max

            if sem_max_marks > 0:
                change_pct = ((mod_score - ex_score) / sem_max_marks) * 100.0
                # Change % is only ever recorded on the Moderator row itself.
                df.loc[moderator_rows.index[0], "Change %"] = round(change_pct, 2)
        except (ValueError, TypeError):
            # Non-numeric scores (e.g. "AB"/"UFM") - leave Change % blank for this student.
            continue

    return df


def reorder_gmr_columns(df):
    """
    Reorders GMR and calculated columns right after TotalObtainedScore in the exact sequence:
    Semester Total Max Marks, Semester Marks, Internal Marks, Total Max Marks, Total Marks Obtained, Grade, Change %, Email
    """
    df = df.copy()
    target_tail = [
        "Semester Total Max Marks",
        "Semester Marks",
        "Internal Marks",
        "Total Max Marks",
        "Total Marks Obtained",
        "Grade",
        "Change %",
        "Email",
    ]
    if "TotalObtainedScore" in df.columns:
        cols = df.columns.tolist()
        tot_idx = cols.index("TotalObtainedScore")
        head = cols[:tot_idx + 1]                              # Everything up to and including TotalObtainedScore.
        tail = [c for c in target_tail if c in cols]            # The GMR/derived columns, in their desired order.
        remaining = [c for c in cols if c not in head and c not in tail]  # Question columns etc. stay where they were.
        df = df[head + tail + remaining]
    return df


# ==================================================
# STEP 2B: PROCESS DATA FOR OUTPUT EXCEL DOWNLOAD
# ==================================================
def process_data_for_download(df, include_all_usertypes=False):
    """Applies download-specific transformations to DataFrame."""
    download_df = df.copy()

    if not include_all_usertypes:
        # --- "Standard" export: only Examiner + (filtered) Moderator rows ---

        # Remove Reval 1 and Reval 2 Rows entirely - they never appear in the standard export.
        if "UserType" in download_df.columns:
            reval_mask = download_df["UserType"].astype(str).str.strip().str.lower().isin(["reval 1", "reval 2"])
            download_df = download_df[~reval_mask]

        # For any student who WAS moderated, only keep their Moderator row (drop the original Examiner row,
        # since the Moderator row now represents the final mark). Students with no moderation keep all rows.
        if "PRNNumber" in download_df.columns and "UserType" in download_df.columns:
            moderator_prns = set(
                download_df[download_df["UserType"].astype(str).str.strip().str.lower() == "moderator"]["PRNNumber"]
            )

            def filter_rows(row):
                prn = row["PRNNumber"]
                u_type = str(row["UserType"]).strip().lower()
                if prn in moderator_prns:
                    return u_type == "moderator"
                return True

            download_df = download_df[download_df.apply(filter_rows, axis=1)]

    # Replace "AB" in TotalObtainedScore with a human-readable label based on attendance.
    if "TotalObtainedScore" in download_df.columns:
        def adjust_score(row):
            score = str(row["TotalObtainedScore"]).strip()
            attendance = str(row.get("MarkAttendance", "")).strip().upper()

            if score.upper() == "AB":
                if attendance == "UFM":
                    return "UFM"      # Unfair means - distinct from a plain absence.
                return "ABSENT"
            return score

        download_df["TotalObtainedScore"] = download_df.apply(adjust_score, axis=1)

    # Drop Unwanted Columns - which set depends on which export flavour this is.
    cols_to_remove = COLUMNS_TO_DROP_FORMATTED if include_all_usertypes else COLUMNS_TO_DROP_STANDARD
    download_df = download_df.drop(columns=cols_to_remove, errors="ignore")

    # Rename Columns to the labels expected in the exported file.
    if "Subject Name-Subject code" in download_df.columns:
        download_df = download_df.rename(columns={"Subject Name-Subject code": "Subject code"})

    if "PRNNumber" in download_df.columns:
        download_df = download_df.rename(columns={"PRNNumber": "SAP ID"})

    download_df = download_df.reset_index(drop=True)

    # Make SrNo Continuous (1, 2, 3, ...) since rows may have been dropped/reordered above.
    if "SrNo" in download_df.columns:
        download_df["SrNo"] = download_df.index + 1

    return download_df


# ==================================================
# STEP 3: DISPLAY FORMATTING HELPERS
# ==================================================
def format_value(value):
    """Formats values cleanly for preview display."""
    text = str(value).strip()

    if text == "" or text.upper() == "NA" or text.upper() == "NAN":
        return "" if text == "" else "NA"

    # Clean leading '+' for formatting display if present
    cleaned_text = re.sub(r"^\+\s*", "", text)

    try:
        number = float(cleaned_text)
        # Show whole numbers without a trailing ".0" (e.g. "8" not "8.0"),
        # but keep decimals for fractional marks (e.g. "7.5").
        if number.is_integer():
            return str(int(number))
        return f"{number:g}"
    except (TypeError, ValueError):
        return text  # Not numeric - show as-is (e.g. "ABSENT", subject names, etc.).


def highlight_rows(data, question_columns):
    """Applies CSS styles for Streamlit preview table safely."""
    styles = pd.DataFrame("", index=data.index, columns=data.columns)

    # Pass 1: whole-row background colour based on UserType (Moderator/Reval 1/Reval 2).
    for idx in data.index:
        if "UserType" in data.columns:
            user_type = str(data.loc[idx, "UserType"]).strip()
            if user_type in ROW_COLOURS:
                styles.loc[idx, :] += ROW_COLOURS[user_type]

    # Pass 2: cell-level green/red highlighting - compare each reviewer row
    # against that student's Examiner row, column by column.
    if "PRNNumber" in data.columns and "UserType" in data.columns:
        for _, group in data.groupby("PRNNumber", dropna=False):
            examiner_rows = group[group["UserType"] == "Examiner"]
            if examiner_rows.empty:
                continue
            examiner_row = examiner_rows.iloc[0]

            for idx in group.index:
                if str(data.loc[idx, "UserType"]) == "Examiner":
                    continue  # Never highlight the baseline Examiner row against itself.

                # Question-wise marks highlighting.
                for col in question_columns:
                    if col not in data.columns:
                        continue

                    examiner_value = str(examiner_row[col]).strip()
                    current_value = str(data.loc[idx, col]).strip()

                    if examiner_value == "" or examiner_value.upper() == "NA":
                        continue  # Examiner didn't mark this question - nothing to compare.

                    if current_value == "" or current_value.upper() == "NA":
                        # Reviewer removed a mark the Examiner had given - treat as a decrease.
                        styles.loc[idx, col] += DECREASE_STYLE
                        continue

                    try:
                        ex_mark = float(examiner_value)
                        cur_mark = float(current_value)
                    except (TypeError, ValueError):
                        continue

                    if cur_mark > ex_mark:
                        styles.loc[idx, col] += INCREASE_STYLE
                    elif cur_mark < ex_mark:
                        styles.loc[idx, col] += DECREASE_STYLE

                # Grade Conditional Formatting (rank comparison, not numeric comparison).
                if "Grade" in data.columns:
                    ex_grade = str(examiner_row.get("Grade", "")).strip().upper()
                    cur_grade = str(data.loc[idx, "Grade"]).strip().upper()

                    if ex_grade in GRADE_RANKS and cur_grade in GRADE_RANKS:
                        ex_rank = GRADE_RANKS[ex_grade]
                        cur_rank = GRADE_RANKS[cur_grade]

                        if cur_rank > ex_rank:
                            styles.loc[idx, "Grade"] += INCREASE_STYLE
                        elif cur_rank < ex_rank:
                            styles.loc[idx, "Grade"] += DECREASE_STYLE

                # Total Marks Obtained Conditional Formatting.
                if "Total Marks Obtained" in data.columns:
                    ex_tot_str = str(examiner_row.get("Total Marks Obtained", "")).strip()
                    cur_tot_str = str(data.loc[idx, "Total Marks Obtained"]).strip()
                    if ex_tot_str != "" and cur_tot_str != "":
                        try:
                            ex_tot = float(ex_tot_str)
                            cur_tot = float(cur_tot_str)
                            if cur_tot > ex_tot:
                                styles.loc[idx, "Total Marks Obtained"] += INCREASE_STYLE
                            elif cur_tot < ex_tot:
                                styles.loc[idx, "Total Marks Obtained"] += DECREASE_STYLE
                        except (ValueError, TypeError):
                            pass

                # Change % Conditional Formatting (sign-based: positive = green, negative = red).
                if "Change %" in data.columns:
                    chg_str = str(data.loc[idx, "Change %"]).strip()
                    if chg_str != "":
                        try:
                            chg_val = float(chg_str)
                            if chg_val > 0:
                                styles.loc[idx, "Change %"] += INCREASE_STYLE
                            elif chg_val < 0:
                                styles.loc[idx, "Change %"] += DECREASE_STYLE
                        except (ValueError, TypeError):
                            pass

    # Pass 3: Semester Marks vs TotalObtainedScore comparison - light green cell
    # fill if they match, light red if they don't. Only applies to rows where
    # Semester Marks was actually populated (Moderator rows, or Examiner rows
    # with no corresponding Moderator record - see apply_gmr_lookup).
    if "Semester Marks" in data.columns and "TotalObtainedScore" in data.columns:
        for idx in data.index:
            sem_val_str = str(data.loc[idx, "Semester Marks"]).strip()
            if sem_val_str == "" or sem_val_str.upper() == "NA":
                continue

            tot_val_str = str(data.loc[idx, "TotalObtainedScore"]).strip()
            try:
                sem_val = float(re.sub(r"^\+\s*", "", sem_val_str))
                tot_val = float(tot_val_str)
            except (ValueError, TypeError):
                continue

            if sem_val == tot_val:
                styles.loc[idx, "Semester Marks"] += "background-color:#C6EFCE;color:#000000;"
            else:
                styles.loc[idx, "Semester Marks"] += "background-color:#ff0000;color:#ffff00;"

    # TotalObtainedScore is always bold, regardless of row type, for quick scanning.
    if "TotalObtainedScore" in data.columns:
        styles.loc[:, "TotalObtainedScore"] += "font-weight:bold;"

    return styles


def style_preview(df, question_columns, hide_columns=PREVIEW_HIDDEN_COLUMNS):
    """Builds styled interactive preview table for web display."""
    display_df = df.drop(columns=hide_columns, errors="ignore").copy()
    display_df.index = range(1, len(display_df) + 1)  # 1-based row numbers look nicer to end users.
    return (
        display_df.style
        .format(format_value)
        .apply(lambda data: highlight_rows(data, question_columns), axis=None)
    )


# ==================================================
# STEP 4: SUMMARY STATISTICS & METADATA
# ==================================================
def count_reviewers(df, user_type):
    """Number of distinct students that have a row of the given UserType."""
    if "UserType" not in df.columns or "PRNNumber" not in df.columns:
        return 0
    return (
        df[df["UserType"].astype(str).str.strip().eq(user_type)]["PRNNumber"]
        .nunique()
    )


def score_change_summary(df, review_type):
    """Compare review_type TotalObtainedScore against Examiner's TotalObtainedScore."""
    increased = decreased = unchanged = 0

    if "TotalObtainedScore" not in df.columns or "UserType" not in df.columns:
        return increased, decreased, unchanged

    for _, group in df.groupby("PRNNumber", dropna=False):
        examiner_rows = group[group["UserType"] == "Examiner"]
        review_rows = group[group["UserType"] == review_type]

        if examiner_rows.empty or review_rows.empty:
            continue  # Skip students who don't have both an Examiner row and a row of this review_type.

        try:
            examiner_score = float(examiner_rows.iloc[0]["TotalObtainedScore"])
            review_score = float(review_rows.iloc[0]["TotalObtainedScore"])
        except (TypeError, ValueError):
            continue  # Non-numeric score (e.g. "AB") - can't compare, skip.

        if review_score > examiner_score:
            increased += 1
        elif review_score < examiner_score:
            decreased += 1
        else:
            unchanged += 1

    return increased, decreased, unchanged


def compute_moderation_10pct_metrics(df):
    """
    Computes:
    - Moderator Count
    - Count of Moderator records with |Change %| > 10%
    - Change % (>10%) ratio & percentage
    - Eligibility for 100% Moderation (>50% threshold)
    """
    if "Change %" not in df.columns or "UserType" not in df.columns:
        return 0, 0, 0.0, False

    mod_rows = df[df["UserType"].astype(str).str.strip() == "Moderator"]
    mod_count = len(mod_rows)

    if mod_count == 0:
        return 0, 0, 0.0, False

    count_gt_10pct = 0
    for val in mod_rows["Change %"]:
        try:
            if abs(float(val)) > 10.0:
                count_gt_10pct += 1
        except (ValueError, TypeError):
            continue  # Blank Change % (no comparable Examiner row) - doesn't count either way.

    ratio_gt_10pct = (count_gt_10pct / mod_count) if mod_count > 0 else 0.0
    # Regulation threshold: if MORE than half of moderated scripts moved by >10%,
    # the whole paper must go through 100% moderation.
    is_eligible_100pct_moderation = ratio_gt_10pct > 0.50

    return mod_count, count_gt_10pct, ratio_gt_10pct, is_eligible_100pct_moderation


def build_paper_info(df):
    """Builds header info line for display."""
    if df.empty:
        return ""

    semester = str(df.iloc[0].get("Semester", "")).strip()
    subject_code = str(df.iloc[0].get("Subject Name-Subject code", "")).strip()
    subject_name = str(df.iloc[0].get("SubjectName", "")).strip()

    return f"{semester}  -  {subject_code}  -  {subject_name}"


def generate_output_filename(df, idx=1, prefix=""):
    """Generates dynamic Excel filename."""
    if df.empty:
        return f"{prefix}Questionwise_Output_{idx}_{datetime.now().strftime('%d-%m-%y')}.xlsx"

    subject_name = str(df.iloc[0].get("SubjectName", "Subject")).strip()
    semester_raw = str(df.iloc[0].get("Semester", "Semester")).strip()

    # Shorten "Semester X" to "Sem X" to keep filenames compact.
    semester_short = re.sub(r"(?i)\bsemester\b", "Sem", semester_raw).strip()

    unique_prns = df["PRNNumber"].nunique() if "PRNNumber" in df.columns else len(df)
    date_str = datetime.now().strftime("%d-%m-%y")

    filename = f"{prefix}{subject_name}_{semester_short}_{unique_prns}_{date_str}.xlsx"
    return filename


# ==================================================
# STEP 5: EXCEL GENERATION & STYLING
# ==================================================
def sanitize_sheet_name(subject_identifier):
    """Sanitizes SubjectName / Subject Code into a valid Excel sheet name (max 31 chars)."""
    if not subject_identifier:
        return "Questionwise Data"

    name = str(subject_identifier)
    name = re.sub(r"\[[^\]]*\]", "", name)          # Drop any bracketed suffix, e.g. "[2024]".
    name = re.sub(r"[\\/?*\[\]:]", "", name)          # Excel forbids these characters in sheet names.
    name = re.sub(r"\s+", " ", name).strip()

    if not name:
        name = "Questionwise Data"

    return name[:31]  # Excel's hard limit on sheet-name length.


def get_sheet_name_for_df(df, default_idx=1):
    """Retrieves sheet name based on 'Subject Name-Subject code' or 'SubjectName'."""
    if not df.empty:
        if "SubjectName" in df.columns and str(df.iloc[0]["SubjectName"]).strip():
            return sanitize_sheet_name(df.iloc[0]["SubjectName"])
        if "Subject Name-Subject code" in df.columns and str(df.iloc[0]["Subject Name-Subject code"]).strip():
            return sanitize_sheet_name(df.iloc[0]["Subject Name-Subject code"])
    return f"Sheet_{default_idx}"


def add_sheet_popup_validation(worksheet, is_eligible):
    """
    Applies Data Validation Input Message to cell A1.
    When a user clicks on/selects cell A1 in Excel, a dynamic floating notification pop-up appears.
    """
    if not is_eligible:
        return

    # A "custom" validation with an always-true formula never blocks input - it exists
    # purely to piggyback on Excel's input-message pop-up as a visible alert banner.
    dv = DataValidation(
        type="custom",
        formula1="TRUE",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=False,
        promptTitle="⚠️ 100% Moderation Alert",
        prompt="The 10% change in marks is observed in more than 50% of the moderated answer books. The paper is eligible for 100% Moderation."
    )
    worksheet.add_data_validation(dv)
    dv.add(worksheet["A1"])


# --------------------------------------------------
# Shared Excel-styling helpers
# --------------------------------------------------

def _thin_border():
    """Returns a Border with a thin black line on all four sides."""
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _center_column_indices(cols, explicit_center_names):
    """
    Works out the 1-based column indices that should be center-aligned.

    Two rules combine:
      1. Any column whose name is in `explicit_center_names`.
      2. Every column from 'SAP ID' to 'TotalObtainedScore' inclusive - this
         is always the block of per-question mark columns.
    """
    indices = [i for i, name in enumerate(cols, start=1) if name in explicit_center_names]

    if "SAP ID" in cols and "TotalObtainedScore" in cols:
        sap_idx = cols.index("SAP ID") + 1
        tot_idx = cols.index("TotalObtainedScore") + 1
        indices.extend(range(sap_idx, tot_idx + 1))

    return set(indices)


def _autosize_or_fixed_width_columns(worksheet, fixed_width_cols=None, fixed_width=15):
    """
    Sets a fixed width for any column named in `fixed_width_cols`; every other
    column auto-sizes to fit its longest cell value (+3 padding, 10 minimum).
    """
    fixed_width_cols = fixed_width_cols or set()
    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = str(col[0].value).strip()

        if col_name in fixed_width_cols:
            worksheet.column_dimensions[col_letter].width = fixed_width
        else:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)


def _apply_grid_styling(worksheet, cols, center_col_indices, total_score_col_idx,
                         download_df=None, user_type_col_idx=None, moderator_fill=None):
    """
    Applies the shared per-cell formatting pass used by every exported sheet:
      - thin border on every cell
      - bold, grey, wrapped & centered header row (row 1)
      - bold font on the 'TotalObtainedScore' column
      - center alignment on any column flagged in `center_col_indices`
      - (formatted sheets only) a light-green fill on rows whose UserType is
        'Moderator' - enabled by passing `moderator_fill`, `user_type_col_idx`
        and `download_df` together; omitted entirely for the plain export.
    """
    thin_border = _thin_border()
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                    min_col=1, max_col=worksheet.max_column):
        is_moderator_row = False
        if moderator_fill is not None and user_type_col_idx is not None and download_df is not None:
            row_idx = row[0].row - 2  # Excel row 2 == DataFrame index 0 (row 1 is the header).
            if 0 <= row_idx < len(download_df):
                u_type = str(download_df.iloc[row_idx].get("UserType", "")).strip().lower()
                is_moderator_row = (u_type == "moderator")

        for cell in row:
            cell.border = thin_border

            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = grey_fill
                cell.alignment = center_wrap_align
            else:
                if is_moderator_row:
                    cell.fill = moderator_fill
                if total_score_col_idx and cell.column == total_score_col_idx:
                    cell.font = Font(bold=True)

            if cell.row > 1 and cell.column in center_col_indices:
                cell.alignment = center_align


def _apply_header_highlight_fills(worksheet, cols, orange_header_cols, dark_grey_header_cols):
    """
    Overrides the default grey header-row fill (set by _apply_grid_styling)
    for specific header cells only:
      - `orange_header_cols` get a medium orange, 60% lighter fill (FCE4D6)
      - `dark_grey_header_cols` get a 35% darker grey fill (999999)
    Only row 1 (the header row) is touched; font/border/alignment set
    elsewhere are left untouched.
    """
    orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    dark_grey_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")

    for col_name in orange_header_cols:
        if col_name in cols:
            col_idx = cols.index(col_name) + 1
            worksheet.cell(row=1, column=col_idx).fill = orange_fill

    for col_name in dark_grey_header_cols:
        if col_name in cols:
            col_idx = cols.index(col_name) + 1
            worksheet.cell(row=1, column=col_idx).fill = dark_grey_fill


def _apply_change_highlighting_and_group_borders(worksheet, download_df, cols, bright_green_font, red_font):
    """
    For every student (grouped by 'SAP ID'), compares each reviewer row
    (Moderator / Reval 1 / Reval 2) against that student's Examiner row and:
      - colors question marks, Grade, Total Marks Obtained and Change % cells
        green if the value improved, red if it dropped or a mark disappeared
      - draws a medium-thickness bottom border under the student's last row
        so each student's block is visually separated on the sheet

    Only used by the "formatted" exports (all UserTypes + GMR columns).
    """
    if "UserType" not in download_df.columns or "SAP ID" not in download_df.columns:
        return

    thin_side = Side(style="thin", color="000000")
    medium_side = Side(style="medium", color="000000")

    question_cols = [c for c in download_df.columns if c not in FIXED_COLUMNS and c not in COLUMNS_TO_DROP_FORMATTED]

    for sap_id, group in download_df.groupby("SAP ID", dropna=False):
        examiner_rows = group[group["UserType"] == "Examiner"]
        if examiner_rows.empty:
            continue
        examiner_row = examiner_rows.iloc[0]

        for row_idx in group.index:
            if str(download_df.loc[row_idx, "UserType"]).strip() == "Examiner":
                continue  # Never highlight the Examiner row against itself.

            excel_row_num = row_idx + 2  # +1 for header row, +1 for 1-based Excel rows.

            # --- Question-wise marks ---
            for q_col in question_cols:
                if q_col not in download_df.columns:
                    continue
                col_idx = cols.index(q_col) + 1
                ex_val_str = str(examiner_row[q_col]).strip()
                cur_val_str = str(download_df.loc[row_idx, q_col]).strip()

                if ex_val_str == "" or ex_val_str.upper() == "NA":
                    continue  # Examiner never marked this question - nothing to compare against.

                cell = worksheet.cell(row=excel_row_num, column=col_idx)

                if cur_val_str == "" or cur_val_str.upper() == "NA":
                    cell.font = red_font  # Reviewer's mark went missing - treat as a decrease.
                    continue

                try:
                    ex_val = float(ex_val_str)
                    cur_val = float(cur_val_str)
                    if cur_val > ex_val:
                        cell.font = bright_green_font
                    elif cur_val < ex_val:
                        cell.font = red_font
                except (ValueError, TypeError):
                    pass

            # --- Grade (rank comparison) ---
            if "Grade" in download_df.columns:
                grade_col_idx = cols.index("Grade") + 1
                ex_grade = str(examiner_row.get("Grade", "")).strip().upper()
                cur_grade = str(download_df.loc[row_idx, "Grade"]).strip().upper()

                if ex_grade in GRADE_RANKS and cur_grade in GRADE_RANKS:
                    cell = worksheet.cell(row=excel_row_num, column=grade_col_idx)
                    if GRADE_RANKS[cur_grade] > GRADE_RANKS[ex_grade]:
                        cell.font = bright_green_font
                    elif GRADE_RANKS[cur_grade] < GRADE_RANKS[ex_grade]:
                        cell.font = red_font

            # --- Total Marks Obtained ---
            if "Total Marks Obtained" in download_df.columns:
                tot_obt_col_idx = cols.index("Total Marks Obtained") + 1
                ex_tot_str = str(examiner_row.get("Total Marks Obtained", "")).strip()
                cur_tot_str = str(download_df.loc[row_idx, "Total Marks Obtained"]).strip()

                if ex_tot_str != "" and cur_tot_str != "":
                    cell = worksheet.cell(row=excel_row_num, column=tot_obt_col_idx)
                    try:
                        ex_tot = float(ex_tot_str)
                        cur_tot = float(cur_tot_str)
                        if cur_tot > ex_tot:
                            cell.font = bright_green_font
                        elif cur_tot < ex_tot:
                            cell.font = red_font
                    except (ValueError, TypeError):
                        pass

            # --- Change % (sign-based) ---
            if "Change %" in download_df.columns:
                chg_col_idx = cols.index("Change %") + 1
                chg_str = str(download_df.loc[row_idx, "Change %"]).strip()

                if chg_str != "":
                    cell = worksheet.cell(row=excel_row_num, column=chg_col_idx)
                    try:
                        chg_val = float(chg_str)
                        if chg_val > 0:
                            cell.font = bright_green_font
                        elif chg_val < 0:
                            cell.font = red_font
                    except (ValueError, TypeError):
                        pass

        # Thick bottom border under this student's last row, marking the end of their block.
        last_row_idx = group.index[-1] + 2
        for c_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=last_row_idx, column=c_idx)
            cell.border = Border(left=thin_side, right=thin_side, top=cell.border.top, bottom=medium_side)


def _apply_semester_marks_highlighting(worksheet, download_df, cols, green_fill, red_fill):
    """
    For every row where 'Semester Marks' is populated (Moderator rows, or an
    Examiner row with no corresponding Moderator record), compares it against
    that same row's TotalObtainedScore and fills the cell:
      - light green if Semester Marks == TotalObtainedScore
      - light red if Semester Marks != TotalObtainedScore
    Rows where Semester Marks is blank are left untouched.
    """
    if "Semester Marks" not in download_df.columns or "TotalObtainedScore" not in download_df.columns:
        return

    sem_col_idx = cols.index("Semester Marks") + 1

    for row_idx in download_df.index:
        sem_val_str = str(download_df.loc[row_idx, "Semester Marks"]).strip()
        if sem_val_str == "" or sem_val_str.upper() == "NA":
            continue

        tot_val_str = str(download_df.loc[row_idx, "TotalObtainedScore"]).strip()
        try:
            sem_val = float(re.sub(r"^\+\s*", "", sem_val_str))
            tot_val = float(tot_val_str)
        except (ValueError, TypeError):
            continue

        excel_row_num = row_idx + 2  # +1 for header row, +1 for 1-based Excel rows.
        cell = worksheet.cell(row=excel_row_num, column=sem_col_idx)
        cell.fill = green_fill if sem_val == tot_val else red_fill


def to_excel_bytes(df):
    """ORIGINAL Standard clean output without GMR columns."""
    download_df = process_data_for_download(df, include_all_usertypes=False)
    buffer = io.BytesIO()

    sheet_name = "Questionwise Data"
    if not df.empty and "SubjectName" in df.columns:
        subj = str(df.iloc[0]["SubjectName"]).strip()
        if subj:
            sheet_name = sanitize_sheet_name(subj)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        download_df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]

        worksheet.views.sheetView[0].showGridLines = False
        worksheet.freeze_panes = "A2"          # Freeze only the header row (column A itself scrolls).
        worksheet.row_dimensions[1].height = 30

        cols = download_df.columns.tolist()
        center_col_indices = _center_column_indices(cols, {"SrNo", "Subject code"})
        total_score_col_idx = cols.index("TotalObtainedScore") + 1 if "TotalObtainedScore" in cols else None

        _apply_grid_styling(worksheet, cols, center_col_indices, total_score_col_idx)
        _autosize_or_fixed_width_columns(worksheet)  # No fixed-width columns in this export - autosize everything.

    return buffer.getvalue()


def to_excel_bytes_formatted(df):
    """
    Formatted output Excel containing ALL UserTypes, GMR columns, Change %,
    center alignment, fixed width 10 for GMR/Change %/Email columns, row height 30 for header with wrap text,
    bright green text (#00B050), light green fill (#E2EFDA) for Moderator rows, thick bottom border per student, 
    F2 freeze pane, conditional formatting, medium-orange (60% lighter) GMR/Change % header cells,
    35%-darker-grey TotalObtainedScore header cell, and interactive Data Validation pop-up if eligible.
    """
    download_df = process_data_for_download(df, include_all_usertypes=True)
    buffer = io.BytesIO()

    sheet_name = get_sheet_name_for_df(df)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        download_df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]

        worksheet.views.sheetView[0].showGridLines = False
        worksheet.freeze_panes = "F2"          # Freeze header row + first 5 identifying columns.
        worksheet.row_dimensions[1].height = 30

        mod_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bright_green_font = Font(color="00B050", bold=True)  # Standard Excel bright green for high visibility
        red_font = Font(color="FF4500", bold=True)
        sem_marks_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        sem_marks_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        cols = download_df.columns.tolist()

        fixed_10_cols = {
            "Semester Total Max Marks", "Semester Marks", "Internal Marks",
            "Total Max Marks", "Total Marks Obtained", "Grade", "Change %"
        }
        align_center_col_names = {
            "SrNo", "Subject code", "Grade", "Internal Marks", "Semester Marks",
            "Total Max Marks", "Semester Total Max Marks", "Total Marks Obtained", "Change %"
        }

        center_col_indices = _center_column_indices(cols, align_center_col_names)
        total_score_col_idx = cols.index("TotalObtainedScore") + 1 if "TotalObtainedScore" in cols else None
        user_type_col_idx = cols.index("UserType") if "UserType" in cols else None

        # Basic styling + Moderator light-green row fill.
        _apply_grid_styling(
            worksheet, cols, center_col_indices, total_score_col_idx,
            download_df=download_df, user_type_col_idx=user_type_col_idx, moderator_fill=mod_fill,
        )

        # Green/red change highlighting + thick per-student borders.
        _apply_change_highlighting_and_group_borders(worksheet, download_df, cols, bright_green_font, red_font)

        # Semester Marks vs TotalObtainedScore match/mismatch cell fill.
        _apply_semester_marks_highlighting(worksheet, download_df, cols, sem_marks_green_fill, sem_marks_red_fill)

        # Header cell fills: GMR/Change % columns -> medium orange (60% lighter);
        # TotalObtainedScore -> 35% darker grey.
        _apply_header_highlight_fills(worksheet, cols, fixed_10_cols, {"TotalObtainedScore"})

        # Interactive pop-up Data Validation message if this paper is eligible for 100% Moderation.
        _, _, _, is_eligible = compute_moderation_10pct_metrics(df)
        add_sheet_popup_validation(worksheet, is_eligible)

        # Auto-size columns, except the GMR/Change % block which gets a fixed width of 10.
        _autosize_or_fixed_width_columns(worksheet, fixed_width_cols=fixed_10_cols, fixed_width=10)

    return buffer.getvalue()


def create_consolidated_multi_sheet_excel(dfs_list):
    """
    Clubs all processed dataframes into a single Excel workbook 
    with separate sheets named by Subject Name-Subject code containing ALL UserTypes,
    Change %, center alignment, fixed width 10 for GMR/Change %/Email columns, F2 freeze pane,
    thick student borders, medium-orange (60% lighter) GMR/Change % header cells, 35%-darker-grey
    TotalObtainedScore header cell, and interactive pop-up message if eligible.
    """
    buffer = io.BytesIO()
    used_sheet_names = set()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for idx, (df, orig_name) in enumerate(dfs_list, start=1):
            download_df = process_data_for_download(df, include_all_usertypes=True)

            # Ensure sheet names are unique within the workbook (Excel forbids duplicates).
            base_sheet_name = get_sheet_name_for_df(df, idx)
            sheet_name = base_sheet_name
            counter = 1
            while sheet_name in used_sheet_names:
                sheet_name = f"{base_sheet_name[:28]}_{counter}"
                counter += 1
            used_sheet_names.add(sheet_name)

            download_df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            worksheet.views.sheetView[0].showGridLines = False
            worksheet.freeze_panes = "F2"
            worksheet.row_dimensions[1].height = 30

            mod_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            bright_green_font = Font(color="00B050", bold=True)  # Standard Excel bright green for high visibility
            red_font = Font(color="FF4500", bold=True)
            sem_marks_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            sem_marks_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            cols = download_df.columns.tolist()

            fixed_10_cols = {
                "Semester Total Max Marks", "Semester Marks", "Internal Marks",
                "Total Max Marks", "Total Marks Obtained", "Grade", "Change %"
            }
            align_center_col_names = {
                "SrNo", "Subject code", "Grade", "Internal Marks", "Semester Marks",
                "Total Max Marks", "Semester Total Max Marks", "Total Marks Obtained", "Change %"
            }

            center_col_indices = _center_column_indices(cols, align_center_col_names)
            total_score_col_idx = cols.index("TotalObtainedScore") + 1 if "TotalObtainedScore" in cols else None
            user_type_col_idx = cols.index("UserType") if "UserType" in cols else None

            # Same styling passes as to_excel_bytes_formatted(), applied per-sheet.
            _apply_grid_styling(
                worksheet, cols, center_col_indices, total_score_col_idx,
                download_df=download_df, user_type_col_idx=user_type_col_idx, moderator_fill=mod_fill,
            )
            _apply_change_highlighting_and_group_borders(worksheet, download_df, cols, bright_green_font, red_font)
            _apply_semester_marks_highlighting(worksheet, download_df, cols, sem_marks_green_fill, sem_marks_red_fill)

            # Header cell fills: GMR/Change % columns -> medium orange (60% lighter);
            # TotalObtainedScore -> 35% darker grey.
            _apply_header_highlight_fills(worksheet, cols, fixed_10_cols, {"TotalObtainedScore"})

            _, _, _, is_eligible = compute_moderation_10pct_metrics(df)
            add_sheet_popup_validation(worksheet, is_eligible)

            _autosize_or_fixed_width_columns(worksheet, fixed_width_cols=fixed_10_cols, fixed_width=10)

    return buffer.getvalue()


def create_zip_file(excel_files):
    """Creates a ZIP archive containing multiple Excel file bytes."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for filename, file_bytes in excel_files:
            zip_file.writestr(filename, file_bytes)
    return zip_buffer.getvalue()


# ==================================================
# MAIN ENTRY POINT
# ==================================================
def show():
    """Render the Questionwise Checker page in Streamlit."""
    st.title("📝 Questionwise Checker")
    st.caption("Upload password-protected or unprotected Excel files.")

    # Side-by-side Upload Controls.
    col_up1, col_up2 = st.columns(2)

    with col_up1:
        uploaded_files = st.file_uploader(
            "Upload Questionwise Excel/Excels",
            type=["xlsx"],
            accept_multiple_files=True
        )

    with col_up2:
        gmr_file = st.file_uploader(
            "Upload GMR Excel",
            type=["xlsx", "xls"]
        )

    if not uploaded_files:
        # Nothing to process yet - still show the Help section so the user can
        # read the documentation before uploading anything.
        st.markdown("---")
        render_help_section()
        return  # Wait for the user to upload at least one Questionwise file.

    # Load GMR File if present (optional - the tool still works without it, just without GMR columns).
    gmr_df = None
    if gmr_file is not None:
        with st.spinner("Loading GMR Excel..."):
            gmr_df = load_gmr_file(gmr_file)
            if gmr_df is not None:
                st.caption("✅ GMR File uploaded and loaded successfully.")
            else:
                st.warning("⚠️ Could not load GMR Excel file.")

    processed_outputs_clean = []   # (filename, bytes) pairs for the bulk ZIP download.
    dfs_for_multi_sheet = []       # (DataFrame, original filename) pairs for the consolidated workbook.

    # Process each uploaded Questionwise file independently - one file's error
    # shouldn't stop the others from being processed (see the try/except below).
    for idx, uploaded_file in enumerate(uploaded_files, start=1):
        st.markdown("---")
        st.subheader(f"📁 File {idx}: {uploaded_file.name}")

        try:
            with st.spinner(f"Decrypting and processing {uploaded_file.name}..."):
                df, used_password = decrypt_and_load(uploaded_file)

            if df is None:
                st.error(f"Unable to open {uploaded_file.name}. Password not recognised.")
                continue

            st.caption(f"Opened successfully using: {used_password}")

            # Clean and Sort.
            df = clean_and_sort(df)

            if df.empty:
                st.warning(f"No valid student rows found in {uploaded_file.name}.")
                continue

            if "PRNNumber" not in df.columns or "UserType" not in df.columns:
                st.error(f"Missing required 'PRNNumber' or 'UserType' columns in {uploaded_file.name}.")
                continue

            # Apply GMR Lookup Match & Change % Calculation.
            df = apply_gmr_lookup(df, gmr_df)
            df = calculate_moderator_change_pct(df)
            df = reorder_gmr_columns(df)

            dfs_for_multi_sheet.append((df, uploaded_file.name))

            question_columns = [col for col in df.columns if col not in FIXED_COLUMNS]

            # --- Summary statistics ---
            total_students = df["PRNNumber"].nunique()
            moderation_count = count_reviewers(df, "Moderator")
            reval1_count = count_reviewers(df, "Reval 1")
            reval2_count = count_reviewers(df, "Reval 2")

            moderator_positive, moderator_negative, moderator_no_change = score_change_summary(df, "Moderator")
            reval1_positive, reval1_negative, reval1_no_change = score_change_summary(df, "Reval 1")
            reval2_positive, reval2_negative, reval2_no_change = score_change_summary(df, "Reval 2")

            mod_cnt, cnt_gt_10, ratio_gt_10, is_eligible = compute_moderation_10pct_metrics(df)

            # --- 5 Summary Cards Grid ---
            col1, col2, col3, col4, col5 = st.columns(5)

            with col1:
                st.info(
                    f"Total Students: {total_students}\n\n"
                    f"Moderation Count: {moderation_count}\n\n"
                    f"Reval 1 Count: {reval1_count}\n\n"
                    f"Reval 2 Count: {reval2_count}"
                )

            with col2:
                st.success(
                    f"Positive Change in Marks:\n\n"
                    f"Moderation: {moderator_positive}\n\n"
                    f"Reval 1: {reval1_positive}\n\n"
                    f"Reval 2: {reval2_positive}"
                )

            with col3:
                st.error(
                    f"Negative Change in Marks:\n\n"
                    f"Moderation: {moderator_negative}\n\n"
                    f"Reval 1: {reval1_negative}\n\n"
                    f"Reval 2: {reval2_negative}"
                )

            with col4:
                st.warning(
                    f"No Change in Marks:\n\n"
                    f"Moderation: {moderator_no_change}\n\n"
                    f"Reval 1: {reval1_no_change}\n\n"
                    f"Reval 2: {reval2_no_change}"
                )

            with col5:
                # 5th Card: Change % Metrics.
                pct_str = f"{ratio_gt_10 * 100:.1f}%"
                color_style = "color:#FF4500;" if is_eligible else "color:#9ACD32;"

                st.markdown(
                    f"""
                    <div style="background-color:#1E293B;padding:12px;border-radius:8px;color:white;border:1px solid #334155;">
                        <span style="font-size:15px;font-weight:600;"> Require 100% Moderation?</span><br/>
                        <span style="font-size:15px;"> Change % > 10%: <b>{cnt_gt_10} / {mod_cnt}</b></span><br/>                        
                        <span style="font-size:60px;font-weight:bold;{color_style}">{pct_str}</span>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            if is_eligible:
                st.error("Above 10% change in marks is observed in more than 50% of the moderated answer books. This module is eligible for 100% Moderation.")

            st.subheader(build_paper_info(df))

            # --- Filter preview subsets ---
            no_attendance_hidden_columns = PREVIEW_HIDDEN_COLUMNS + ["MarkAttendance"]

            review_prns = df[
                df["UserType"].astype(str).str.strip().isin(["Moderator", "Reval 1", "Reval 2"])
            ]["PRNNumber"].unique()
            review_df = df[df["PRNNumber"].isin(review_prns)].copy()

            # Students where ANY reviewer's TotalObtainedScore actually differs from the Examiner's.
            changed_prns = []
            if "TotalObtainedScore" in df.columns and "UserType" in df.columns:
                for prn, group in df.groupby("PRNNumber", dropna=False, sort=False):
                    examiner_rows = group[group["UserType"] == "Examiner"]
                    reviewer_rows = group[group["UserType"].isin(["Moderator", "Reval 1", "Reval 2"])]
                    if not examiner_rows.empty and not reviewer_rows.empty:
                        try:
                            ex_score = float(examiner_rows.iloc[0]["TotalObtainedScore"])
                            for _, rev_row in reviewer_rows.iterrows():
                                rev_score = float(rev_row["TotalObtainedScore"])
                                if rev_score != ex_score:
                                    changed_prns.append(prn)
                                    break  # One differing reviewer is enough to flag this student.
                        except (TypeError, ValueError):
                            continue  # Non-numeric score - can't determine "changed", skip this student.

            changes_df = df[df["PRNNumber"].isin(changed_prns)].copy() if changed_prns else pd.DataFrame()

            # --- Display Tabs ---
            tab1, tab2, tab3 = st.tabs([
                "📋 All Data Preview",
                "🔍 Moderation / Revaluation Cases Only",
                "⚡ Changes Only"
            ])

            with tab1:
                st.subheader("All Data Preview")
                st.write(style_preview(df, question_columns))

            with tab2:
                st.subheader("Moderation / Revaluation Cases Only")
                if not review_df.empty:
                    st.write(style_preview(review_df, question_columns, hide_columns=no_attendance_hidden_columns))
                else:
                    st.info("No moderation or revaluation cases found in this file.")

            with tab3:
                st.subheader("Changes Only")
                if not changes_df.empty:
                    st.write(style_preview(changes_df, question_columns, hide_columns=no_attendance_hidden_columns))
                else:
                    st.info("No score changes detected between examiner and reviewers.")

            # --- Generate individual Excel Outputs ---
            out_bytes_clean = to_excel_bytes(df)
            out_bytes_fmt = to_excel_bytes_formatted(df)
            out_filename = generate_output_filename(df, idx)
            out_filename_fmt = generate_output_filename(df, idx, prefix="Formatted_")

            # Save original clean outputs for the bulk-download ZIP.
            processed_outputs_clean.append((out_filename, out_bytes_clean))

            # Side-by-Side Individual File Download Buttons.
            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    label=f"📥 Download Cleaned File: {out_filename}",
                    data=out_bytes_clean,
                    file_name=out_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_single_clean_{idx}"
                )
            with dl_col2:
                st.download_button(
                    label=f"🎨 Download Formatted File: {out_filename_fmt}",
                    data=out_bytes_fmt,
                    file_name=out_filename_fmt,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_single_fmt_{idx}"
                )

        except Exception as error:
            # Isolate failures per-file so one bad upload doesn't block the rest of the batch.
            st.error(f"Error processing file {uploaded_file.name}: {error}")

    # --- Bulk Download Options ---
    if processed_outputs_clean:
        st.markdown("---")
        st.header("📦 Bulk Download Options")

        bulk_col1, bulk_col2 = st.columns(2)

        # 1. Original ZIP File Download (Clean Files).
        with bulk_col1:
            zip_filename = f"Questionwise_Processed_{datetime.now().strftime('%d-%m-%y_%H%M%S')}.zip"
            zip_bytes = create_zip_file(processed_outputs_clean)

            st.download_button(
                label="📦 Download All Output Excel Files (ZIP)",
                data=zip_bytes,
                file_name=zip_filename,
                mime="application/zip",
                key="dl_zip_all"
            )

        # 2. Consolidated Multi-Sheet Excel Download (only worth offering for 2+ files).
        with bulk_col2:
            if len(dfs_for_multi_sheet) > 1:
                multi_sheet_bytes = create_consolidated_multi_sheet_excel(dfs_for_multi_sheet)
                multi_sheet_filename = f"Consolidated_Questionwise_{datetime.now().strftime('%d-%m-%y_%H%M%S')}.xlsx"

                st.download_button(
                    label="📊 Download Consolidated Multi-Sheet Excel",
                    data=multi_sheet_bytes,
                    file_name=multi_sheet_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_multi_sheet_all"
                )

    # --- Tool Description / Help (collapsed by default) ---
    st.markdown("---")
    render_help_section()


def render_help_section():
    """
    Renders a collapsible 'Tool Description / Help' section at the bottom of
    the page. Collapsed by default (expanded=False) - it only opens when the
    user clicks on it. Purely documentation; no processing logic lives here.
    """
    with st.expander("ℹ️ Tool Description / Help - click to expand", expanded=False):
        st.markdown(
            """
### 1. Purpose of the Tool
The Questionwise Checker helps exam cell / evaluation staff process **Questionwise mark-sheet
exports** from the examination system. It opens password-protected or unprotected Excel files,
cleans and sorts them, optionally cross-references a **GMR (Grade/Marks Register) Excel file** to
pull in Internal Marks, Max Marks, Semester Marks and computed Grades, calculates the **Change %**
introduced by moderation, flags papers that qualify for **100% moderation**, and produces several
ready-to-use preview tabs and downloadable Excel outputs.

---
### 2. Required Input Files & Formats
- **Questionwise Excel/Excels** (`.xlsx`, one or more files) - the raw export from the examination
  system. May be password-protected (the tool automatically tries a small list of known passwords)
  or unprotected. Must contain at minimum a `PRNNumber` and `UserType` column after cleaning, plus
  the standard identifier columns (`SrNo`, `Subject Name-Subject code`, `SubjectName`, `Semester`,
  `RollNo`, `Name`, `CampusName`, `MarkAttendance`, `Email`, `TotalObtainedScore`) and one column
  per question (headers like `Q1 (10)` where the number in parentheses is that question's max mark).
- **GMR Excel** (`.xlsx` / `.xls`, single file, optional) - the Grade/Marks Register export. Must
  contain `Student Number`, `Modulecode`, `AGR type`, and `Grades` columns for any lookups to run;
  `MAX Marks` is additionally needed for the Max Marks lookups. If the GMR file is not uploaded, or
  is missing these columns, the tool still works - the GMR-derived columns are simply left blank.

---
### 3. Upload Options
- **Upload Questionwise Excel/Excels** - accepts multiple files at once; each is processed and
  displayed independently, one section per file, so one bad/corrupt file does not block the rest.
- **Upload GMR Excel** - a single optional file used to enrich *every* uploaded Questionwise file
  with Internal Marks, Total Max Marks, Semester Total Max Marks, Semester Marks and Grade.

---
### 4. Matching Logic Between Questionwise and GMR Files
Every lookup uses the same composite key: **PRNNumber + Subject Name-Subject code** (Questionwise
side) matched against **Student Number + Modulecode** (GMR side), both stripped of surrounding
whitespace and concatenated into a single lookup string. A Questionwise row is only enriched if its
key finds an exact match in the GMR file.

---
### 5. Conditions & Filters Applied During Processing
- Rows with a blank `PRNNumber`, fully-empty rows, and the export's trailing `Count: N` footer row
  are removed during cleaning.
- `PRNNumber` is coerced to numeric; any row that isn't a valid number is dropped.
- Rows are sorted per student in the fixed order **Examiner → Moderator → Reval 1 → Reval 2**.

---
### 6. Grade, Internal Marks, Internal Total Marks, Semester Marks & Other Lookups
- **Internal Marks** - `Grades` value from GMR rows where `AGR type` contains "Internal Total".
- **Total Max Marks** - `MAX Marks` value from GMR rows where `AGR type` contains "Composite ".
- **Semester Total Max Marks** - `MAX Marks` value from GMR rows where `AGR type` contains
  "Semester Total".
- **Semester Marks** *(new)* - `Grades` value from GMR rows where `AGR type` contains "Semester
  Total Marks" (e.g. `Semester Total Marks(60 )`). Uses the same PRN + Subject code matching key
  as the other lookups. It is placed immediately after **Semester Total Max Marks** and is only
  ever populated on:
  - **every Moderator row**, and
  - an **Examiner row that has no corresponding Moderator row** for that same student + module.

  If both an Examiner and a Moderator row exist for a student/module, Semester Marks is shown only
  on the Moderator row; the Examiner row is left blank. Once fetched, Semester Marks is compared to
  that row's `TotalObtainedScore`: a **match is filled light green**, a **mismatch is filled light
  red**. This colouring appears in every preview tab and in both the Formatted File and the
  Consolidated Multi-Sheet Excel - it is intentionally excluded from the Cleaned File and the ZIP
  of Output Excel Files.
- **Total Marks Obtained** - `TotalObtainedScore` (exam script) + `Internal Marks` (GMR). If both
  are blank, this is left blank; otherwise a missing side is treated as 0.
- **Grade** - derived from `Total Marks Obtained` / `Total Max Marks` as a percentage:
  O ≥ 90%, A+ ≥ 80%, A ≥ 70%, B+ ≥ 60%, B ≥ 55%, C ≥ 50%, P ≥ 40%, otherwise F.

---
### 7. Moderator & Examiner Processing Logic
- Each student's rows are grouped and sorted so Examiner always appears first, followed by
  Moderator, Reval 1, and Reval 2 (if present).
- For the **Cleaned File** / **ZIP** export, Reval rows are dropped entirely, and for any student
  who has a Moderator row, only that Moderator row is kept (the Examiner row is dropped, since the
  Moderator mark is treated as final). Students with no moderation keep their Examiner row as-is.
- The **Formatted File** and **Consolidated Multi-Sheet Excel** keep **all** UserTypes
  (Examiner, Moderator, Reval 1, Reval 2) so the full review trail is visible.

---
### 8. Change % Calculation & Moderation Analysis
`Change % = (Moderator TotalObtainedScore − Examiner TotalObtainedScore) / Denominator × 100`,
recorded only on the Moderator row. The denominator is chosen in priority order: **Semester Total
Max Marks** (from GMR) → **Total Max Marks** (from GMR) → sum of question max marks parsed from the
column headers (used only when no GMR file is available). A paper is flagged **"Eligible for 100%
Moderation"** when more than 50% of Moderator rows show an absolute Change % greater than 10%.

---
### 9. Conditional Formatting Rules & Colour Meanings
- **Question marks, Grade, Total Marks Obtained, Change %** (reviewer vs Examiner comparison):
  bright green text = increased/improved, red text = decreased/dropped, red text = a mark that
  existed for the Examiner but is missing for the reviewer.
- **Semester Marks vs TotalObtainedScore**: light green cell fill = values match, light red cell
  fill = values differ.
- **Row background** (previews only): Moderator rows are shown on a dark green background, Reval 1
  on dark blue, Reval 2 on dark purple, for quick visual scanning.
- **Moderator row fill** (Formatted File / Consolidated Excel only): light green row fill, plus a
  medium-thickness bottom border marking the end of each student's block.
- **100% Moderation pop-up**: if a paper is eligible, opening the Excel file and clicking cell A1
  shows an on-screen alert message.

---
### 10. Preview Tabs
- **📋 All Data Preview** - every row of the processed file, fully styled and colour-coded.
- **🔍 Moderation / Revaluation Cases Only** - only students who have at least one Moderator, Reval
  1, or Reval 2 row.
- **⚡ Changes Only** - only students where a reviewer's `TotalObtainedScore` actually differs from
  the Examiner's.

---
### 11. Downloadable Outputs
- **📥 Download Cleaned File** - legacy/simple export. Reval rows removed, Moderator-only rows kept
  where applicable, GMR columns (Internal Marks, Total/Semester Max Marks, **Semester Marks**,
  Total Marks Obtained, Grade, Change %) **not included**.
- **🎨 Download Formatted File** - all UserTypes, all GMR columns including **Semester Marks**,
  Change %, centered alignment, header freeze, thick per-student borders, green/red highlighting,
  Semester Marks match/mismatch fill, and the 100% Moderation pop-up if eligible.
- **📦 Download All Output Excel Files (ZIP)** - one Cleaned File per uploaded Questionwise file,
  bundled into a single ZIP (Semester Marks **not included**, matching the Cleaned File format).
- **📊 Download Consolidated Multi-Sheet Excel** - one workbook with a separate sheet per uploaded
  file (only offered when 2+ files are uploaded), each sheet formatted the same way as the
  Formatted File, including **Semester Marks**.

---
### 12. Validation Rules, Assumptions & Processing Workflow
1. Upload one or more Questionwise files (and optionally a GMR file).
2. Each file is decrypted (if needed) and loaded.
3. Rows are cleaned (blank/footer rows removed) and sorted into student order.
4. If a GMR file was supplied, Internal Marks, Total Max Marks, Semester Total Max Marks, Semester
   Marks, Total Marks Obtained, and Grade are looked up/calculated; otherwise these columns stay
   blank.
5. Change % is calculated for Moderator rows, and GMR/derived columns are reordered directly after
   `TotalObtainedScore`.
6. Summary cards, paper info, and the three preview tabs are rendered.
7. Individual Cleaned/Formatted downloads are generated per file, plus bulk ZIP and Consolidated
   Multi-Sheet downloads across all uploaded files.

**Assumptions**: a Questionwise file represents a single paper/module; `PRNNumber` uniquely
identifies a student within that file; non-numeric scores (e.g. "AB", "UFM") are treated as
absent/unfair-means and excluded from numeric comparisons rather than causing errors.
            """
        )