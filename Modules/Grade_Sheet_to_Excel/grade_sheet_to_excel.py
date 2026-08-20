import io
import re
from datetime import datetime

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# HELPER PARSING & CLEANING FUNCTIONS
# ==============================================================================

SUBJECT_CODE_RE = re.compile(r"^[A-Z]{2,10}\d{2,4}[A-Z]?$")


def degap(raw):
    """
    Some fonts used for bold summary figures (e.g. PERCENTAGE) are extracted
    by pdfplumber with a stray space between every character, e.g. '7 1 . 1 7'
    instead of '71.17'. This strips whitespace from a captured numeric span
    without touching the rest of the text.
    """
    if raw is None:
        return ""
    return re.sub(r"\s+", "", raw).strip()


def clean_val(val):
    """Strip currency-style '$' grace/adjustment markers, '~', '#' and whitespace."""
    if val is None:
        return ""
    return str(val).replace("$", "").replace("~", "").replace("#", "").strip(" \n\t")


def has_marker(val):
    """True if a cell carries a '$' marker (grace marks / marks adjusted / moderated)."""
    return "$" in str(val or "")


def extract_field(pattern, text, group=1, default=""):
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(group).strip() if m else default


def extract_numeric_field(label_pattern, text, stop_pattern, default=""):
    """
    Grabs everything between a label (e.g. 'PERCENTAGE\\s*:') and the next known
    label / end-of-line, then strips internal whitespace. Handles the
    letter-spaced-digit rendering quirk described in `degap`.
    """
    m = re.search(label_pattern + r"\s*:?\s*([0-9A-Za-z\.\-/\+\s]+?)\s*(?=" + stop_pattern + r"|\n|$)", text, re.IGNORECASE)
    return degap(m.group(1)) if m else default


# ==============================================================================
# PAGE-LEVEL PARSER
# ==============================================================================

STOP_LABELS = r"AGGREGATE MARKS|PERCENTAGE|CREDITS EARNED|SGPA|CGPA|GRADE|REMARK|TOTAL CREDITS|FINAL GRADE|OVERALL PERCENTAGE"


def parse_header_footer(text):
    """Extracts demographic info + summary/result metrics from a page's plain text."""

    name_m = re.search(r"Name of the Student\s*:?\s*(.+)", text, re.IGNORECASE)
    raw_name = name_m.group(1).strip() if name_m else ""
    # Heuristic observed in this data set: a leading '/' before the name marks
    # the record as female (mother's-name-first convention). Treat as a best
    # effort signal, not a certainty.
    gender = "Female" if raw_name.startswith("/") else "Male" if raw_name else ""
    clean_name = raw_name.lstrip("/").strip()

    sap_roll_m = re.search(
        r"Student No\.?\s*/\s*Roll No\.?\s*:?\s*([0-9]+)\s*/?\s*([A-Za-z0-9]+)",
        text, re.IGNORECASE
    )
    sap_no = sap_roll_m.group(1).strip() if sap_roll_m else ""
    roll_no = sap_roll_m.group(2).strip() if sap_roll_m else ""

    programme = extract_field(r"Programme\s*:?\s*(.+)", text)
    year_sem = extract_field(r"Year\s*&\s*Semester\s*:?\s*(.+)", text)
    academic_year = extract_field(r"Academic Year\s*:?\s*([0-9]{4}\s*-\s*[0-9]{4})", text)

    prn_no = extract_field(
        r"Uni\.?\s*PRN\s*/?\s*Reg\.?\s*No\s*\.?\s*:?\s*([A-Za-z0-9]+)", text
    )

    exam_session = extract_field(
        r"Month\s*&\s*Year of Exam\s*:?\s*([A-Za-z]+,?\s*\d{4})", text
    )

    apaar_id = extract_field(r"(?:APAAR ID|ABC ID)\s*:?\s*([0-9]{10,14})", text)

    # ---- Summary / result metrics (footer line) ----
    agg_raw = extract_numeric_field(r"AGGREGATE MARKS", text, STOP_LABELS)
    marks_obt, max_marks = "", ""
    if "/" in agg_raw:
        parts = agg_raw.split("/")
        marks_obt, max_marks = parts[0].strip(), parts[1].strip()

    percentage = extract_numeric_field(r"\bPERCENTAGE", text, STOP_LABELS)
    sem_credits = extract_numeric_field(r"CREDITS EARNED", text, STOP_LABELS)
    sem_sgpa = extract_numeric_field(r"\bSGPA", text, STOP_LABELS)
    cgpa = extract_numeric_field(r"\bCGPA", text, STOP_LABELS)
    total_credits = extract_numeric_field(r"TOTAL CREDITS EARNED", text, STOP_LABELS)
    final_grade = extract_numeric_field(r"FINAL GRADE", text, STOP_LABELS)
    overall_pct = extract_numeric_field(r"OVERALL PERCENTAGE", text, STOP_LABELS)

    sem_grade_m = re.search(r"(?<!FINAL )\bGRADE\s*:\s*([A-O\+\-]+|--)", text, re.IGNORECASE)
    sem_grade = sem_grade_m.group(1).strip() if sem_grade_m else ""

    # Remark is always a single word (SUCCESSFUL/UNSUCCESSFUL); on consolidated
    # multi-semester cards it shares a line with "TOTAL CREDITS EARNED ..." so
    # we must not let the capture run past the first word. A trailing '$'
    # marker (rare) is captured separately.
    remark_m = re.search(r"REMARK\s*:?\s*([A-Za-z]+)(\${0,2})", text, re.IGNORECASE)
    remark = remark_m.group(1).strip() if remark_m else ""
    remark_marked = bool(remark_m and remark_m.group(2))

    return {
        "SAP No": sap_no,
        "Roll No": roll_no,
        "Candidate Full Name": clean_name,
        "Gender": gender,
        "Uni PRN": prn_no,
        "APAAR ID": apaar_id,
        "Programme": programme,
        "Year & Semester": year_sem,
        "Academic Year": academic_year,
        "Exam Session": exam_session,
        "Marks Obtained": marks_obt,
        "Max Total Marks": max_marks,
        "Percentage": percentage,
        "Sem Credits Earned": sem_credits,
        "SGPA": sem_sgpa,
        "Sem Grade": sem_grade,
        "Total Credits Earned": total_credits,
        "CGPA": cgpa,
        "Final Grade": final_grade,
        "Overall Percentage": overall_pct,
        "Result Remark": remark,
        "Remark Adjustment Marked": "Yes" if remark_marked else "No",
    }


def parse_subject_rows_from_table(table):
    """Primary path: parse the subject grid using pdfplumber's detected table."""
    subjects = []
    for row in table:
        if not row or not row[0]:
            continue
        code = clean_val(row[0]).replace("\n", " ")
        if not SUBJECT_CODE_RE.match(code):
            continue  # skips header row / stray rows

        cells = [c if c is not None else "" for c in row]
        # Pad defensively in case a row has fewer columns than expected
        while len(cells) < 9:
            cells.append("")

        title = str(cells[1]).replace("\n", " ").strip()
        credits = clean_val(cells[2])
        ca_max = clean_val(cells[3])
        ca_obt_raw = cells[4]
        ese_max = clean_val(cells[5])
        ese_obt_raw = cells[6]
        total_raw = cells[7]
        grade_raw = str(cells[8]).strip()

        # Grade cell sometimes carries a trailing '$' note, e.g. "B+ $"
        grade = grade_raw.replace("$", "").strip()
        if grade == "0":  # OCR/extraction sometimes renders 'O' grade as zero
            grade = "O"

        subjects.append({
            "Subject Code": code,
            "Course Title": title,
            "Credits": credits,
            "CA Max": ca_max,
            "CA Marks": clean_val(ca_obt_raw),
            "ESE Max": ese_max,
            "ESE Marks": clean_val(ese_obt_raw),
            "Total Marks": clean_val(total_raw),
            "Grade": grade,
            "Grace/Adjustment Marked": "Yes" if (
                has_marker(ca_obt_raw) or has_marker(ese_obt_raw)
                or has_marker(total_raw) or has_marker(grade_raw)
            ) else "No",
        })
    return subjects


LINE_SUBJECT_RE = re.compile(
    r"^([A-Z]{2,10}\d{2,4}[A-Z]?)\s+(.+?)\s+(\d+\.\d+)\s+"
    r"(\d+|--)\s+([\d\$]+|--)\s+(\d+|--)\s+([\d\$]+|--)\s+([\d\$]+)\s+"
    r"([A-O\+\-]+(?:\s*\$)?)\s*$"
)


def parse_subject_rows_from_text(text):
    """Fallback path used only if pdfplumber can't detect a table grid on the page."""
    subjects = []
    for line in text.splitlines():
        m = LINE_SUBJECT_RE.match(line.strip())
        if not m:
            continue
        code, title, credits, ca_max, ca_obt, ese_max, ese_obt, total, grade_raw = m.groups()
        grade = grade_raw.replace("$", "").strip()
        subjects.append({
            "Subject Code": code,
            "Course Title": title.strip(),
            "Credits": credits,
            "CA Max": ca_max,
            "CA Marks": clean_val(ca_obt),
            "ESE Max": ese_max,
            "ESE Marks": clean_val(ese_obt),
            "Total Marks": clean_val(total),
            "Grade": grade,
            "Grace/Adjustment Marked": "Yes" if "$" in (ca_obt + ese_obt + total + grade_raw) else "No",
        })
    return subjects


def parse_semester_history(tables):
    """
    Some grade cards (final-year consolidated statements) carry a second
    small table breaking down Credits Earned / SGPA per semester (I, II, ...)
    plus a merged 'OVERALL PERCENTAGE: xx.xx' cell. Returns a flat dict like
    {"Sem I Credits": "22.00", "Sem I SGPA": "7.09", ...} plus overall %.
    Returns ({}, "") if this table isn't present on the page.
    """
    history = {}
    overall_pct_hist = ""
    for t in tables:
        if not t or not t[0] or not t[0][0]:
            continue
        header = [c.strip() if c else "" for c in t[0]]
        if header[0].upper() != "SEMESTER":
            continue

        sem_labels = header[1:]
        credits_row = next(
            (r for r in t if r and r[0] and "CREDIT" in r[0].strip().upper()), None
        )
        sgpa_row = next(
            (r for r in t if r and r[0] and r[0].strip().upper() == "SGPA"), None
        )

        for idx, sem in enumerate(sem_labels):
            sem = (sem or "").strip()
            if not sem:
                continue
            col = idx + 1
            credits_val = clean_val(credits_row[col]) if credits_row and col < len(credits_row) else ""
            sgpa_val = clean_val(sgpa_row[col]) if sgpa_row and col < len(sgpa_row) else ""
            history[f"Sem {sem} Credits"] = credits_val
            history[f"Sem {sem} SGPA"] = sgpa_val

        for r in t:
            if r and r[0] and "OVERALL PERCENTAGE" in r[0].upper():
                m = re.search(r"OVERALL PERCENTAGE\s*:?\s*([\d\.\-\s]+)", r[0], re.IGNORECASE)
                if m:
                    overall_pct_hist = degap(m.group(1))
    return history, overall_pct_hist


def parse_grade_sheet_page(page, p_no, source_file):
    """Parses a single student grade-sheet page into a structured profile dict."""
    text = page.extract_text(x_tolerance=2) or ""
    profile = parse_header_footer(text)

    tables = page.extract_tables()
    subjects = []
    used_fallback = False
    if tables:
        subjects = parse_subject_rows_from_table(tables[0])
    if not subjects:
        subjects = parse_subject_rows_from_text(text)
        used_fallback = bool(subjects)

    sem_history, overall_pct_hist = parse_semester_history(tables)
    # The per-line "OVERALL PERCENTAGE:" match (from parse_header_footer) is
    # usually enough, but fall back to the table-derived value if that was
    # blank (defensive - covers layout variants where it isn't a standalone line).
    if not profile.get("Overall Percentage") and overall_pct_hist:
        profile["Overall Percentage"] = overall_pct_hist

    profile["Page No"] = p_no
    profile["Source File"] = source_file
    profile["Subjects"] = subjects
    profile["Semester History"] = sem_history
    profile["Used Fallback Parser"] = used_fallback
    return profile


# ==============================================================================
# DATAFRAME BUILDERS
# ==============================================================================

def build_datasets(all_student_profiles):
    """Constructs Consolidated, Detailed, and Subject-Wise DataFrames from student profiles."""
    consolidated_rows = []
    detailed_rows = []
    subject_wise_rows = []

    # Union of all semester-history columns seen across the batch (only
    # populated for consolidated/final-year grade cards that carry a
    # per-semester SGPA breakdown table).
    all_hist_cols = sorted({
        k for p in all_student_profiles for k in p.get("Semester History", {}).keys()
    })

    for p in all_student_profiles:
        base_id = {
            "Source File": p["Source File"],
            "SAP No": p["SAP No"],
            "Roll No": p["Roll No"],
            "Candidate Full Name": p["Candidate Full Name"],
            "Gender": p["Gender"],
            "Uni PRN": p["Uni PRN"],
            "APAAR ID": p["APAAR ID"],
        }

        # 1. Consolidated result row
        con_row = dict(base_id)
        con_row.update({
            "Programme": p["Programme"],
            "Semester": p["Year & Semester"],
            "Academic Year": p["Academic Year"],
            "Exam Session": p["Exam Session"],
            "Marks Obtained": p["Marks Obtained"],
            "Max Marks": p["Max Total Marks"],
            "Percentage": p["Percentage"],
            "Credits Earned": p["Sem Credits Earned"],
            "SGPA": p["SGPA"],
            "Sem Grade": p["Sem Grade"],
            "Total Credits": p["Total Credits Earned"],
            "CGPA": p["CGPA"],
            "Final Grade": p["Final Grade"],
            "Overall Percentage": p["Overall Percentage"],
            "Result Remark": p["Result Remark"],
            "Remark Adjustment Marked": p.get("Remark Adjustment Marked", "No"),
            "Subjects Detected": len(p["Subjects"]),
        })
        hist = p.get("Semester History", {})
        for col in all_hist_cols:
            con_row[col] = hist.get(col, "")
        consolidated_rows.append(con_row)

        # 2. Detailed horizontal (wide) row - one row per student, subjects across columns
        det_row = dict(base_id)
        for idx, sub in enumerate(p["Subjects"]):
            pre = f"Sub{idx + 1}"
            det_row[f"{pre} Code"] = sub["Subject Code"]
            det_row[f"{pre} Title"] = sub["Course Title"]
            det_row[f"{pre} Credits"] = sub["Credits"]
            det_row[f"{pre} CA Marks"] = sub["CA Marks"]
            det_row[f"{pre} ESE Marks"] = sub["ESE Marks"]
            det_row[f"{pre} Total Marks"] = sub["Total Marks"]
            det_row[f"{pre} Grade"] = sub["Grade"]

            # 3. Normalized subject-wise row (one row per student per subject)
            subject_wise_rows.append({
                **base_id,
                "Subject Code": sub["Subject Code"],
                "Course Title": sub["Course Title"],
                "Credits": sub["Credits"],
                "CA Max": sub["CA Max"],
                "CA Marks": sub["CA Marks"],
                "ESE Max": sub["ESE Max"],
                "ESE Marks": sub["ESE Marks"],
                "Total Marks": sub["Total Marks"],
                "Grade": sub["Grade"],
                "Grace/Adjustment Marked": sub["Grace/Adjustment Marked"],
            })

        det_row["Marks Obtained"] = p["Marks Obtained"]
        det_row["Max Total Marks"] = p["Max Total Marks"]
        det_row["Percentage"] = p["Percentage"]
        det_row["SGPA"] = p["SGPA"]
        det_row["CGPA"] = p["CGPA"]
        det_row["Sem Grade"] = p["Sem Grade"]
        det_row["Result Remark"] = p["Result Remark"]
        detailed_rows.append(det_row)

    return (
        pd.DataFrame(consolidated_rows),
        pd.DataFrame(detailed_rows),
        pd.DataFrame(subject_wise_rows),
    )


# ==============================================================================
# EXCEL EXPORT
# ==============================================================================

def save_to_formatted_excel(sheets_dict):
    """Exports structured data to Excel with color styling, borders, freeze panes, and highlights."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            fail_font = Font(name="Calibri", size=11, color="9C0006", bold=True)
            grace_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            thin_side = Side(style="thin", color="D9D9D9")
            thick_right_side = Side(style="medium", color="1F4E79")

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

            if sheet_name == "Detailed_Result":
                sub_grade_cols = [
                    idx + 1 for idx, col in enumerate(df.columns)
                    if str(col).endswith(" Grade") and not str(col).startswith("Final")
                    and not str(col).startswith("Sem")
                ]
                sub_blocks = [(g_col, g_col - 6, g_col) for g_col in sub_grade_cols]

                for row_idx in range(2, ws.max_row + 1):
                    failed_cols = set()
                    for g_col, start_c, end_c in sub_blocks:
                        val = str(ws.cell(row=row_idx, column=g_col).value or "").strip()
                        if val == "F":
                            for c in range(max(1, start_c), end_c + 1):
                                failed_cols.add(c)

                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        r_side = thick_right_side if col_idx in sub_grade_cols else thin_side
                        cell.border = Border(left=thin_side, right=r_side, top=thin_side, bottom=thin_side)
                        cell.alignment = Alignment(vertical="center", horizontal="center")
                        if col_idx in failed_cols:
                            cell.fill = fail_fill
                            cell.font = fail_font

                for g_col in sub_grade_cols:
                    ws.cell(row=1, column=g_col).border = Border(
                        left=thin_side, right=thick_right_side, top=thin_side, bottom=thin_side
                    )
            else:
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                grade_col_idx = None
                grace_col_idx = None
                for idx, col in enumerate(df.columns):
                    if col in ("Grade", "Sem Grade"):
                        grade_col_idx = idx + 1
                    if col == "Grace/Adjustment Marked":
                        grace_col_idx = idx + 1

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center", horizontal="center")
                        val = str(cell.value).strip() if cell.value is not None else ""
                        if val in ("F", "UNSUCCESSFUL", "Unsuccessful"):
                            cell.fill = fail_fill
                            cell.font = fail_font
                        elif grace_col_idx and cell.column == grace_col_idx and val == "Yes":
                            cell.fill = grace_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 10)

    output.seek(0)
    return output


# ==============================================================================
# STREAMLIT UI APPLICATION
# ==============================================================================

def show():
    st.title("🎓 Student Grade Sheet Extractor")
    st.caption("Parses multi-page student grade sheet PDFs into structured, analysis-ready Excel sheets.")

    uploaded_files = st.file_uploader(
        "Upload Student Grade Sheet PDF(s)",
        type=["pdf"],
        accept_multiple_files=True
    )

    if not uploaded_files:
        st.info("Upload one or more grade sheet PDFs (each page = one student) to get started.")
        return

    all_profiles = []
    logs = []

    progress = st.progress(0.0, text="Starting extraction...")
    total_pages_all = 0

    file_page_counts = []
    for uploaded_file in uploaded_files:
        with pdfplumber.open(uploaded_file) as pdf:
            file_page_counts.append(len(pdf.pages))
    total_pages_all = sum(file_page_counts) or 1

    pages_done = 0
    for f_idx, uploaded_file in enumerate(uploaded_files):
        with pdfplumber.open(uploaded_file) as pdf:
            for p_idx, page in enumerate(pdf.pages):
                p_num = p_idx + 1
                profile = parse_grade_sheet_page(page, p_num, uploaded_file.name)
                all_profiles.append(profile)

                status = "Processed"
                if not profile["Candidate Full Name"]:
                    status = "⚠️ Name not found"
                elif not profile["Subjects"]:
                    status = "⚠️ No subjects detected"
                elif profile["Used Fallback Parser"]:
                    status = "Processed (fallback parser)"

                logs.append({
                    "Source File": uploaded_file.name,
                    "Page No": p_num,
                    "SAP No": profile["SAP No"],
                    "Student Name": profile["Candidate Full Name"],
                    "Subjects Detected": len(profile["Subjects"]),
                    "Status": status,
                })

                pages_done += 1
                progress.progress(
                    pages_done / total_pages_all,
                    text=f"Extracting {uploaded_file.name} — page {p_num}/{file_page_counts[f_idx]}"
                )

    progress.empty()

    df_con, df_det, df_sub = build_datasets(all_profiles)
    df_logs = pd.DataFrame(logs)

    n_students = len(all_profiles)
    n_pass = (df_con["Result Remark"].str.upper() == "SUCCESSFUL").sum() if "Result Remark" in df_con else 0
    n_warn = sum(1 for l in logs if "⚠️" in l["Status"])

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Students Parsed", n_students)
    c2.metric("Successful Results", int(n_pass))
    c3.metric("Unsuccessful / Other", n_students - int(n_pass))
    c4.metric("Pages Needing Review", n_warn)

    if n_warn:
        st.warning(f"{n_warn} page(s) had extraction issues — check the Extraction Log tab.")

    t1, t2, t3, t4, t5 = st.tabs([
        "📊 Consolidated Results",
        "📑 Detailed Subject Matrix",
        "📈 Subject-Wise (Normalized)",
        "📉 Analytics",
        "📝 Extraction Log",
    ])

    with t1:
        search = st.text_input("Search by name / roll no / SAP no", key="search_con")
        view = df_con
        if search:
            mask = df_con.apply(lambda r: search.lower() in " ".join(str(v) for v in r).lower(), axis=1)
            view = df_con[mask]
        st.dataframe(view, use_container_width=True)

    with t2:
        st.dataframe(df_det, use_container_width=True)

    with t3:
        st.dataframe(df_sub, use_container_width=True)

    with t4:
        if not df_con.empty:
            pct_numeric = pd.to_numeric(df_con["Percentage"], errors="coerce")
            colA, colB = st.columns(2)
            with colA:
                st.write("**Class Average %:**", round(pct_numeric.mean(), 2) if pct_numeric.notna().any() else "N/A")
                st.write("**Topper %:**", pct_numeric.max() if pct_numeric.notna().any() else "N/A")
            with colB:
                if "Sem Grade" in df_con:
                    st.write("**Overall Grade Distribution**")
                    st.bar_chart(df_con["Sem Grade"].value_counts())
            if not df_sub.empty and "Grade" in df_sub:
                st.write("**Subject-Wise Grade Distribution**")
                st.bar_chart(df_sub["Grade"].value_counts())
            if not df_sub.empty:
                fail_rows = df_sub[df_sub["Grade"] == "F"]
                if not fail_rows.empty:
                    st.write("**Subjects with most fails**")
                    st.bar_chart(fail_rows["Course Title"].value_counts())
        else:
            st.info("No data to analyze yet.")

    with t5:
        st.dataframe(df_logs, use_container_width=True)

    sheets_to_save = {
        "Consolidated_Results": df_con,
        "Detailed_Result": df_det,
        "Student_Subject_Wise": df_sub,
        "Extraction_Log": df_logs,
    }
    excel_data = save_to_formatted_excel(sheets_to_save)

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    export_name = f"GradeSheets_Combined_{stamp}.xlsx" if len(uploaded_files) > 1 else \
        f"{uploaded_files[0].name.rsplit('.', 1)[0]}_GradeSheets.xlsx"

    st.download_button(
        label=f"⬇️ Download Structured Excel ({export_name})",
        data=excel_data,
        file_name=export_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    st.set_page_config(
        page_title="Grade Sheet PDF Extractor",
        page_icon="🎓",
        layout="wide"
    )
    show()