import io
import re
import time
from datetime import datetime
from functools import lru_cache

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st


# ==========================================
# OPTIONAL DEPENDENCY: deep_translator
# ==========================================
_translator = None
_translator_unavailable = False


def _get_translator():
    """Lazily create and cache the GoogleTranslator instance."""
    global _translator, _translator_unavailable
    if _translator is not None:
        return _translator
    if _translator_unavailable:
        return None
    try:
        from deep_translator import GoogleTranslator
        _translator = GoogleTranslator(source="en", target="mr")
    except Exception:
        _translator_unavailable = True
        _translator = None
    return _translator


# ==========================================
# 1. REQUIRED OUTPUT COLUMNS SPECIFICATION
# ==========================================
TARGET_COLUMNS = [
    "LOTNO", "CONVID", "FACULTY", "PRNERN", "APAAR ID", "PROGTYPE", "APPL_NO", "SEAT_NO",
    "COLL_NO", "COLL_NAME", "COLL_NAMEM", "STUDLASTNAME", "STUDFIRSTNAME",
    "STUDMIDDDLENAME", "STUDMOTHERNAME", "NAME", "NAME_MARAT", "SEX", "ABBR",
    "CLASS", "CGPA", "MCLASS", "SUB1", "SUB1_NAME", "SUB1_NAMEM", "SUB2", "SUB2_NAME",
    "SUB2_NAMEM", "DEGNM", "MDEGNM", "SUBDEGNM", "MSUBDEGNM", "MONTH", "MMONTH",
    "CGPA Source File"
]

_CGPA_SOURCE_FILE_INTERNAL_COL = "__CGPA_SOURCE_FILE__"

MKCL_SOURCE_MAPPING = {
    "Student Number": "SEAT_NO",
    "Student Name": "NAME",
    "Gender": "SEX",
    "Final Grade": "CLASS",
    "CGPA": "CGPA",
    "Highest month of passing": "MONTH"
}


# ==========================================
# 2. ROBUST DATA CLEANING & MATCHING HELPERS
# ==========================================
def sanitize_key(val) -> str:
    if pd.isna(val) or val is None:
        return ""
    
    clean_str = str(val).replace("\xa0", " ").strip()
    
    if "." in clean_str:
        try:
            float_val = float(clean_str)
            if float_val.is_integer():
                return str(int(float_val))
        except ValueError:
            pass
        clean_str = clean_str.split(".")[0].strip()

    return clean_str.upper()


# ==========================================
# NAME_MARAT & MMONTH TRANSLITERATION
# ==========================================
@lru_cache(maxsize=4096)
def _translate_cached(clean_name: str) -> str:
    translator = _get_translator()
    if translator is None:
        return ""

    max_attempts = 6
    backoff_seconds = 1.0
    for attempt in range(1, max_attempts + 1):
        try:
            result = translator.translate(clean_name)
            if result and str(result).strip():
                return str(result).strip()
        except Exception:
            pass
        if attempt < max_attempts:
            time.sleep(backoff_seconds)
            backoff_seconds = min(backoff_seconds * 2, 15.0)
    return ""


_OFFLINE_VOWELS_INDEPENDENT = [
    ("aa", "आ"), ("ee", "ई"), ("oo", "ऊ"), ("ai", "ऐ"), ("au", "औ"),
    ("a", "अ"), ("i", "इ"), ("u", "उ"), ("e", "ए"), ("o", "ओ"),
]
_OFFLINE_VOWELS_MATRA = [
    ("aa", "ा"), ("ee", "ी"), ("oo", "ू"), ("ai", "ै"), ("au", "ौ"),
    ("a", ""), ("i", "ि"), ("u", "ु"), ("e", "े"), ("o", "ो"),
]
_OFFLINE_CONSONANTS = [
    ("kh", "ख"), ("gh", "घ"), ("chh", "छ"), ("ch", "च"), ("jh", "झ"),
    ("th", "थ"), ("dh", "ध"), ("ph", "फ"), ("bh", "भ"), ("sh", "श"),
    ("ng", "ङ"), ("ny", "ञ"),
    ("k", "क"), ("g", "ग"), ("j", "ज"), ("t", "त"), ("d", "द"),
    ("n", "न"), ("p", "प"), ("f", "फ"), ("b", "ब"), ("m", "म"),
    ("y", "य"), ("r", "र"), ("l", "ल"), ("v", "व"), ("w", "व"),
    ("s", "स"), ("h", "ह"), ("z", "ज़"), ("x", "क्स"), ("q", "क"),
    ("c", "क"),
]


def _offline_transliterate_word(word: str) -> str:
    w = word.lower()
    n = len(w)
    i = 0
    out = []
    pending_consonant = None

    def flush_pending():
        if pending_consonant is not None:
            out.append(pending_consonant)

    while i < n:
        matched = False

        for length in (3, 2, 1):
            chunk = w[i:i + length]
            for latin, dev in _OFFLINE_CONSONANTS:
                if latin == chunk:
                    flush_pending()
                    pending_consonant = dev
                    i += length
                    matched = True
                    break
            if matched:
                break
        if matched:
            continue

        for length in (2, 1):
            chunk = w[i:i + length]
            for latin, dev in _OFFLINE_VOWELS_MATRA:
                if latin == chunk:
                    if pending_consonant is not None:
                        out.append(pending_consonant + dev)
                        pending_consonant = None
                    else:
                        for latin2, dev2 in _OFFLINE_VOWELS_INDEPENDENT:
                            if latin2 == chunk:
                                out.append(dev2)
                                break
                    i += length
                    matched = True
                    break
            if matched:
                break
        if matched:
            continue

        i += 1

    if pending_consonant is not None:
        out.append(pending_consonant + "्")

    return "".join(out)


@lru_cache(maxsize=4096)
def _offline_transliterate_cached(clean_name: str) -> str:
    words = [w for w in re.split(r"\s+", clean_name.strip()) if w]
    converted = []
    for word in words:
        alpha_only = re.sub(r"[^A-Za-z]", "", word)
        if not alpha_only:
            converted.append(word)
            continue
        converted.append(_offline_transliterate_word(alpha_only))
    return " ".join(converted)


def transliterate_name_to_marathi(name_str: str) -> str:
    if pd.isna(name_str) or not str(name_str).strip():
        return ""
    clean_name = str(name_str).strip()

    online_result = _translate_cached(clean_name)
    if online_result:
        return online_result

    return _offline_transliterate_cached(clean_name)


def derive_faculty(degnm: str) -> str:
    if pd.isna(degnm) or not str(degnm).strip():
        return ""
    
    degnm_str = str(degnm).strip()
    
    commerce_pattern = r'(?i)\b(commerce|accounting|management|finance|financial|marketing|business|economics|b\.?com|m\.?com)\b'
    if re.search(commerce_pattern, degnm_str):
        return "Commerce & Management"
        
    science_pattern = r'(?i)\b(science|artificial|intelligent|biochemistry|data|technology|b\.?sc\.?|m\.?sc\.?)\b'
    if re.search(science_pattern, degnm_str):
        return "Science & Technology"
        
    Humanities_pattern = r'(?i)\b(arts|entertainment|media|multimedia|film|television|advertising|b\.?a\.?|m\.?a\.?)\b'
    if re.search(Humanities_pattern, degnm_str):
        return "Humanities"
        
    return ""


def derive_subdegnm(degnm: str) -> str:
    if pd.isna(degnm) or not str(degnm).strip():
        return ""
    
    degnm_str = str(degnm).strip()
    if re.search(r'\bMASTERS?\b', degnm_str, re.IGNORECASE) or degnm_str.upper().startswith('M'):
        return "Two Year Degree Course"
    if re.search(r'\bBACHELORS?\b', degnm_str, re.IGNORECASE) or degnm_str.upper().startswith('B'):
        return "Three Year Degree Course"
        
    return ""


def map_gender(val) -> str:
    if pd.isna(val):
        return ""
    
    val_str = str(val).strip().lower()
    if val_str in ["male", "m", "1"]:
        return "1"
    elif val_str in ["female", "f", "2"]:
        return "2"
    return str(val)


MARATHI_DIGIT_MAP = {
    "0": "०", "1": "१", "2": "२", "3": "३", "4": "४",
    "5": "५", "6": "६", "7": "७", "8": "८", "9": "९",
}


def convert_digits_to_marathi(val) -> str:
    """
    Convert English digits in a string into Marathi (Devanagari) numerals.
    e.g. MARCH 2026 / मार्च 2026 -> मार्च २०२६
    """
    if pd.isna(val) or not str(val).strip():
        return ""

    val_str = str(val).strip()
    return "".join(MARATHI_DIGIT_MAP.get(ch, ch) for ch in val_str)


def format_class_grade(val) -> str:
    if pd.isna(val) or not str(val).strip():
        return ""
    
    val_str = str(val).strip()
    clean_val = re.sub(r"(?i)\bgrade\b", "", val_str).strip(" '\"")
    if clean_val:
        return f"'{clean_val}' Grade"
    return ""


# ==========================================
# 3. DYNAMIC MASTER LOOKUP MAP CREATION
# ==========================================
def find_priority_column(columns, keyword_groups):
    cols_lower = {c: c.lower() for c in columns}
    for keywords in keyword_groups:
        for c in columns:
            cl = cols_lower[c]
            if any(k in cl for k in keywords):
                return c
    return None


def build_master_lookup(master_files, progress_callback=None):
    """
    Reads all uploaded Student Master Data files across all sheets.
    Builds Student Number -> {"program": ..., "abc_id": ...} dictionary.
    """
    student_lookup = {}
    master_records_count = 0
    total_master_files = len(master_files) or 1

    for file_idx, file in enumerate(master_files, start=1):
        excel_obj = pd.ExcelFile(file)
        for sheet_name in excel_obj.sheet_names:
            df_master = pd.read_excel(excel_obj, sheet_name=sheet_name, dtype=str)
            master_records_count += len(df_master)

            df_master.columns = [str(c).replace("\xa0", " ").strip() for c in df_master.columns]

            stu_col = find_priority_column(
                df_master.columns,
                [["student number"], ["student_number"]]
            )
            if not stu_col and len(df_master.columns) > 0:
                stu_col = df_master.columns[0]

            prog_col = find_priority_column(
                df_master.columns,
                [
                    ["program name"],
                    ["program_name"],
                    ["programme name"],
                    ["degree name"],
                ]
            )
            if not prog_col and len(df_master.columns) >= 204:
                prog_col = df_master.columns[203]

            abc_col = find_priority_column(
                df_master.columns,
                [
                    ["abc id"], ["abc_id"], ["abcid"],
                    ["apaar id"], ["apaar_id"], ["apaarid"],
                    ["abc id/apaar id"], ["abc"]
                ]
            )

            if stu_col and prog_col:
                for _, row in df_master.iterrows():
                    key = sanitize_key(row[stu_col])
                    prog_val = row[prog_col]
                    abc_val = str(row[abc_col]).strip() if (abc_col and pd.notna(row[abc_col])) else ""
                    if abc_val.lower() == "nan":
                        abc_val = ""

                    if key and pd.notna(prog_val) and str(prog_val).strip():
                        student_lookup[key] = {
                            "program": str(prog_val).strip(),
                            "abc_id": abc_val
                        }

        if progress_callback:
            progress_callback(file_idx / total_master_files)

    return student_lookup, master_records_count


# ==========================================
# 4. CORE ETL & MERGE PROCESSING
# ==========================================
def process_data(mkcl_files, student_lookup, college_choice, progress_callback=None):
    mkcl_dfs = []
    total_mkcl_raw_rows = 0
    total_mkcl_files = len(mkcl_files) or 1

    for file_idx, file in enumerate(mkcl_files, start=1):
        excel_obj = pd.ExcelFile(file)
        for sheet_name in excel_obj.sheet_names:
            df = pd.read_excel(excel_obj, sheet_name=sheet_name, dtype=str)
            total_mkcl_raw_rows += len(df)

            df.columns = [str(c).replace("\xa0", " ").strip() for c in df.columns]

            source_file_name = getattr(file, "name", str(file))
            df[_CGPA_SOURCE_FILE_INTERNAL_COL] = source_file_name

            # 1. Convocation Number filter: must be blank/null/empty
            conv_col = next((c for c in df.columns if c.lower() == "convocation number"), None)
            if conv_col:
                df = df[df[conv_col].isna() | (df[conv_col].astype(str).str.strip() == "")]

            # 2. CGPA filter: must contain a value
            cgpa_col = next((c for c in df.columns if c.lower() == "cgpa"), None)
            if cgpa_col:
                df = df[df[cgpa_col].notna() & (df[cgpa_col].astype(str).str.strip() != "")]
            else:
                df = df.iloc[0:0]

            # 3. GPA Filtering (Aware of Lateral Admission & UG vs PG programs)
            student_re_col = next((c for c in df.columns if c.lower() == "student re"), None)
            stu_num_col = next((c for c in df.columns if "student number" in c.lower()), None)
            if not stu_num_col and len(df.columns) > 0:
                stu_num_col = df.columns[0]

            has_sem5_col = any("sem 5" in c.lower() for c in df.columns)
            has_sem6_col = any("sem 6" in c.lower() for c in df.columns)

            sem3_gpa_col = next((c for c in df.columns if c.lower() == "sem 3_gpa"), None)
            sem4_gpa_col = next((c for c in df.columns if c.lower() == "sem 4_gpa"), None)
            sem5_gpa_col = next((c for c in df.columns if c.lower() == "sem 5_gpa"), None)
            sem6_gpa_col = next((c for c in df.columns if c.lower() == "sem 6_gpa"), None)

            target_sem_gpas = [c for c in ["SEM 1_GPA", "SEM 2_GPA", "SEM 3_GPA", "SEM 4_GPA", "SEM 5_GPA", "SEM 6_GPA"] if c in df.columns]

            def row_passes_gpa_filter(row):
                stu_key = sanitize_key(row[stu_num_col]) if stu_num_col else ""
                student_info = student_lookup.get(stu_key, {})
                prog_name = student_info.get("program", "") if isinstance(student_info, dict) else str(student_info)

                # Post Graduate (PG) programs start with 'M' or have only 4 semesters in the report
                is_pg = prog_name.strip().upper().startswith("M") or (not has_sem5_col and not has_sem6_col)

                re_val = str(row[student_re_col]).strip() if (student_re_col and pd.notna(row[student_re_col])) else ""
                is_lateral = "lateral admission" in re_val.lower()

                if is_lateral:
                    if is_pg:
                        # PG Lateral Admission (Admitted directly in Semester 3):
                        # SEM 3_GPA and SEM 4_GPA must be populated
                        val3 = str(row[sem3_gpa_col]).strip() if (sem3_gpa_col and pd.notna(row[sem3_gpa_col])) else ""
                        val4 = str(row[sem4_gpa_col]).strip() if (sem4_gpa_col and pd.notna(row[sem4_gpa_col])) else ""
                        
                        if not val3 or not val4 or val3.lower() == "nan" or val4.lower() == "nan":
                            return False
                        return True
                    else:
                        # UG Lateral Admission (Admitted directly in Semester 3 or Semester 5):
                        # 1) SEM 5_GPA and SEM 6_GPA must be populated
                        val5 = str(row[sem5_gpa_col]).strip() if (sem5_gpa_col and pd.notna(row[sem5_gpa_col])) else ""
                        val6 = str(row[sem6_gpa_col]).strip() if (sem6_gpa_col and pd.notna(row[sem6_gpa_col])) else ""
                        if not val5 or not val6 or val5.lower() == "nan" or val6.lower() == "nan":
                            return False

                        # 2) Check Sem 3 & Sem 4 GPAs:
                        val3 = str(row[sem3_gpa_col]).strip() if (sem3_gpa_col and pd.notna(row[sem3_gpa_col])) else ""
                        val4 = str(row[sem4_gpa_col]).strip() if (sem4_gpa_col and pd.notna(row[sem4_gpa_col])) else ""
                        
                        if val3.lower() == "nan": val3 = ""
                        if val4.lower() == "nan": val4 = ""

                        # If one of Sem 3 or Sem 4 GPA is present, both must be present
                        if bool(val3) != bool(val4):
                            return False

                        return True
                else:
                    # Standard student: All present SEM X_GPA columns must contain a value
                    for sem_gpa_name in target_sem_gpas:
                        sem_col = next((c for c in df.columns if c.lower() == sem_gpa_name.lower()), None)
                        if sem_col:
                            val = str(row[sem_col]).strip() if pd.notna(row[sem_col]) else ""
                            if not val or val.lower() == "nan":
                                return False
                    return True

            df = df[df.apply(row_passes_gpa_filter, axis=1)]

            if not df.empty:
                mkcl_dfs.append(df)

        if progress_callback:
            progress_callback("read", file_idx / total_mkcl_files)

    if not mkcl_dfs:
        if progress_callback:
            progress_callback("match", 1.0)
        empty_df = pd.DataFrame(columns=TARGET_COLUMNS)
        return empty_df, pd.DataFrame(), total_mkcl_raw_rows, 0, 0, 0

    merged_mkcl = pd.concat(mkcl_dfs, ignore_index=True)
    total_filtered_rows = len(merged_mkcl)

    matched_count = 0
    unmatched_count = 0
    deg_names = []
    apaar_ids = []
    unmatched_rows = []

    stu_num_col = next((c for c in merged_mkcl.columns if "student number" in c.lower()), None)
    if not stu_num_col and len(merged_mkcl.columns) > 0:
        stu_num_col = merged_mkcl.columns[0]

    total_match_rows = len(merged_mkcl) or 1
    progress_interval = max(1, total_match_rows // 50)

    for row_idx, (idx, row) in enumerate(merged_mkcl.iterrows(), start=1):
        stu_id = sanitize_key(row[stu_num_col]) if stu_num_col else ""
        student_info = student_lookup.get(stu_id)

        if student_info:
            matched_count += 1
            deg_names.append(student_info.get("program", ""))
            apaar_ids.append(student_info.get("abc_id", ""))
        else:
            unmatched_count += 1
            deg_names.append("")
            apaar_ids.append("")
            unmatched_row_dict = row.to_dict()
            unmatched_row_dict.pop(_CGPA_SOURCE_FILE_INTERNAL_COL, None)
            unmatched_rows.append(unmatched_row_dict)

        if progress_callback and (row_idx % progress_interval == 0 or row_idx == total_match_rows):
            progress_callback("match", row_idx / total_match_rows)

    merged_mkcl["DEGNM"] = deg_names
    merged_mkcl["APAAR ID"] = apaar_ids

    out_df = pd.DataFrame(columns=TARGET_COLUMNS)

    for src_col, target_col in MKCL_SOURCE_MAPPING.items():
        matched_src = next((c for c in merged_mkcl.columns if c.lower() == src_col.lower()), None)
        if matched_src:
            out_df[target_col] = merged_mkcl[matched_src].apply(sanitize_key) if target_col == "SEAT_NO" else merged_mkcl[matched_src].astype(str).str.strip()

    prn_col = next((c for c in merged_mkcl.columns if "prn" in c.lower()), None)
    if prn_col:
        out_df["PRNERN"] = merged_mkcl[prn_col].apply(sanitize_key)

    out_df["APAAR ID"] = merged_mkcl["APAAR ID"].astype(str).str.strip()
    out_df["DEGNM"] = merged_mkcl["DEGNM"].astype(str).str.strip()
    out_df["PROGTYPE"] = "Degree"

    if _CGPA_SOURCE_FILE_INTERNAL_COL in merged_mkcl.columns:
        out_df["CGPA Source File"] = merged_mkcl[_CGPA_SOURCE_FILE_INTERNAL_COL].astype(str).str.strip()

    if college_choice == "Mithibai College":
            out_df["COLL_NO"] = "132"
            out_df["COLL_NAME"] = "Shri Vile Parle Kelvani Mandal's Mithibai College of Arts, Chauhan Institute of Science and Amrutben Jivanlal College of Commerce and Economics (Autonomous)"
            out_df["COLL_NAMEM"] = "श्री विलेपार्ले केळवणी संचालित, मिठीबाई कॉलेज ऑफ आर्टस्, चौहान इन्स्टिटयूट ऑफ सायन्स अँन्ड अमृतबेन जीवनलाल कॉलेज ऑफ कॉमर्स अँन्ड इकॉनॉमिक्स (स्वायत्त)"

    elif college_choice == "Narsee Monjee College":
        out_df["COLL_NO"] = "205"
        out_df["COLL_NAME"] = "Narsee Monjee College of Commerce & Economics (Empowered Autonomous)"
        out_df["COLL_NAMEM"] = "नरसी मोनजी कॉलेज ऑफ कॉमर्स आणि इकोनॉमिक्स (स्वायत्त)"

    elif college_choice == "Usha Pravin Gandhi College":
        out_df["COLL_NO"] = "598"
        out_df["COLL_NAME"] = "Usha Pravin Gandhi College of Arts, Science & Commerce (Autonomous)"
        out_df["COLL_NAMEM"] = "उषा प्रविण गांधी कॉलेज ऑफ आर्टस्, सायन्स अँन्ड कॉमर्स (स्वायत्त)"

    out_df["SEX"] = out_df["SEX"].apply(map_gender)
    out_df["CLASS"] = out_df["CLASS"].apply(format_class_grade)
    out_df["MCLASS"] = out_df["CGPA"].apply(convert_digits_to_marathi)
    out_df["FACULTY"] = out_df["DEGNM"].apply(derive_faculty)
    out_df["SUBDEGNM"] = out_df["DEGNM"].apply(derive_subdegnm)

    out_df["MONTH"] = out_df["MONTH"].astype(str).str.replace(",", " ").str.strip()
    out_df["MONTH"] = out_df["MONTH"].replace("nan", "")

    # Transliterate MONTH to Marathi and convert digits to Marathi numerals
    out_df["MMONTH"] = out_df["MONTH"].apply(transliterate_name_to_marathi).apply(convert_digits_to_marathi)
    out_df["NAME_MARAT"] = out_df["NAME"].apply(transliterate_name_to_marathi)

    out_df = out_df.fillna("").astype(str).replace("nan", "")
    out_df = out_df[TARGET_COLUMNS]

    out_df = out_df.drop_duplicates().reset_index(drop=True)
    out_df = out_df.sort_values(by="SEAT_NO", ascending=True).reset_index(drop=True)

    unmatched_df = pd.DataFrame(unmatched_rows)
    if not unmatched_df.empty:
        unmatched_df = unmatched_df.drop_duplicates().reset_index(drop=True)

    return out_df, unmatched_df, total_mkcl_raw_rows, total_filtered_rows, matched_count, unmatched_count


# ==========================================
# 5. OPENPYXL EXCEL GENERATION & STYLING
# ==========================================
def generate_formatted_excel(df: pd.DataFrame, progress_callback=None) -> bytes:
    wb = openpyxl.Workbook()
    
    # SHEET 1: Convocation Data
    ws = wb.active
    ws.title = "Convocation Data"
    ws.views.sheetView[0].showGridLines = False

    font_regular = Font(name="Times New Roman", size=11, bold=False)
    font_header = Font(name="Times New Roman", size=11, bold=True)
    font_cgpa_source_file = Font(name="Times New Roman", size=11, italic=True, color="808080")
    
    # Header fills
    header_fill_default = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill_orange = PatternFill(start_color="f1c232", end_color="f1c232", fill_type="solid")  # Light Orange
    
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border_side = Side(style="thin", color="D3D3D3")
    cell_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side, bottom=thin_border_side
    )

    orange_header_columns = {"APAAR ID", "CGPA", "CGPA Source File"}

    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill_orange if cell.value in orange_header_columns else header_fill_default
        cell.alignment = header_align
        cell.border = cell_border

    for row_values in df.itertuples(index=False):
        ws.append(list(row_values))

    cgpa_source_file_col_idx = (
        list(df.columns).index("CGPA Source File") + 1
        if "CGPA Source File" in df.columns else None
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = (
                font_cgpa_source_file
                if cgpa_source_file_col_idx is not None and cell.column == cgpa_source_file_col_idx
                else font_regular
            )
            cell.border = cell_border
            if cell.column_letter in ["A", "B", "E", "F", "H", "I", "R", "AF"]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    if progress_callback:
        progress_callback(0.5)

    # SHEET 2: Statistics
    ws_stats = wb.create_sheet(title="Statistics")
    ws_stats.views.sheetView[0].showGridLines = False

    stats_headers = ["PROGRAM NAME", "MALE (1)", "FEMALE (2)", "TOTAL COUNT"]
    ws_stats.append(stats_headers)

    for col_idx in range(1, len(stats_headers) + 1):
        cell = ws_stats.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill_default
        cell.alignment = header_align
        cell.border = cell_border

    if not df.empty and "DEGNM" in df.columns and "SEX" in df.columns:
        program_groups = df.groupby("DEGNM")
        
        grand_male = 0
        grand_female = 0
        grand_total = 0

        for prog_name, group in program_groups:
            display_prog = prog_name if prog_name.strip() else "[UNMATCHED PROGRAM]"
            male_cnt = (group["SEX"] == "1").sum()
            female_cnt = (group["SEX"] == "2").sum()
            total_cnt = len(group)

            grand_male += male_cnt
            grand_female += female_cnt
            grand_total += total_cnt

            row_data = [display_prog, male_cnt, female_cnt, total_cnt]
            ws_stats.append(row_data)

        total_row_idx = ws_stats.max_row + 1
        ws_stats.append(["TOTAL", grand_male, grand_female, grand_total])

        for row in ws_stats.iter_rows(min_row=2, max_row=ws_stats.max_row, min_col=1, max_col=4):
            is_total_row = (row[0].row == total_row_idx)
            for idx, cell in enumerate(row):
                cell.font = Font(name="Times New Roman", size=11, bold=is_total_row)
                cell.border = cell_border
                if idx == 0:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

    for r in range(1, ws_stats.max_row + 1):
        ws_stats.row_dimensions[r].height = 20

    for col in ws_stats.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_stats.column_dimensions[col_letter].width = max(max_len + 4, 15)

    if progress_callback:
        progress_callback(0.9)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    if progress_callback:
        progress_callback(1.0)

    return output_buffer.getvalue()


# ==========================================
# 6. MODULE ENTRY POINT FOR STREAMLIT ROUTING
# ==========================================
def show():
    st.title("📄 Convocation Data Generator")
    st.markdown(
        "Upload **MKCL Reports** and **Student Master Data** files. The app automatically handles "
        "Excel text/number mismatches to fetch degree/program names accurately, standardizes columns, "
        "and generates formatted convocation reports."
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        st.text("🏫 Select College")
        college_option = st.selectbox(
            "Select College",
            options=[
                "Mithibai College",
                "Narsee Monjee College",
                "Usha Pravin Gandhi College"                
            ],
            index=None,
            placeholder="Select College",
            label_visibility="collapsed"
        )

    with col2:
        mkcl_files = st.file_uploader(
            "Upload MKCL Report Files (.xlsx / .xls)",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="mkcl_uploader"
        )

    with col3:
        master_files = st.file_uploader(
            "Upload Student Master Data Files (.xlsx / .xls)",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="master_uploader"
        )

    st.markdown("---")

    process_ready = bool(college_option and mkcl_files and master_files)

    btn_col, progress_col = st.columns([1, 2])
    with btn_col:
        process_btn = st.button("🚀 Process & Generate Convocation File", type="primary", disabled=not process_ready)
    with progress_col:
        progress_placeholder = st.empty()

    if not process_ready:
        st.info("💡 Select a college and upload at least one MKCL Report file AND one Student Master Data file to enable processing.")

    if process_btn and process_ready:
        try:
            STAGE_RANGES = {
                "master": (0, 30, "📚 Building Student Master Lookup Map"),
                "read": (30, 60, "📥 Reading & Filtering MKCL Reports"),
                "match": (60, 88, "🔗 Matching Students & Deriving Data"),
                "excel": (88, 100, "📊 Generating Formatted Excel File"),
            }
            progress_bar = progress_placeholder.progress(0, text="Starting…  0%")

            def update_progress(stage, fraction):
                start, end, label = STAGE_RANGES[stage]
                fraction = max(0.0, min(1.0, fraction))
                pct = int(start + (end - start) * fraction)
                pct = max(0, min(100, pct))
                progress_bar.progress(pct / 100, text=f"{label}…  {pct}%")

            update_progress("master", 0.0)
            student_lookup, master_total_records = build_master_lookup(
                master_files,
                progress_callback=lambda frac: update_progress("master", frac)
            )

            final_df, unmatched_df, mkcl_raw_count, filtered_count, matched_count, unmatched_count = process_data(
                mkcl_files, student_lookup, college_option,
                progress_callback=update_progress
            )

            st.subheader("📊 Processing Summary Statistics")
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Master Records Loaded", f"{master_total_records:,}")
            m2.metric("MKCL Raw Records", f"{mkcl_raw_count:,}")
            m3.metric("Filtered Records (Conv/CGPA/GPAs)", f"{filtered_count:,}")
            m4.metric("Program Matches Found", f"{matched_count:,}")
            m5.metric("Unmatched Records", f"{unmatched_count:,}")

            if unmatched_count > 0:
                st.warning(f"⚠️ **Exception Report:** {unmatched_count} record(s) could not find a matching Program Name in Master Data.")
                with st.expander("🔍 View Unmatched Exceptions"):
                    st.dataframe(unmatched_df, use_container_width=True)

            if final_df.empty:
                progress_bar.progress(1.0, text="⚠️ Stopped — no records to export  100%")
                st.error("❌ No valid output records generated. Please verify your filter criteria or uploaded files.")
            else:
                st.subheader("Final Processed Data Preview")
                if "CGPA Source File" in final_df.columns:
                    styler = final_df.style
                    style_fn = getattr(styler, "map", None) or styler.applymap
                    preview_df = style_fn(
                        lambda _: "font-style: italic; color: grey;",
                        subset=["CGPA Source File"]
                    )
                else:
                    preview_df = final_df
                st.dataframe(preview_df, use_container_width=True)

                excel_data = generate_formatted_excel(
                    final_df,
                    progress_callback=lambda frac: update_progress("excel", frac)
                )

                progress_bar.progress(1.0, text="✅ Complete!  100%")

                college_tag = "NMC" if "Narsee Monjee" in college_option else "UPG"
                current_date_str = datetime.now().strftime("%d-%m-%Y")
                dynamic_filename = f"Convocation_Data_{college_tag}_{current_date_str}.xlsx"

                st.download_button(
                    label="📥 Download Standardized Convocation Excel File",
                    data=excel_data,
                    file_name=dynamic_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            progress_placeholder.progress(0, text="❌ Failed")
            st.error(f"❌ An error occurred during processing: {str(e)}")
            st.exception(e)


main = show

if __name__ == "__main__":
    show()