"""
==================================================
QUESTIONWISE CHECKER
==================================================

PURPOSE
--------------------------------------------------
This Streamlit tool helps exam staff:

1. Open password-protected or unprotected Excel mark sheets exported from
   the examination system.
2. Clean the raw export (remove blank/footer rows, drop unused
   columns).
3. Sort records so each student's rows appear in a fixed order:
   Examiner -> Moderator -> Reval 1 -> Reval 2.
4. Compare each Moderator / Revaluer's question-wise marks against
   the original Examiner's marks and highlight increases, decreases,
   and missing marks in the preview.
5. Summarise how many students were moderated / revalued, and how
   many scores went up, down, or stayed the same.
6. Export the processed output Excel file formatted with continuous SrNo,
   borders, bold scores, renamed sheet, removed Reval 1/2 rows, and dynamic output filename.
==================================================
"""

from datetime import datetime
import io
import os
import re
import tempfile

import msoffcrypto
from openpyxl.styles import Border, Font, Side
import pandas as pd
import streamlit as st


# ==================================================
# CONFIGURATION
# ==================================================

# Passwords tried in order to unlock the uploaded Excel file.
PASSWORDS = [
    "Exam@108",
    "UPG@123",
    "Exam@105"
]

# Non-question metadata columns
FIXED_COLUMNS = [
    "SrNo",
    "ExamAssingment_ID",
    "Subject Name-Subject code",
    "OnScreenID",
    "SubjectName",
    "Semester",
    "RollNo",
    "PRNNumber",
    "UserType",
    "MarkAttendance",
    "Email",
]

# Columns to completely remove from the final Excel download file
COLUMNS_TO_DROP = [
    "ExamAssingment_ID",
    "OnScreenID",
    "RollNo",
    "Name",
    "UserType",
    "CampusName",
    "MarkAttendance",
    "Email",
]

# Columns hidden from the on-screen Streamlit preview tables
PREVIEW_HIDDEN_COLUMNS = [
    "Subject Name-Subject code",
    "SubjectName",
    "Semester",
]

# Order of Examiner/Moderator/Reval rows per student
USER_TYPE_ORDER = {
    "Examiner": 1,
    "Moderator": 2,
    "Reval 1": 3,
    "Reval 2": 4,
}

# Row background colours per UserType for on-screen web display
ROW_COLOURS = {
    "Moderator": "background-color:#032E15;",
    "Reval 1": "background-color:#162456;color:white;",
    "Reval 2": "background-color:#290245;color:white;",
}

INCREASE_STYLE = "color:#9ACD32;font-weight:bold;"  # mark went up
DECREASE_STYLE = "color:#FF4500;font-weight:bold;"  # mark went down / missing


# ==================================================
# STEP 1: DECRYPT + LOAD THE UPLOADED FILE
# ==================================================
def decrypt_and_load(uploaded_file):
    """
    Attempts to load unprotected Excel files directly first.
    If protected, iterates through PASSWORDS list to decrypt.

    Returns:
        (dataframe, password_used) on success
        (None, None) if loading fails
    """
    # 1. Try reading directly (for unprotected Excel files)
    try:
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, dtype=object, keep_default_na=False)
        return df, "Unprotected (Direct)"
    except Exception:
        pass

    # 2. Try decrypting using known passwords
    for password in PASSWORDS:
        decrypted_path = None
        try:
            uploaded_file.seek(0)

            office_file = msoffcrypto.OfficeFile(uploaded_file)
            office_file.load_key(password=password)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                office_file.decrypt(tmp)
                decrypted_path = tmp.name

            df = pd.read_excel(decrypted_path, dtype=object, keep_default_na=False)
            return df, password

        except Exception:
            continue

        finally:
            if decrypted_path and os.path.exists(decrypted_path):
                os.remove(decrypted_path)

    return None, None


# ==================================================
# STEP 2: CLEAN + SORT THE DATA (FOR WEB APP)
# ==================================================
def clean_and_sort(df):
    """
    Clean up the raw export and sort rows for full web app analysis.
    """
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    if "PRNNumber" in df.columns:
        df = df[df["PRNNumber"].astype(str).str.strip() != ""]

    df = df.dropna(how="all")

    # Remove rows having "Count:" in SrNo column
    if "SrNo" in df.columns:
        df = df[~df["SrNo"].astype(str).str.contains("Count:", case=False, na=False)]
    else:
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains("Count:", case=False, na=False)]

    if "PRNNumber" in df.columns:
        df["PRNNumber"] = pd.to_numeric(df["PRNNumber"], errors="coerce")
        df = df[df["PRNNumber"].notna()]

    if "PRNNumber" in df.columns and "UserType" in df.columns:
        df["SortOrder"] = df["UserType"].map(USER_TYPE_ORDER).fillna(99)
        df = df.sort_values(by=["PRNNumber", "SortOrder"]).drop(columns=["SortOrder"])

    df = df.reset_index(drop=True)
    return df


# ==================================================
# STEP 2B: PROCESS DATA FOR OUTPUT EXCEL DOWNLOAD
# ==================================================
def process_data_for_download(df):
    """
    Applies download-specific transformations:
    - Removes 'Reval 1' and 'Reval 2' rows completely
    - Filters Moderator rows (drops Examiner rows for PRNs with Moderator)
    - Replaces "AB" with "UFM" or "ABSENT" based on MarkAttendance
    - Drops unwanted columns specified in COLUMNS_TO_DROP
    - Makes SrNo numbers continuous starting from 1
    """
    download_df = df.copy()

    # --- Rule: Remove Reval 1 and Reval 2 Rows ---
    if "UserType" in download_df.columns:
        reval_mask = download_df["UserType"].astype(str).str.strip().str.lower().isin(["reval 1", "reval 2"])
        download_df = download_df[~reval_mask]

    # --- Rule: Filter Moderator Rows ---
    if "PRNNumber" in download_df.columns and "UserType" in download_df.columns:
        moderator_prns = set(
            download_df[download_df["UserType"].astype(str).str.strip().str.lower() == "moderator"]["PRNNumber"]
        )

        def filter_rows(row):
            prn = row["PRNNumber"]
            u_type = str(row["UserType"]).strip().lower()
            if prn in moderator_prns:
                return u_type == "moderator"
            return True

        download_df = download_df[download_df.apply(filter_rows, axis=1)]

    # --- Rule: Replace "AB" in TotalObtainedScore ---
    if "TotalObtainedScore" in download_df.columns:
        def adjust_score(row):
            score = str(row["TotalObtainedScore"]).strip()
            attendance = str(row.get("MarkAttendance", "")).strip().upper()
            
            if score.upper() == "AB":
                if attendance == "UFM":
                    return "UFM"
                return "ABSENT"
            return score

        download_df["TotalObtainedScore"] = download_df.apply(adjust_score, axis=1)

    # --- Rule: Drop Unwanted Columns ---
    download_df = download_df.drop(columns=COLUMNS_TO_DROP, errors="ignore")

    # Reset index cleanly
    download_df = download_df.reset_index(drop=True)

    # --- Rule: Make SrNo Continuous (1, 2, 3, ...) ---
    if "SrNo" in download_df.columns:
        download_df["SrNo"] = download_df.index + 1

    return download_df


# ==================================================
# STEP 3: DISPLAY FORMATTING HELPERS
# ==================================================
def format_value(value):
    """Formats values cleanly for preview display."""
    text = str(value).strip()

    if text == "" or text.upper() == "NA":
        return "NA"

    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
        return f"{number:g}"
    except (TypeError, ValueError):
        return text


def highlight_rows(data, question_columns):
    """Applies CSS styles for Streamlit preview table."""
    styles = pd.DataFrame("", index=data.index, columns=data.columns)

    for idx in data.index:
        if "UserType" in data.columns:
            user_type = str(data.loc[idx, "UserType"]).strip()
            if user_type in ROW_COLOURS:
                styles.loc[idx, :] += ROW_COLOURS[user_type]

    if "PRNNumber" in data.columns and "UserType" in data.columns:
        for _, group in data.groupby("PRNNumber", dropna=False):
            examiner_rows = group[group["UserType"] == "Examiner"]
            if examiner_rows.empty:
                continue
            examiner_row = examiner_rows.iloc[0]

            for idx in group.index:
                if str(data.loc[idx, "UserType"]) == "Examiner":
                    continue

                for col in question_columns:
                    examiner_value = str(examiner_row[col]).strip()
                    current_value = str(data.loc[idx, col]).strip()

                    if examiner_value == "" or examiner_value.upper() == "NA":
                        continue

                    if current_value == "" or current_value.upper() == "NA":
                        styles.loc[idx, col] += DECREASE_STYLE
                        continue

                    try:
                        ex_mark = float(examiner_value)
                        cur_mark = float(current_value)
                    except (TypeError, ValueError):
                        continue

                    if cur_mark > ex_mark:
                        styles.loc[idx, col] += INCREASE_STYLE
                    elif cur_mark < ex_mark:
                        styles.loc[idx, col] += DECREASE_STYLE

    if "TotalObtainedScore" in data.columns:
        styles.loc[:, "TotalObtainedScore"] += "font-weight:bold;"

    return styles


def style_preview(df, question_columns, hide_columns=PREVIEW_HIDDEN_COLUMNS):
    """Builds styled interactive preview table for web display."""
    display_df = df.drop(columns=hide_columns, errors="ignore")
    return (
        display_df.style
        .hide(axis="index")
        .format(format_value)
        .apply(lambda data: highlight_rows(data, question_columns), axis=None)
    )


# ==================================================
# STEP 4: SUMMARY STATISTICS & METADATA
# ==================================================
def count_reviewers(df, user_type):
    """Number of distinct students (PRNNumber) that have a row of the given UserType."""
    if "UserType" not in df.columns or "PRNNumber" not in df.columns:
        return 0
    return (
        df[df["UserType"].astype(str).str.strip().eq(user_type)]["PRNNumber"]
        .nunique()
    )


def score_change_summary(df, review_type):
    """
    Compare each student's review_type TotalObtainedScore against Examiner's TotalObtainedScore.
    """
    increased = decreased = unchanged = 0

    if "TotalObtainedScore" not in df.columns or "UserType" not in df.columns:
        return increased, decreased, unchanged

    for _, group in df.groupby("PRNNumber", dropna=False):
        examiner_rows = group[group["UserType"] == "Examiner"]
        review_rows = group[group["UserType"] == review_type]

        if examiner_rows.empty or review_rows.empty:
            continue

        try:
            examiner_score = float(examiner_rows.iloc[0]["TotalObtainedScore"])
            review_score = float(review_rows.iloc[0]["TotalObtainedScore"])
        except (TypeError, ValueError):
            continue

        if review_score > examiner_score:
            increased += 1
        elif review_score < examiner_score:
            decreased += 1
        else:
            unchanged += 1

    return increased, decreased, unchanged


def build_paper_info(df):
    """Builds header info line for display."""
    if df.empty:
        return ""

    semester = str(df.iloc[0].get("Semester", "")).strip()
    subject_code = str(df.iloc[0].get("Subject Name-Subject code", "")).strip()
    subject_name = str(df.iloc[0].get("SubjectName", "")).strip()

    return f"{semester}  -  {subject_code}  -  {subject_name}"


def generate_output_filename(df):
    """
    Generates dynamic excel filename:
    <SubjectName>_<Semester>_<Unique PRN Count>_<DD-MM-YY>.xlsx
    Example: Laws for Cyber Security [Re-Exam]_Sem 6_3_24-07-26.xlsx
    """
    if df.empty:
        return f"Questionwise_Output_{datetime.now().strftime('%d-%m-%y')}.xlsx"

    subject_name = str(df.iloc[0].get("SubjectName", "Subject")).strip()
    semester_raw = str(df.iloc[0].get("Semester", "Semester")).strip()

    # Shorten 'Semester 6' -> 'Sem 6'
    semester_short = re.sub(r"(?i)\bsemester\b", "Sem", semester_raw).strip()

    unique_prns = df["PRNNumber"].nunique() if "PRNNumber" in df.columns else len(df)
    date_str = datetime.now().strftime("%d-%m-%y")

    filename = f"{subject_name}_{semester_short}_{unique_prns}_{date_str}.xlsx"
    return filename


# ==================================================
# STEP 5: EXCEL GENERATION & STYLING
# ==================================================
def sanitize_sheet_name(subject_name):
    """Sanitizes SubjectName into a valid Excel sheet name (max 31 chars, no forbidden symbols)."""
    if not subject_name:
        return "Questionwise Data"

    name = re.sub(r"\[[^\]]*\]", "", subject_name)
    name = re.sub(r"[\\/?*\[\]:]", "", name)
    name = re.sub(r"\s+", " ", name).strip()

    if not name:
        name = "Questionwise Data"

    return name[:31]


def to_excel_bytes(df):
    """
    Converts processed DataFrame to styled Excel file bytes:
    - Applies download-specific column drops, Moderator filters, and removes Reval rows
    - Ensures SrNo is sequential (1, 2, 3...)
    - Renames worksheet to SubjectName
    - Bold TotalObtainedScore column values
    - Applies thin borders across all data and header cells
    """
    download_df = process_data_for_download(df)
    buffer = io.BytesIO()

    sheet_name = "Questionwise Data"
    if not df.empty and "SubjectName" in df.columns:
        subj = str(df.iloc[0]["SubjectName"]).strip()
        if subj:
            sheet_name = sanitize_sheet_name(subj)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        download_df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        total_score_col_idx = None
        for col_idx, col_name in enumerate(download_df.columns, start=1):
            if col_name == "TotalObtainedScore":
                total_score_col_idx = col_idx
                break

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                
                if cell.row == 1:
                    cell.font = Font(bold=True)
                elif total_score_col_idx and cell.column == total_score_col_idx:
                    cell.font = Font(bold=True)

    return buffer.getvalue()


# ==================================================
# MAIN ENTRY POINT
# ==================================================
def show():
    """
    Render the Questionwise Checker page in Streamlit.
    """
    st.title("📝 Questionwise Checker")
    st.caption("Upload a password protected or unprotected Excel file.")

    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

    if uploaded_file is None:
        return

    try:
        file_id = f"{uploaded_file.name}_{uploaded_file.size}"
        if st.session_state.get("qc_last_file_id") != file_id:
            st.session_state.qc_last_file_id = file_id
            st.session_state.files_uploaded = st.session_state.get("files_uploaded", 0) + 1

        # Step 1: Decrypt + load
        with st.spinner("Decrypting and loading file..."):
            df, used_password = decrypt_and_load(uploaded_file)

        if df is None:
            st.error("Unable to open file. Password not recognised.")
            return

        st.caption(f"File opened successfully using password: {used_password}")

        # Step 2: Clean + sort
        df = clean_and_sort(df)

        if df.empty:
            st.warning("No valid student rows were found in this file after cleaning.")
            return

        if "PRNNumber" not in df.columns or "UserType" not in df.columns:
            st.error(
                "This file is missing the expected 'PRNNumber' or 'UserType' "
                "columns, so it can't be processed by this tool."
            )
            return

        question_columns = [col for col in df.columns if col not in FIXED_COLUMNS]

        # Step 3: Summary statistics
        total_students = df["PRNNumber"].nunique()
        moderation_count = count_reviewers(df, "Moderator")
        reval1_count = count_reviewers(df, "Reval 1")
        reval2_count = count_reviewers(df, "Reval 2")

        moderator_positive, moderator_negative, moderator_no_change = (
            score_change_summary(df, "Moderator")
        )
        reval1_positive, reval1_negative, reval1_no_change = (
            score_change_summary(df, "Reval 1")
        )
        reval2_positive, reval2_negative, reval2_no_change = (
            score_change_summary(df, "Reval 2")
        )

        st.session_state.students_processed = int(total_students)

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.info(
                f"Total Students: {total_students}\n\n"
                f"Moderation Count: {moderation_count}\n\n"
                f"Reval 1 Count: {reval1_count}\n\n"
                f"Reval 2 Count: {reval2_count}"
            )

        with col2:
            st.success(
                f"Positive Change in Marks:\n\n"
                f"Moderation: {moderator_positive}\n\n"
                f"Reval 1: {reval1_positive}\n\n"
                f"Reval 2: {reval2_positive}"
            )

        with col3:
            st.error(
                f"Negative Change in Marks:\n\n"
                f"Moderation: {moderator_negative}\n\n"
                f"Reval 1: {reval1_negative}\n\n"
                f"Reval 2: {reval2_negative}"
            )

        with col4:
            st.warning(
                f"No Change in Marks:\n\n"
                f"Moderation: {moderator_no_change}\n\n"
                f"Reval 1: {reval1_no_change}\n\n"
                f"Reval 2: {reval2_no_change}"
            )

        # Step 4: All-data preview
        st.header(build_paper_info(df))
        st.subheader("All Data Preview")
        st.write(style_preview(df, question_columns))

        # Step 5: Moderation / Revaluation only preview
        review_prns = df[
            df["UserType"].astype(str).str.strip().isin(["Moderator", "Reval 1", "Reval 2"])
        ]["PRNNumber"].unique()

        review_df = df[df["PRNNumber"].isin(review_prns)].copy()

        if not review_df.empty:
            st.subheader("Moderation / Revaluation Cases Only")
            st.write(style_preview(review_df, question_columns))
            st.write(f"Students Requiring Review: {review_df['PRNNumber'].nunique()}")
            st.write(f"Total Records: {len(review_df)}")

        # Step 6: Downloads
        st.markdown("---")
        download_col1, download_col2 = st.columns(2)

        out_filename = generate_output_filename(df)

        with download_col1:
            st.download_button(
                label=f"📥 Download Cleaned File (Excel)",
                data=to_excel_bytes(df),
                file_name=out_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as error:
        st.error(f"Something went wrong while processing this file: {error}")