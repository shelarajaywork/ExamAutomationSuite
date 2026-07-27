"""
==================================================
GRACING CHECKER MODULE
==================================================

PURPOSE
--------------------------------------------------
This Streamlit tool processes examination gracing sheets:

1. Loads uploaded Excel files containing pre/post gracing records.
2. Restructures student records into a consolidated tabular layout with exact column order:
   - Module Code, Module Desc, Credit, Internal Actual (Max), Internal Actual,
     Sem Actual (Max), Sem Actual, Sem Grace, Composite Score (Max), 
     Composite Score, Overall Grade, Completion Status after Gracing, Remarks & Scale.
   - Merged student header row across all columns:
     "Student Number | Student Name | Additional ID | Gender"
   - Converts marks to numeric values; replaces Z1 with "AB" and Z3 with "UFM" (0.0 calculation).
   - Highlights Z1 ("AB"), Z3 ("UFM"), marks < 40% of Max, and "F" grades in BOLD RED.
   - Summary row showing Total Composite Score, Total Composite Score (Max), 
     and Percentage in Overall Grade with thick bottom border for student separation.
   - Center alignment for columns Credit through Overall Grade.
   - Highlights 'Completed Unsuccessfuly' statuses in BOLD RED.
3. Identifies all instances where grace marks were awarded and students who failed.
4. Computes high-level statistics & summary reports.
5. Displays tab-based previews with 1-based indexing:
   - 📊 All Students (First Tab)
   - ⚠️ Failed Students
   - ✅ Graced Students
   - 📋 Graced Cases Details
   - 📚 Subject Summary
   - 👤 Student Summary
6. Generates a formatted multi-sheet Excel report ("All Students", "Failed Students", "Graced Students", etc.).
==================================================
"""

from datetime import datetime
import io
import os
import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st


# ==================================================
# HELPER FUNCTIONS
# ==================================================
def parse_scale_num(scale_val):
    """Parses maximum mark numeric float from scale strings (e.g., 'R050' -> 50.0)."""
    s = str(scale_val).strip()
    digits = re.findall(r'\d+', s)
    if digits:
        return float(digits[0])
    return 0.0


def parse_scale_str(scale_val):
    """Parses maximum mark integer display from scale strings (e.g., 'R050' -> '50')."""
    s = str(scale_val).strip()
    digits = re.findall(r'\d+', s)
    if digits:
        return str(int(digits[0]))
    return s if s.upper() not in ['NAN', 'NONE', 'NA'] else ''


def parse_mark(val):
    """
    Converts mark values to numeric or replaced text strings:
    - 'Z1' -> ('AB', 0.0, True)
    - 'Z3' -> ('UFM', 0.0, True)
    - Numeric values -> (int/float display, float calculation, False)
    """
    s = str(val).strip().upper()
    if s == 'Z1':
        return 'AB', 0.0, True
    elif s == 'Z3':
        return 'UFM', 0.0, True
    elif s in ['', 'NAN', 'NONE', 'NA']:
        return '', 0.0, False
    else:
        try:
            num = float(s)
            out_val = int(num) if num.is_integer() else num
            return out_val, float(num), False
        except (TypeError, ValueError):
            return val, 0.0, False


# ==================================================
# STEP 1: PROCESSED DATA GENERATION
# ==================================================
def create_processed_data(df, target_student_ids=None):
    """
    Restructures raw examination records into a clean student-wise dataset:
    - Merged Student Info Row across all columns: "Student Number | Name | Add ID | Gender"
    - Replaces Z1 with "AB" and Z3 with "UFM"
    - Checks 40% minimum passing criteria on max marks
    - Formats numbers, totals, and percentages
    - If target_student_ids is provided, filters strictly for those students.
    """
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    if target_student_ids is not None:
        target_set = set(str(sid).strip() for sid in target_student_ids)
        student_ids = [sid for sid in df['Student Number'].dropna().unique() if str(sid).strip() in target_set]
    else:
        student_ids = df['Student Number'].dropna().unique()

    all_processed_rows = []

    for s_num in student_ids:
        s_df = df[df['Student Number'] == s_num]

        # Non-summary module rows
        module_df = s_df[
            s_df['Module Desc'].notna() & 
            (~s_df['Student Name'].astype(str).str.upper().isin(['TOTAL', 'PERCENTAGE']))
        ]

        # Summary rows
        total_row = s_df[s_df['Student Name'].astype(str).str.upper() == 'TOTAL']
        perc_row = s_df[s_df['Student Name'].astype(str).str.upper() == 'PERCENTAGE']

        std_name = module_df.iloc[0]['Student Name'] if not module_df.empty else (s_df.iloc[0]['Student Name'] if not s_df.empty else '')
        add_id = module_df.iloc[0]['Additional ID'] if not module_df.empty else (s_df.iloc[0]['Additional ID'] if not s_df.empty else '')
        gender = module_df.iloc[0]['Gender'] if not module_df.empty else (s_df.iloc[0]['Gender'] if not s_df.empty else '')

        # 1. Clean Merged Student Header Row
        student_header_str = f"{s_num} | {std_name} | {add_id} | {gender}"
        all_processed_rows.append({
            'Module Code': student_header_str,
            'Module Desc': '',
            'Credit': '',
            'Internal Actual (Max)': '',
            'Internal Actual': '',
            'Sem Actual (Max)': '',
            'Sem Actual': '',
            'Sem Grace': '',
            'Composite Score (Max)': '',
            'Composite Score': '',
            'Overall Grade': '',
            'Completion Status after Gracing': '',
            'Remarks & Scale': '',
            'Row_Type': 'HEADER',
            'Comp_Mismatch': False,
            'Red_Internal': False,
            'Red_Sem': False,
            'Red_Grace': False,
            'Red_Composite': False
        })

        modules = module_df['Module Code'].dropna().unique()
        calc_total_comp = 0.0
        calc_total_max_comp = 0.0

        for m_code in modules:
            m_sub = module_df[module_df['Module Code'] == m_code]
            m_desc = m_sub.iloc[0]['Module Desc']

            # Credit parsing
            credit_row = m_sub[m_sub['Appraisal Type Code'].astype(str).str.strip() == 'Overall Grade']
            credit_raw = credit_row.iloc[0]['Credit'] if not credit_row.empty else m_sub.iloc[0]['Credit']
            credit_val, _, _ = parse_mark(credit_raw)

            # Helper for extracting appraisal type values and scale max marks
            def get_val_and_scale(code):
                sub = m_sub[m_sub['Appraisal Type Code'].astype(str).str.strip() == code]
                if not sub.empty:
                    val = sub['Marks after Gracing'].iloc[0]
                    scale = sub['Scale'].iloc[0] if 'Scale' in sub.columns else ''
                    return val, scale
                return '', ''

            int_raw, int_scale_raw = get_val_and_scale('Internal Actual')
            sem_raw, sem_scale_raw = get_val_and_scale('Sem Actual')
            grace_raw, _ = get_val_and_scale('Sem Grace')
            comp_raw, comp_scale_raw = get_val_and_scale('Composite Score')
            overall_grade, _ = get_val_and_scale('Overall Grade')

            int_display, i_num, int_is_abs_ufm = parse_mark(int_raw)
            sem_display, s_num_val, sem_is_abs_ufm = parse_mark(sem_raw)
            grace_display, g_num_val, grace_is_abs_ufm = parse_mark(grace_raw)
            comp_display, c_num_val, comp_is_abs_ufm = parse_mark(comp_raw)

            int_max_num = parse_scale_num(int_scale_raw)
            int_max_display = parse_scale_str(int_scale_raw)

            sem_max_num = parse_scale_num(sem_scale_raw)
            sem_max_display = parse_scale_str(sem_scale_raw)

            comp_max_num = parse_scale_num(comp_scale_raw)
            comp_max_display = parse_scale_str(comp_scale_raw)

            sub_overall = m_sub[m_sub['Appraisal Type Code'].astype(str).str.strip() == 'Overall Grade']
            completion_status = sub_overall['Completion Status after Gracing'].iloc[0] if not sub_overall.empty else ''

            # Merge Appraisal Remarks, General Remarks, Scale Type
            app_rem = str(sub_overall['Appraisal Remarks'].iloc[0]).strip() if not sub_overall.empty else ''
            gen_rem = str(sub_overall['General Remarks'].iloc[0]).strip() if not sub_overall.empty else ''
            scale_t = str(sub_overall['Scale Type'].iloc[0]).strip() if not sub_overall.empty else ''

            remarks_list = [r for r in [app_rem, gen_rem, scale_t] if r and r.upper() not in ['NAN', 'NONE', 'NA']]
            remarks_combined = " | ".join(remarks_list)

            # Verification: Internal Actual + Sem Actual + Sem Grace == Composite Score
            sum_components = i_num + s_num_val + g_num_val
            calc_total_comp += c_num_val
            calc_total_max_comp += comp_max_num

            is_mismatch = False
            if str(comp_raw).strip() not in ['', 'nan', 'NA']:
                if abs(sum_components - c_num_val) >= 0.01:
                    is_mismatch = True

            # 40% Minimum Passing Criteria Checks
            fail_int = (int_max_num > 0) and (i_num < 0.40 * int_max_num or int_is_abs_ufm)
            fail_sem = (sem_max_num > 0) and (s_num_val < 0.40 * sem_max_num or sem_is_abs_ufm)
            fail_comp = (comp_max_num > 0) and (c_num_val < 0.40 * comp_max_num or comp_is_abs_ufm)

            all_processed_rows.append({
                'Module Code': m_code,
                'Module Desc': m_desc,
                'Credit': credit_val,
                'Internal Actual (Max)': int_max_display,
                'Internal Actual': int_display,
                'Sem Actual (Max)': sem_max_display,
                'Sem Actual': sem_display,
                'Sem Grace': grace_display,
                'Composite Score (Max)': comp_max_display,
                'Composite Score': comp_display,
                'Overall Grade': overall_grade,
                'Completion Status after Gracing': completion_status,
                'Remarks & Scale': remarks_combined,
                'Row_Type': 'DATA',
                'Comp_Mismatch': is_mismatch,
                'Red_Internal': int_is_abs_ufm or fail_int,
                'Red_Sem': sem_is_abs_ufm or fail_sem,
                'Red_Grace': grace_is_abs_ufm,
                'Red_Composite': comp_is_abs_ufm or fail_comp or is_mismatch
            })

        # 2. Total & Percentage Summary Row
        tot_val_raw = total_row.iloc[0]['Marks after Gracing'] if not total_row.empty else (total_row.iloc[0]['Marks before Gracing'] if not total_row.empty else '')
        perc_val_raw = perc_row.iloc[0]['Marks after Gracing'] if not perc_row.empty else (perc_row.iloc[0]['Marks before Gracing'] if not perc_row.empty else '')

        tot_display, tot_num, _ = parse_mark(tot_val_raw)

        total_mismatch = False
        if str(tot_val_raw).strip() not in ['', 'nan', 'NA']:
            if abs(calc_total_comp - tot_num) >= 0.01:
                total_mismatch = True

        try:
            p_num = float(perc_val_raw)
            perc_display = f"{p_num:g}%"
        except (TypeError, ValueError):
            perc_display = f"{perc_val_raw}%" if str(perc_val_raw).strip() != '' else ''

        tot_max_display = int(calc_total_max_comp) if calc_total_max_comp.is_integer() else calc_total_max_comp

        all_processed_rows.append({
            'Module Code': '',
            'Module Desc': '',
            'Credit': '',
            'Internal Actual (Max)': '',
            'Internal Actual': '',
            'Sem Actual (Max)': '',
            'Sem Actual': '',
            'Sem Grace': '',
            'Composite Score (Max)': tot_max_display,
            'Composite Score': tot_display,
            'Overall Grade': perc_display,
            'Completion Status after Gracing': '',
            'Remarks & Scale': '',
            'Row_Type': 'SUMMARY',
            'Comp_Mismatch': total_mismatch,
            'Red_Internal': False,
            'Red_Sem': False,
            'Red_Grace': False,
            'Red_Composite': total_mismatch
        })

    return pd.DataFrame(all_processed_rows)


# ==================================================
# STEP 2: GRACED CASES & SUMMARIES
# ==================================================
def process_gracing_data(df):
    """Extracts cases strictly where grace marks (>0) were awarded."""
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    non_summary = df[
        df['Module Desc'].notna() & 
        (~df['Student Name'].astype(str).str.upper().isin(['TOTAL', 'PERCENTAGE']))
    ].copy()

    grace_rows = non_summary[
        non_summary['Appraisal Type Code'].astype(str).str.strip().isin(['Sem Grace', 'Grace on Total marks'])
    ].copy()

    grace_rows['Grace_Marks_Val'] = pd.to_numeric(grace_rows['Marks after Gracing'], errors='coerce').fillna(0)
    awarded_grace = grace_rows[grace_rows['Grace_Marks_Val'] > 0].copy()

    records = []
    for _, g_row in awarded_grace.iterrows():
        s_num = g_row['Student Number']
        m_code = g_row['Module Code']

        sub_df = non_summary[
            (non_summary['Student Number'] == s_num) & 
            (non_summary['Module Code'] == m_code)
        ]

        sem_actual = sub_df[sub_df['Appraisal Type Code'].astype(str).str.strip() == 'Sem Actual']
        sem_tot = sub_df[sub_df['Appraisal Type Code'].astype(str).str.strip() == 'Semester Total Marks']
        comp_score = sub_df[sub_df['Appraisal Type Code'].astype(str).str.strip() == 'Composite Score']
        grade_row = sub_df[sub_df['Appraisal Type Code'].astype(str).str.strip() == 'Overall Grade']

        marks_before = sem_actual.iloc[0]['Marks before Gracing'] if not sem_actual.empty else g_row.get('Marks before Gracing', '')
        marks_after = sem_tot.iloc[0]['Marks after Gracing'] if not sem_tot.empty else ''
        comp_before = comp_score.iloc[0]['Marks before Gracing'] if not comp_score.empty else ''
        comp_after = comp_score.iloc[0]['Marks after Gracing'] if not comp_score.empty else ''
        grade = grade_row.iloc[0]['Marks after Gracing'] if not grade_row.empty else ''
        status = grade_row.iloc[0]['Completion Status after Gracing'] if not grade_row.empty else ''

        records.append({
            'Student Number': s_num,
            'Student Name': g_row.get('Student Name', ''),
            'Additional ID': g_row.get('Additional ID', ''),
            'Module Code': m_code,
            'Module Desc': g_row.get('Module Desc', ''),
            'Appraisal Type': g_row.get('Appraisal Type Code', ''),
            'Marks Before Gracing': marks_before,
            'Grace Marks': g_row['Grace_Marks_Val'],
            'Marks After Gracing': marks_after,
            'Composite Score Before': comp_before,
            'Composite Score After': comp_after,
            'Grade': grade,
            'Status': status
        })

    return pd.DataFrame(records)


def get_subject_summary(graced_df):
    """Generates subject-wise summary for graced marks."""
    if graced_df.empty:
        return pd.DataFrame(columns=[
            'Subject Code', 'Subject Name', 'Graced Students Count', 
            'Total Grace Marks', 'Max Grace Marks', 'Avg Grace Marks'
        ])

    subj_summary = graced_df.groupby(['Module Code', 'Module Desc']).agg(
        Graced_Students=('Student Number', 'nunique'),
        Total_Grace=('Grace Marks', 'sum'),
        Max_Grace=('Grace Marks', 'max'),
        Avg_Grace=('Grace Marks', 'mean')
    ).reset_index()

    subj_summary['Avg_Grace'] = subj_summary['Avg_Grace'].round(2)

    return subj_summary.rename(columns={
        'Module Code': 'Subject Code',
        'Module Desc': 'Subject Name',
        'Graced_Students': 'Graced Students Count',
        'Total_Grace': 'Total Grace Marks',
        'Max_Grace': 'Max Grace Marks',
        'Avg_Grace': 'Avg Grace Marks'
    })


def get_student_summary(graced_df):
    """Generates student-wise summary for graced marks."""
    if graced_df.empty:
        return pd.DataFrame(columns=[
            'Student Number', 'Student Name', 'Additional ID', 
            'Graced Subjects Count', 'Total Grace Marks', 'Graced Subjects & Marks'
        ])

    student_summary = graced_df.groupby(['Student Number', 'Student Name', 'Additional ID']).agg(
        Graced_Subjects_Count=('Module Code', 'count'),
        Total_Grace_Marks=('Grace Marks', 'sum'),
        Subjects_List=('Module Desc', lambda x: ', '.join([
            f"{subj} (+{int(g) if float(g).is_integer() else g})" 
            for subj, g in zip(x, graced_df.loc[x.index, 'Grace Marks'])
        ]))
    ).reset_index()

    return student_summary.rename(columns={
        'Graced_Subjects_Count': 'Graced Subjects Count',
        'Total_Grace_Marks': 'Total Grace Marks',
        'Subjects_List': 'Graced Subjects & Marks'
    })


# ==================================================
# STEP 3: DISPLAY & STYLING HELPERS
# ==================================================
def format_value(value):
    """Formats cell values cleanly for Streamlit interactive preview."""
    text = str(value).strip()
    if text == "" or text.upper() in ["NA", "NAN", "NONE"]:
        return ""
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
        return f"{number:g}"
    except (TypeError, ValueError):
        return text


def style_processed_preview(proc_df):
    """
    Styles the Processed Data table:
    - Bold max columns
    - Center alignment for Credit through Overall Grade
    - BOLD Red font for marks < 40%, Z1 ("AB"), Z3 ("UFM"), or mismatched Composite Scores
    - BOLD Red font for 'F' grades
    - BOLD Red font for 'Completed Unsuccessfuly'
    - Thick bottom border on student summary rows
    """
    internal_flags = [
        'Row_Type', 'Comp_Mismatch', 'Fail_Internal', 'Fail_Sem', 
        'Fail_Composite', 'Red_Internal', 'Red_Sem', 'Red_Grace', 'Red_Composite'
    ]
    display_df = proc_df.drop(columns=internal_flags, errors='ignore').copy()
    display_df.index = range(1, len(display_df) + 1)

    def apply_custom_styles(data):
        styles = pd.DataFrame("", index=data.index, columns=data.columns)

        center_cols = [
            'Credit', 'Internal Actual (Max)', 'Internal Actual',
            'Sem Actual (Max)', 'Sem Actual', 'Sem Grace', 
            'Composite Score (Max)', 'Composite Score', 'Overall Grade'
        ]

        max_cols = ['Internal Actual (Max)', 'Sem Actual (Max)', 'Composite Score (Max)']

        for idx in data.index:
            orig_idx = idx - 1
            p_row = proc_df.iloc[orig_idx]
            row_type = p_row.get('Row_Type', '')

            status_val = str(data.loc[idx, 'Completion Status after Gracing']).strip() if 'Completion Status after Gracing' in data.columns else ''

            # Center align specified columns
            for col in center_cols:
                if col in data.columns:
                    styles.loc[idx, col] += "text-align: center;"

            # Bold max columns
            for col in max_cols:
                if col in data.columns:
                    styles.loc[idx, col] += "font-weight: bold;"

            # 1. Student Header Row (Soft Blue Background + Bold text)
            if row_type == 'HEADER':
                styles.loc[idx, :] += "background-color: #D9E1F2; font-weight: bold; color: #1F4E78;"
                continue

            # 2. Total Summary Row (Soft Yellow Background + Bold text + Thick Bottom Border)
            if row_type == 'SUMMARY':
                styles.loc[idx, :] += "background-color: #FFF2CC; font-weight: bold; color: #000000; border-bottom: 3px solid #000000;"
                if p_row.get('Red_Composite', False) and 'Composite Score' in data.columns:
                    styles.loc[idx, 'Composite Score'] += "color: red; font-weight: bold;"
                continue

            # 3. BOLD Red font for < 40%, Z1 ("AB"), Z3 ("UFM")
            if p_row.get('Red_Internal', False) and 'Internal Actual' in data.columns:
                styles.loc[idx, 'Internal Actual'] += "color: red; font-weight: bold;"
            if p_row.get('Red_Sem', False) and 'Sem Actual' in data.columns:
                styles.loc[idx, 'Sem Actual'] += "color: red; font-weight: bold;"
            if p_row.get('Red_Grace', False) and 'Sem Grace' in data.columns:
                styles.loc[idx, 'Sem Grace'] += "color: red; font-weight: bold;"
            if p_row.get('Red_Composite', False) and 'Composite Score' in data.columns:
                styles.loc[idx, 'Composite Score'] += "color: red; font-weight: bold;"

            # 4. BOLD Red font for Grade 'F'
            grade_val = str(data.loc[idx, 'Overall Grade']).strip().upper() if 'Overall Grade' in data.columns else ''
            if grade_val == 'F' and 'Overall Grade' in data.columns:
                styles.loc[idx, 'Overall Grade'] += "color: red; font-weight: bold;"

            # 5. BOLD Red font for 'Completed Unsuccessfuly'
            if status_val.lower() == 'completed unsuccessfuly' and 'Completion Status after Gracing' in data.columns:
                styles.loc[idx, 'Completion Status after Gracing'] += "color: red; font-weight: bold;"

        return styles

    return (
        display_df.style
        .format(format_value)
        .apply(apply_custom_styles, axis=None)
    )


def style_preview(df):
    """Prepares dataframe with 1-based indexing for preview display."""
    display_df = df.copy()
    display_df.index = range(1, len(display_df) + 1)
    return display_df.style.format(format_value)


# ==================================================
# STEP 4: EXCEL REPORT GENERATION
# ==================================================
def format_excel_processed_sheet(worksheet, clean_proc, proc_df):
    """Helper to apply formatting, borders, bold fonts, and red highlights to processed student sheets."""
    cols = clean_proc.columns.tolist()

    column_widths = {
        'Module Code': 15,
        'Module Desc': 40,
        'Credit': 7,
        'Internal Actual (Max)': 11,
        'Internal Actual': 11,
        'Sem Actual (Max)': 11,
        'Sem Actual': 11,
        'Sem Grace': 11,
        'Composite Score (Max)': 11,
        'Composite Score': 11,
        'Overall Grade': 11,
        'Completion Status after Gracing': 25,
        'Remarks & Scale': 25
    }

    center_cols_indices = [
        i + 1 for i, c in enumerate(cols) if c in [
            'Credit', 'Internal Actual (Max)', 'Internal Actual',
            'Sem Actual (Max)', 'Sem Actual', 'Sem Grace', 
            'Composite Score (Max)', 'Composite Score', 'Overall Grade'
        ]
    ]
    max_cols_indices = [
        i + 1 for i, c in enumerate(cols) if c in [
            'Internal Actual (Max)', 'Sem Actual (Max)', 'Composite Score (Max)'
        ]
    ]

    int_act_idx = cols.index('Internal Actual') + 1 if 'Internal Actual' in cols else None
    sem_act_idx = cols.index('Sem Actual') + 1 if 'Sem Actual' in cols else None
    sem_grace_idx = cols.index('Sem Grace') + 1 if 'Sem Grace' in cols else None
    comp_score_idx = cols.index('Composite Score') + 1 if 'Composite Score' in cols else None
    grade_idx = cols.index('Overall Grade') + 1 if 'Overall Grade' in cols else None
    status_col_idx = cols.index('Completion Status after Gracing') + 1 if 'Completion Status after Gracing' in cols else None

    thin_side = Side(style="thin", color="000000")
    thick_side = Side(style="medium", color="000000")

    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    summary_thick_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_side)

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    student_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    red_bold_font = Font(color="FF0000", bold=True)
    bold_font = Font(bold=True)
    bold_blue_font = Font(color="1F4E78", bold=True)

    worksheet.views.sheetView[0].showGridLines = False
    worksheet.freeze_panes = "A2"

    for row_idx in range(1, len(clean_proc) + 2):
        if row_idx == 1:
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = header_align
            continue

        proc_idx = row_idx - 2
        p_row = proc_df.iloc[proc_idx]
        row_type_val = p_row.get('Row_Type', '')

        is_student_header = (row_type_val == 'HEADER')
        is_summary_row = (row_type_val == 'SUMMARY')

        if is_student_header:
            worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(cols))
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = bold_blue_font
                cell.fill = student_header_fill
        else:
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)

                if is_summary_row:
                    cell.border = summary_thick_border
                    cell.font = bold_font
                    cell.fill = summary_fill
                else:
                    cell.border = thin_border
                    if col_idx in max_cols_indices:
                        cell.font = bold_font

                if col_idx in center_cols_indices:
                    cell.alignment = center_align

                # BOLD RED font for marks < 40%, Z1 ("AB"), Z3 ("UFM")
                if p_row.get('Red_Internal', False) and col_idx == int_act_idx:
                    cell.font = red_bold_font
                if p_row.get('Red_Sem', False) and col_idx == sem_act_idx:
                    cell.font = red_bold_font
                if p_row.get('Red_Grace', False) and col_idx == sem_grace_idx:
                    cell.font = red_bold_font
                if p_row.get('Red_Composite', False) and col_idx == comp_score_idx:
                    cell.font = red_bold_font

                # BOLD RED font for Overall Grade 'F'
                grade_val = str(cell.value).strip().upper() if col_idx == grade_idx else ''
                if col_idx == grade_idx and grade_val == 'F':
                    cell.font = red_bold_font

                # BOLD RED font for unsuccessful completion status
                status_val = str(worksheet.cell(row=row_idx, column=status_col_idx).value).strip() if status_col_idx else ''
                if status_col_idx and col_idx == status_col_idx and status_val.lower() == 'completed unsuccessfuly':
                    cell.font = red_bold_font

    # Apply specified fixed column widths
    for col_name, width in column_widths.items():
        if col_name in cols:
            c_idx = cols.index(col_name) + 1
            col_letter = get_column_letter(c_idx)
            worksheet.column_dimensions[col_letter].width = width


def to_excel_bytes(all_proc_df, failed_proc_df, graced_proc_df, graced_df, subject_summary, student_summary):
    """
    Generates styled multi-sheet Excel output file:
    - Sheets: "All Students", "Failed Students", "Graced Students", 
              "Graced Cases Details", "Subject Summary", "Student Summary"
    """
    buffer = io.BytesIO()
    internal_flags = [
        'Row_Type', 'Comp_Mismatch', 'Fail_Internal', 'Fail_Sem', 
        'Fail_Composite', 'Red_Internal', 'Red_Sem', 'Red_Grace', 'Red_Composite'
    ]

    clean_all = all_proc_df.drop(columns=internal_flags, errors='ignore')
    clean_failed = failed_proc_df.drop(columns=internal_flags, errors='ignore')
    clean_graced_students = graced_proc_df.drop(columns=internal_flags, errors='ignore')

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        clean_all.to_excel(writer, index=False, sheet_name="All Students")
        clean_failed.to_excel(writer, index=False, sheet_name="Failed Students")
        clean_graced_students.to_excel(writer, index=False, sheet_name="Graced Students")

        if not graced_df.empty:
            graced_df.to_excel(writer, index=False, sheet_name="Graced Cases Details")
            subject_summary.to_excel(writer, index=False, sheet_name="Subject Summary")
            student_summary.to_excel(writer, index=False, sheet_name="Student Summary")

        workbook = writer.book

        # Format All Students, Failed Students, Graced Students sheets
        format_excel_processed_sheet(workbook["All Students"], clean_all, all_proc_df)
        format_excel_processed_sheet(workbook["Failed Students"], clean_failed, failed_proc_df)
        format_excel_processed_sheet(workbook["Graced Students"], clean_graced_students, graced_proc_df)

        # Format remaining summary sheets
        thin_side = Side(style="thin", color="000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold_font = Font(bold=True)

        other_sheets = ["Graced Cases Details", "Subject Summary", "Student Summary"] if not graced_df.empty else []
        for sheet_name in other_sheets:
            worksheet = workbook[sheet_name]
            worksheet.views.sheetView[0].showGridLines = False
            worksheet.freeze_panes = "A2"

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:
                        cell.font = bold_font
                        cell.fill = header_fill
                        cell.alignment = header_align

            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

    return buffer.getvalue()


# ==================================================
# MAIN ENTRY POINT
# ==================================================
def show():
    """Renders the Gracing Checker tool page in Streamlit."""
    st.title("✅ Gracing Checker")
    st.caption("Upload an examination result Excel file to analyze awarded grace marks.")

    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx", "xls"])

    if uploaded_file is None:
        return

    try:
        # Load raw file
        with st.spinner("Loading and processing gracing data..."):
            df = pd.read_excel(uploaded_file, dtype=object, keep_default_na=False)

        if df.empty:
            st.warning("The uploaded file is empty.")
            return

        # Clean column headers
        df.columns = df.columns.astype(str).str.strip()

        # Check required columns
        required_cols = ['Student Number', 'Module Desc', 'Marks after Gracing', 'Appraisal Type Code']
        missing_cols = [c for c in required_cols if c not in df.columns]
        if missing_cols:
            st.error(f"The uploaded file is missing expected columns: {', '.join(missing_cols)}")
            return

        # Filter student lists
        failed_prns = df[
            df['Completion Status after Gracing'].astype(str).str.strip().str.lower() == 'completed unsuccessfuly'
        ]['Student Number'].dropna().unique().tolist()

        graced_prns = df[
            (df['Appraisal Type Code'].astype(str).str.strip().isin(['Sem Grace', 'Grace on Total marks'])) &
            (pd.to_numeric(df['Marks after Gracing'], errors='coerce').fillna(0) > 0)
        ]['Student Number'].dropna().unique().tolist()

        # Generate Processed Datasets
        all_proc_df = create_processed_data(df)
        failed_proc_df = create_processed_data(df, target_student_ids=failed_prns)
        graced_proc_df = create_processed_data(df, target_student_ids=graced_prns)

        # Generate Summaries
        graced_df = process_gracing_data(df)
        subject_summary = get_subject_summary(graced_df)
        student_summary = get_student_summary(graced_df)

        # Statistics
        total_students = df['Student Number'].nunique()
        graced_students_count = len(graced_prns)
        total_graced_cases = len(graced_df)
        total_grace_marks = graced_df['Grace Marks'].sum() if not graced_df.empty else 0

        # Update session state
        st.session_state.students_processed = int(total_students)

        st.markdown("---")

        # Metric Cards
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Total Candidates Evaluated", total_students)

        with col2:
            st.metric("Candidates Receiving Grace", graced_students_count)

        with col3:
            st.metric("Total Gracing Cases", total_graced_cases)

        with col4:
            st.metric("Total Grace Marks Awarded", f"{total_grace_marks:g}")

        st.markdown("---")

        # Tab-based Preview Windows
        tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
            "📊 All Students", 
            "⚠️ Failed Students", 
            "✅ Graced Students",
            "📋 Graced Cases Details", 
            "📚 Subject Summary", 
            "👤 Student Summary"
        ])

        with tab1:
            st.subheader("All Students Preview")
            if not all_proc_df.empty:
                st.write(style_processed_preview(all_proc_df))
            else:
                st.info("No student data generated.")

        with tab2:
            st.subheader("Failed Students Preview")
            if not failed_proc_df.empty:
                st.write(style_processed_preview(failed_proc_df))
            else:
                st.info("No failed students found in this file.")

        with tab3:
            st.subheader("Graced Students Preview")
            if not graced_proc_df.empty:
                st.write(style_processed_preview(graced_proc_df))
            else:
                st.info("No graced students found in this file.")

        with tab4:
            st.subheader("Graced Cases Details")
            if not graced_df.empty:
                st.write(style_preview(graced_df))
            else:
                st.info("No grace marks were awarded in this file.")

        with tab5:
            st.subheader("Subject-Wise Gracing Summary")
            if not subject_summary.empty:
                st.write(style_preview(subject_summary))
            else:
                st.info("No subject-wise gracing data available.")

        with tab6:
            st.subheader("Student-Wise Gracing Summary")
            if not student_summary.empty:
                st.write(style_preview(student_summary))
            else:
                st.info("No student-wise gracing data available.")

        # Export Section
        st.markdown("---")
        st.subheader("📥 Export Gracing Report")

        excel_bytes = to_excel_bytes(all_proc_df, failed_proc_df, graced_proc_df, graced_df, subject_summary, student_summary)
        date_str = datetime.now().strftime("%d-%m-%y")
        output_filename = f"Gracing_Analysis_Report_{date_str}.xlsx"

        st.download_button(
            label=f"📥 Download Complete Gracing Report ({output_filename})",
            data=excel_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as error:
        st.error(f"Something went wrong while processing the file: {error}")